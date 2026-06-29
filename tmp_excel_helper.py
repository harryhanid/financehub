import sys, csv
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH     = r'C:\Users\25010160\Downloads\UNIQLO_Week7_Analytics.xlsx'
PATH_OUT = r'C:\Users\25010160\Downloads\UNIQLO_Week7_StepByStep.xlsx'

print("Loading workbook...")
wb = load_workbook(PATH)
ws = wb['Data + Risk Score']

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, bg, fg="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def dat(cell, bg="FFFFFF", bold=False, center=False):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(size=9, bold=bold, color="222222")
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = border

# ── Column positions (1-based) ────────────────────────────────────────────
# Existing columns 1-29. We add helpers starting at column 30.
COL_START = 30  # column AD

headers_helper = [
    ("LANGKAH 1\nJual Rugi?\n(cost > price)", "FF4C4C"),   # AD = 30
    ("LANGKAH 2\nDiskon Besar?\n(≥20%)", "FF9900"),        # AE = 31
    ("LANGKAH 3\nToko Medan?\n", "9933CC"),                # AF = 32
    ("LANGKAH 4\nAccessories?\n", "0066CC"),               # AG = 33
    ("LANGKAH 5\nBulan Rawan?\n(2,5,8,10)", "006633"),     # AH = 34
    ("TOTAL\nRISK SCORE\n(0-100)", "1F3864"),              # AI = 35
]

# Write headers
ws.row_dimensions[1].height = 45
for i, (label, color) in enumerate(headers_helper):
    col = COL_START + i
    cell = ws.cell(row=1, column=col, value=label)
    hdr(cell, color)
    ws.column_dimensions[get_column_letter(col)].width = 14

print("Writing helper formulas for all rows...")
max_row = ws.max_row  # 49801

for r in range(2, max_row + 1):
    # LANGKAH 1: cost_price (N) > list_price (O) ?
    c1 = ws.cell(row=r, column=30)
    c1.value = f'=IF(N{r}>O{r},"YA (+40 poin)","Tidak")'
    bg1 = "FFD7D7" if ws.cell(row=r, column=14).value else "FFFFFF"

    # LANGKAH 2: discount >= 20% ?
    c2 = ws.cell(row=r, column=31)
    c2.value = f'=IF(OR(G{r}="20%",G{r}="30%"),"YA (+20 poin)",IF(G{r}="10%","YA (+10 poin)","Tidak"))'

    # LANGKAH 3: Medan Center ?
    c3 = ws.cell(row=r, column=32)
    c3.value = f'=IF(T{r}="Medan Center","YA (+15 poin)","Tidak")'

    # LANGKAH 4: Accessories ?
    c4 = ws.cell(row=r, column=33)
    c4.value = f'=IF(I{r}="Accessories","YA (+15 poin)","Tidak")'

    # LANGKAH 5: Peak return month ?
    c5 = ws.cell(row=r, column=34)
    c5.value = f'=IF(OR(MONTH(B{r})=2,MONTH(B{r})=5,MONTH(B{r})=8,MONTH(B{r})=10),"YA (+10 poin)","Tidak")'

    # TOTAL SCORE
    c6 = ws.cell(row=r, column=35)
    c6.value = (
        f'=MIN(100,'
        f'IF(N{r}>O{r},40,0)'
        f'+IF(OR(G{r}="20%",G{r}="30%"),20,IF(G{r}="10%",10,0))'
        f'+IF(T{r}="Medan Center",15,0)'
        f'+IF(I{r}="Accessories",15,0)'
        f'+IF(OR(MONTH(B{r})=2,MONTH(B{r})=5,MONTH(B{r})=8,MONTH(B{r})=10),10,0)'
        f')'
    )

    # basic styling (just border + font, skip heavy fill for speed)
    for cell in (c1, c2, c3, c4, c5):
        cell.font = Font(size=9, color="222222")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    c6.font = Font(size=9, bold=True, color="FFFFFF")
    c6.alignment = Alignment(horizontal="center", vertical="center")
    c6.border = border
    # color c6 based on computed score value in existing col 29
    try:
        sc = ws.cell(row=r, column=29).value
        if sc is not None:
            s = int(sc)
            if s >= 60:   c6.fill = PatternFill("solid", fgColor="C0392B")
            elif s >= 30: c6.fill = PatternFill("solid", fgColor="E67E22")
            else:         c6.fill = PatternFill("solid", fgColor="27AE60")
    except:
        c6.fill = PatternFill("solid", fgColor="1F3864")

    if r % 5000 == 0:
        print(f"  {r:,}/{max_row:,} rows...")

# ── Add a "CARA BACA" sheet ───────────────────────────────────────────────
if "Cara Baca Formula" in wb.sheetnames:
    del wb["Cara Baca Formula"]

ws_guide = wb.create_sheet("Cara Baca Formula", 1)
ws_guide.sheet_view.showGridLines = False
ws_guide.column_dimensions['A'].width = 8
ws_guide.column_dimensions['B'].width = 28
ws_guide.column_dimensions['C'].width = 35
ws_guide.column_dimensions['D'].width = 20
ws_guide.column_dimensions['E'].width = 18

steps = [
    # (row_height, col_a, col_b, col_c, col_d, col_e, bg, fg, bold)
    (30, "CARA MEMBACA FORMULA RISK SCORE", "", "", "", "", "1F3864", "FFFFFF", True),
    (16, "File: UNIQLO_Week7_Analytics.xlsx  |  Sheet: Data + Risk Score", "", "", "", "", "2E4A8B", "FFFFFF", False),
    (10, "", "", "", "", "", "FFFFFF", "000000", False),
    (20, "KONSEP DASAR", "", "", "", "", "E8EFF7", "1F3864", True),
    (16, "Setiap transaksi mendapat 'nilai risiko' seperti kuis:", "", "", "", "", "FFFFFF", "333333", False),
    (16, "Semakin besar nilainya, semakin berisiko transaksi itu merugikan perusahaan.", "", "", "", "", "FFFFFF", "333333", False),
    (10, "", "", "", "", "", "FFFFFF", "000000", False),
    (22, "LANGKAH", "PERTANYAAN", "FORMULA DI EXCEL", "KOLOM DI FILE", "NILAI JIKA 'YA'", "2E4A8B", "FFFFFF", True),
    (20, "Langkah 1", "Apakah produk dijual di bawah harga modal?", '=IF(N2>O2,"YA (+40 poin)","Tidak")', "Kolom AD", "+40 poin", "FFD7D7", "333333", False),
    (20, "Langkah 2", "Apakah diskon 20% atau 30%?", '=IF(OR(G2="20%",G2="30%"),"YA (+20 poin)",IF(G2="10%","YA (+10 poin)","Tidak"))', "Kolom AE", "+20 poin (atau +10)", "FFE8CC", "333333", False),
    (20, "Langkah 3", "Apakah dari toko Medan Center?", '=IF(T2="Medan Center","YA (+15 poin)","Tidak")', "Kolom AF", "+15 poin", "EDD7FF", "333333", False),
    (20, "Langkah 4", "Apakah kategori Accessories?", '=IF(I2="Accessories","YA (+15 poin)","Tidak")', "Kolom AG", "+15 poin", "D7E8FF", "333333", False),
    (20, "Langkah 5", "Apakah bulan Feb/Mei/Ags/Okt?", "=IF(OR(MONTH(B2)=2,MONTH(B2)=5,\nMONTH(B2)=8,MONTH(B2)=10),\n\"YA (+10 poin)\",\"Tidak\")", "Kolom AH", "+10 poin", "D7FFE8", "333333", False),
    (20, "TOTAL SCORE", "Jumlahkan semua poin, maksimum 100", "=MIN(100, total semua langkah)", "Kolom AI", "0 - 100", "1F3864", "FFFFFF", True),
    (10, "", "", "", "", "", "FFFFFF", "000000", False),
    (22, "ARTI SCORE", "", "", "", "", "E8EFF7", "1F3864", True),
    (18, "Score >= 60  =  RISIKO TINGGI", "97% transaksi ini jual di bawah modal", "", "", "", "FFCCCC", "333333", True),
    (18, "Score 30-59  =  RISIKO SEDANG", "49% transaksi ini jual di bawah modal", "", "", "", "FFF2CC", "333333", False),
    (18, "Score < 30   =  RISIKO RENDAH", "0% transaksi ini jual di bawah modal", "", "", "", "E2EFDA", "333333", False),
    (10, "", "", "", "", "", "FFFFFF", "000000", False),
    (22, "CONTOH PERHITUNGAN", "", "", "", "", "E8EFF7", "1F3864", True),
    (16, "Transaksi: Accessories, diskon 20%, dari Medan, bulan Agustus, cost > price", "", "", "", "", "FFFFFF", "444444", False),
    (16, "  Langkah 1: cost > price  → +40", "", "", "", "", "FFD7D7", "333333", False),
    (16, "  Langkah 2: diskon 20%   → +20", "", "", "", "", "FFE8CC", "333333", False),
    (16, "  Langkah 3: Medan Center → +15", "", "", "", "", "EDD7FF", "333333", False),
    (16, "  Langkah 4: Accessories  → +15", "", "", "", "", "D7E8FF", "333333", False),
    (16, "  Langkah 5: bulan Agustus→ +10", "", "", "", "", "D7FFE8", "333333", False),
    (20, "  TOTAL = 40+20+15+15+10 = 100  →  RISIKO TINGGI", "", "", "", "", "FFCCCC", "C0392B", True),
    (10, "", "", "", "", "", "FFFFFF", "000000", False),
    (22, "CARA VERIFIKASI DI EXCEL", "", "", "", "", "E8EFF7", "1F3864", True),
    (16, "1. Buka sheet 'Data + Risk Score'", "", "", "", "", "FFFFFF", "333333", False),
    (16, "2. Lihat kolom AD sampai AI — itu adalah langkah 1-5 dan total score", "", "", "", "", "FFFFFF", "333333", False),
    (16, "3. Klik auto-filter di header → filter kolom AI (risk_score) >= 60", "", "", "", "", "FFFFFF", "333333", False),
    (16, "4. Perhatikan kolom AD — hampir semua baris HIGH risk = 'YA (+40 poin)'", "", "", "", "", "FFFFFF", "333333", False),
    (16, "5. Itu membuktikan: transaksi HIGH risk = transaksi jual di bawah modal", "", "", "", "", "FFFFFF", "333333", False),
]

thin2 = Side(style='thin', color="DDDDDD")
b2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

for r_idx, step in enumerate(steps, 1):
    row_h, a, b, c, d, e, bg, fg, bold = step
    ws_guide.row_dimensions[r_idx].height = row_h
    for c_idx, val in enumerate([a, b, c, d, e], 1):
        cell = ws_guide.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=10 if bold else 9)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = b2

print("Saving updated workbook...")
wb.save(PATH_OUT)
print(f"Done! Saved to: {PATH_OUT}")
