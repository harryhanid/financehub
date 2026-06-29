import sys, csv
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

SRC  = r'C:\Users\25010160\Downloads\Cleaned Data - combined_data.csv'
DEST = r'C:\Users\25010160\Downloads\UNIQLO_Week7_FINAL.xlsx'

# ── Number parsers ────────────────────────────────────────────────────────────
def parse_idr(s):
    """'Rp1,234,567' / '-Rp1,234,567' / '0' → float or None"""
    if not s or s.strip() in ('', 'NULL', '#N/A', '0'):
        return 0.0
    s = s.strip().replace('Rp', '').replace(',', '').replace(' ', '')
    # handle negative like "-1234567"
    try:
        return float(s)
    except:
        return None

def parse_int(s, fallback=None):
    """'3' / '0.00' / '1.00' → int or fallback"""
    if not s or s.strip() in ('', 'NULL', '#N/A'):
        return fallback
    try:
        return int(float(s.strip()))
    except:
        return fallback

def parse_float_plain(s, fallback=None):
    """'0.00' / '1.00' → float or fallback"""
    if not s or s.strip() in ('', 'NULL', '#N/A'):
        return fallback
    try:
        return float(s.strip())
    except:
        return fallback

# Column name → converter function
# discount_% stays as TEXT (formulas reference "20%", "10%" as text strings)
NUMERIC_COLS = {
    'quantity':         lambda v: parse_int(v, 0),
    'returned':         lambda v: parse_int(v, 0),
    'store_size_m2':    lambda v: parse_int(v, None),
    'age':              lambda v: parse_int(v, None),
    'cost_price':       lambda v: parse_idr(v),
    'list_price':       lambda v: parse_idr(v),
    'Total_Sales':      lambda v: parse_idr(v),
    'Return':           lambda v: parse_idr(v),
    'Total_Discount':   lambda v: parse_idr(v),
    'Gross_Profit':     lambda v: parse_idr(v),
    'Total_Cost_Price': lambda v: parse_idr(v),
    'Net_Sales':        lambda v: parse_idr(v),
}

# ── Style helpers ─────────────────────────────────────────────────────────────
thin  = Side(style='thin', color="CCCCCC")
bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
NAVY  = "1F3864"; WHITE = "FFFFFF"; LIGHT = "DCE6F1"
IDR_FMT = '#,##0'          # no Rp prefix, just comma-thousands (e.g. 1,234,567)
IDR_FMT_NEG = '#,##0;[Red]-#,##0'

def hdr_cell(cell, bg=NAVY, fg=WHITE):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr

def dat_cell(cell, bg=WHITE, num_fmt=None, center=False):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(size=9, color="222222")
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = bdr
    if num_fmt:
        cell.number_format = num_fmt

# ── Read CSV ──────────────────────────────────────────────────────────────────
print("Reading CSV...")
with open(SRC, encoding='utf-8') as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    rows = list(reader)
print(f"  {len(rows):,} rows, {len(fieldnames)} columns")

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Data + Risk Score
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Data + Risk Score"

new_headers = list(fieldnames) + ['risk_score']
total_cols   = len(new_headers)

# Money columns (for number format)
MONEY_COLS = {'cost_price','list_price','Total_Sales','Return',
              'Total_Discount','Gross_Profit','Total_Cost_Price','Net_Sales'}
INT_COLS   = {'quantity','returned','store_size_m2','age'}

# Column letter lookup
col_letter = {h: get_column_letter(i+1) for i, h in enumerate(new_headers)}

# ── Header row ────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 30
for c_idx, h in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=c_idx, value=h)
    if h == 'risk_score':
        hdr_cell(cell, bg="1F5C2E")
    else:
        hdr_cell(cell)

# ── Data rows ─────────────────────────────────────────────────────────────────
print("Writing data rows...")
for r_idx, row_dict in enumerate(rows, 2):
    alt_bg = LIGHT if r_idx % 2 == 0 else WHITE

    for c_idx, h in enumerate(fieldnames, 1):
        raw = row_dict[h]

        # Convert to correct type
        if h in NUMERIC_COLS:
            val = NUMERIC_COLS[h](raw)
        else:
            val = raw.strip() if raw else None
            # Clean up garbage values
            if val in ('NULL', '#N/A', 'null', '???'):
                val = None

        cell = ws.cell(row=r_idx, column=c_idx, value=val)

        # Number format
        if h in MONEY_COLS and isinstance(val, (int, float)):
            dat_cell(cell, bg=alt_bg, num_fmt=IDR_FMT_NEG)
        elif h in INT_COLS:
            dat_cell(cell, bg=alt_bg, num_fmt='0', center=True)
        elif h == 'discount_%':
            dat_cell(cell, bg=alt_bg, center=True)
        else:
            dat_cell(cell, bg=alt_bg)

    # ── Risk score (column 29) ────────────────────────────────────────────────
    # Formula: reads cost_price(N) and list_price(O) — now numeric, so N2>O2 works
    R = r_idx
    B = col_letter['date'];        G = col_letter['discount_%']
    I = col_letter['category'];    N = col_letter['cost_price']
    O = col_letter['list_price'];  T = col_letter['store_name']

    formula = (
        f'=MIN(100,'
        f'IF({N}{R}>{O}{R},40,0)'
        f'+IF(OR({G}{R}="20%",{G}{R}="30%"),20,IF({G}{R}="10%",10,0))'
        f'+IF({T}{R}="Medan Center",15,0)'
        f'+IF({I}{R}="Accessories",15,0)'
        f'+IF(OR(MONTH(DATEVALUE({B}{R}))=2,MONTH(DATEVALUE({B}{R}))=5,'
        f'MONTH(DATEVALUE({B}{R}))=8,MONTH(DATEVALUE({B}{R}))=10),10,0)'
        f')'
    )

    sc_cell = ws.cell(row=r_idx, column=total_cols, value=formula)
    sc_cell.number_format = '0'
    sc_cell.font = Font(size=9, bold=True, color="FFFFFF")
    sc_cell.alignment = Alignment(horizontal="center", vertical="center")
    sc_cell.border = bdr
    # color by pre-computed score (from original NUMERIC_COLS logic for display)
    cost  = NUMERIC_COLS['cost_price'](row_dict['cost_price']) or 0
    lst   = NUMERIC_COLS['list_price'](row_dict['list_price']) or 0
    disc  = row_dict['discount_%'].strip()
    store = row_dict['store_name'].strip()
    cat   = row_dict['category'].strip()
    try: mo = int(row_dict['date'].split('-')[1])
    except: mo = 0
    sc = 0
    if cost > lst > 0: sc += 40
    if disc in ('20%','30%'): sc += 20
    elif disc == '10%': sc += 10
    if store == 'Medan Center': sc += 15
    if cat == 'Accessories': sc += 15
    if mo in (2,5,8,10): sc += 10
    sc = min(sc, 100)
    if sc >= 60:   sc_cell.fill = PatternFill("solid", fgColor="C0392B")
    elif sc >= 30: sc_cell.fill = PatternFill("solid", fgColor="E67E22")
    else:          sc_cell.fill = PatternFill("solid", fgColor="27AE60")

    if r_idx % 5000 == 0:
        print(f"  {r_idx:,}/{len(rows)+1:,} rows...")

# ── Column widths & freeze ────────────────────────────────────────────────────
col_widths = {
    1:14, 2:12, 3:10, 4:8, 5:10, 6:8, 7:9, 8:8, 9:13, 10:8,
    11:6, 12:9, 13:12, 14:14, 15:14, 16:6, 17:8, 18:12, 19:24,
    20:18, 21:12, 22:12, 23:16, 24:14, 25:16, 26:16, 27:16, 28:16,
    29:12,
}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Cara Baca Formula
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Cara Baca Formula")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 42

guide = [
    ("CARA MEMBACA FORMULA RISK SCORE", "", True, NAVY, WHITE, 14),
    ("File: UNIQLO_Week7_StepByStep.xlsx  |  Sheet: Data + Risk Score", "", False, "2E4A8B", WHITE, 9),
    ("", "", False, WHITE, "000000", 9),
    ("KONSEP DASAR", "", True, "E8EFF7", "1F3864", 11),
    ("Setiap transaksi dapat 'nilai risiko' seperti kuis — makin tinggi makin berisiko rugi.", "", False, WHITE, "333333", 9),
    ("", "", False, WHITE, "000000", 9),
    ("LANGKAH", "PERTANYAAN / KONDISI", True, NAVY, WHITE, 10),
    ("Langkah 1 (+40)  cost_price > list_price?", "Produk dijual DI BAWAH harga modal → kerugian terjamin", False, "FFD7D7", "333333", 9),
    ("Langkah 2 (+20)  discount >= 20%?", "Margin hancur, volume tetap sama → diskon tidak efektif", False, "FFE8CC", "333333", 9),
    ("Langkah 2 (+10)  discount = 10%?", "Erosi margin sedang (alternatif jika bukan 20%/30%)", False, "FFF5E0", "333333", 9),
    ("Langkah 3 (+15)  store = Medan Center?", "Return rate tertinggi (10.48%) + produktivitas terendah", False, "EDD7FF", "333333", 9),
    ("Langkah 4 (+15)  category = Accessories?", "Return rate 10.38% — tertinggi di semua kategori", False, "D7E8FF", "333333", 9),
    ("Langkah 5 (+10)  Bulan Feb/Mei/Ags/Okt?", "Pola return spike konsisten 5 tahun berturut-turut", False, "D7FFE8", "333333", 9),
    ("TOTAL SCORE (0–100)", "Jumlah semua poin, maksimum 100", True, NAVY, WHITE, 10),
    ("", "", False, WHITE, "000000", 9),
    ("ARTI SCORE", "", True, "E8EFF7", "1F3864", 11),
    ("Score >= 60  =  RISIKO TINGGI  (merah)", "97.2% transaksi ini adalah below-cost SKU", True, "FFCCCC", "C0392B", 9),
    ("Score 30–59  =  RISIKO SEDANG  (kuning)", "49% transaksi ini adalah below-cost SKU", False, "FFF2CC", "8B6914", 9),
    ("Score < 30   =  RISIKO RENDAH  (hijau)", "0% transaksi ini adalah below-cost SKU", False, "E2EFDA", "1F5C2E", 9),
    ("", "", False, WHITE, "000000", 9),
    ("FORMULA DI EXCEL (kolom risk_score / kolom AC)", "", True, "E8EFF7", "1F3864", 11),
    (" =MIN(100,IF(N2>O2,40,0)", "", False, "F0FFF0", "1F5C2E", 8),
    (" +IF(OR(G2=\"20%\",G2=\"30%\"),20,IF(G2=\"10%\",10,0))", "", False, "F0FFF0", "1F5C2E", 8),
    (" +IF(T2=\"Medan Center\",15,0)", "", False, "F0FFF0", "1F5C2E", 8),
    (" +IF(I2=\"Accessories\",15,0)", "", False, "F0FFF0", "1F5C2E", 8),
    (" +IF(OR(MONTH(DATEVALUE(B2))=2,MONTH(DATEVALUE(B2))=5,", "", False, "F0FFF0", "1F5C2E", 8),
    ("   MONTH(DATEVALUE(B2))=8,MONTH(DATEVALUE(B2))=10),10,0))", "", False, "F0FFF0", "1F5C2E", 8),
    ("", "", False, WHITE, "000000", 9),
    ("CATATAN PENTING", "", True, "FFF9E6", "8B6914", 10),
    ("• cost_price dan list_price sekarang ANGKA (bukan text) — formula perbandingan bekerja", "", False, WHITE, "333333", 9),
    ("• discount_% sengaja dibiarkan text ('0%','10%','20%','30%') karena formula IF pakai string", "", False, WHITE, "333333", 9),
    ("• DATEVALUE(B2) digunakan karena kolom date tersimpan sebagai text (YYYY-MM-DD)", "", False, WHITE, "333333", 9),
    ("• Kolom uang (cost_price, Total_Sales, dll) format '#,##0' — angka riil, tampil dengan koma", "", False, WHITE, "333333", 9),
]

for r_idx, (a, b, bold, bg, fg, fsize) in enumerate(guide, 1):
    ws2.row_dimensions[r_idx].height = 18
    for c_idx, val in enumerate([a, b], 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=fsize)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = bdr

# ── Save ──────────────────────────────────────────────────────────────────────
print("Saving...")
wb.save(DEST)
print(f"Done! → {DEST}")
