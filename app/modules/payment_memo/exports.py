"""PAM export helpers — PDF (ReportLab) and Excel (openpyxl)."""
import io
import config
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, BaseDocTemplate, PageTemplate, Frame, NextPageTemplate,
    Table, TableStyle, Paragraph, Spacer, PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.payment_memo.service import get_pam_detail, get_pam_payments, get_pam_payments_detail

_BLUE   = colors.HexColor("#1e3a5f")
_GOLD   = colors.HexColor("#f59e0b")
_GRAY   = colors.HexColor("#e2e8f0")
_LGRAY  = colors.HexColor("#f8fafc")
_WHITE  = colors.white
_AMBER  = colors.HexColor("#fffbeb")


def _fmt_date(s: str) -> str:
    try:
        dt = datetime.strptime(s[:10], "%Y-%m-%d")
        return dt.strftime("%d %B %Y").lstrip("0")
    except Exception:
        return s or ""


def _fmt_rp(v) -> str:
    try:
        return f"Rp {float(v):,.0f}"
    except Exception:
        return str(v)


def _style(name, **kw):
    defaults = {"fontName": "Helvetica", "fontSize": 8, "leading": 10}
    defaults.update(kw)
    return ParagraphStyle(name, **defaults)


_S_LABEL   = _style("lbl",  fontName="Helvetica-Bold", fontSize=8,
                     textColor=colors.HexColor("#374151"))
_S_VAL     = _style("val",  fontSize=8.5, textColor=colors.HexColor("#1e293b"))
_S_VAL_HI  = _style("valhi", fontName="Helvetica-Bold", fontSize=9,
                     textColor=colors.HexColor("#1e3a5f"))
_S_SECTION = _style("sec",  fontName="Helvetica-Bold", fontSize=7.5,
                     textColor=colors.HexColor("#374151"))
_S_CB      = _style("cb",   fontSize=8, textColor=colors.HexColor("#1e293b"))
_S_TITLE   = _style("ttl",  fontName="Helvetica-Bold", fontSize=13,
                     alignment=TA_CENTER, textColor=_WHITE)
_S_ATCH_H  = _style("atch", fontName="Helvetica-Bold", fontSize=9, textColor=_WHITE)
_S_TH      = _style("th",   fontName="Helvetica-Bold", fontSize=8,
                     textColor=_WHITE, alignment=TA_CENTER)
_S_TD      = _style("td",   fontSize=8, textColor=colors.HexColor("#1e293b"))
_S_TD_R    = _style("tdr",  fontSize=8, alignment=2,
                     textColor=colors.HexColor("#1e293b"))

_CW = [3.8*cm, 4.7*cm, 3.8*cm, 4.7*cm]


def _p(s, style=None):
    return Paragraph(str(s) if s is not None else "", style or _S_VAL)


def _terlampir():
    return _p("Terlampir", _style("tl", fontName="Helvetica-Bold",
                                  fontSize=8, textColor=colors.HexColor("#92400e"),
                                  backColor=colors.HexColor("#fef3c7")))


def _build_pam_table(pam: dict, approved_by_1: str, approved_by_2: str) -> Table:
    pam_date    = _fmt_date(pam.get("pam_date", ""))
    due_date    = _fmt_date(pam.get("due_date", ""))
    invoice_amt = _fmt_rp(pam.get("total_amount", 0))

    data = [
        [_p("PAM No.", _S_LABEL), _p(pam.get("pam_no", ""), _S_VAL_HI),
         _p("Cost Center", _S_LABEL), _p(pam.get("cost_center", ""), _S_VAL)],
        [_p("Date", _S_LABEL), _p(pam_date, _S_VAL),
         _p("GL Account", _S_LABEL), _p(str(pam.get("gl_account", "")), _S_VAL)],
        [_p("Requestor's Name", _S_LABEL), _p(pam.get("requestors_name", ""), _S_VAL),
         _p("SO / SC", _S_LABEL), _p("", _S_VAL)],
        [_p("Department", _S_LABEL), _p("-", _S_VAL),
         _p("Company", _S_LABEL), _p(pam.get("pt", ""), _S_VAL)],
        [_p("Business Unit", _S_SECTION), "", "", ""],
        [_p("☐  Upstream          ☐  Downstream          ☑  Corporate", _S_CB),
         "", "", ""],
        [_p("Type of Request", _S_SECTION), "", "", ""],
        [_p("☐  Downpayment to vendor", _S_CB), "", "", ""],
        [_p("☑  Invoice Payment – Non PO Invoice", _S_CB), "", "", ""],
        [_p("☐  Employee Advance / Reimbursement (Fund Transfer)", _S_CB), "", "", ""],
        [_p("Invoice Information", _S_SECTION), "", "", ""],
        [_p("Vendor Name", _S_LABEL), _terlampir(),
         _p("Invoice / Memo No", _S_LABEL), _p("-", _S_VAL)],
        [_p("Invoice Amount", _S_LABEL), _p(invoice_amt, _S_VAL_HI),
         _p("Expected Due Date", _S_LABEL), _p(due_date, _S_VAL)],
        [_p("Vendor Bank Account Details", _S_SECTION), "", "", ""],
        [_p("Bank Account Name", _S_LABEL),  _terlampir(), "", ""],
        [_p("Bank Name", _S_LABEL),          _terlampir(), "", ""],
        [_p("Bank Account Number", _S_LABEL), _terlampir(), "", ""],
    ]

    span_rows = [4, 5, 6, 7, 8, 9, 10, 13]
    style_cmds = [
        ("FONTNAME",      (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("BACKGROUND",    (0, 0), (-1, -1), _LGRAY),
        ("BACKGROUND",    (0, 1), (-1, 1),  _WHITE),
        ("BACKGROUND",    (0, 3), (-1, 3),  _WHITE),
        ("BACKGROUND",    (0, 4),  (-1, 4),  _GRAY),
        ("BACKGROUND",    (0, 6),  (-1, 6),  _GRAY),
        ("BACKGROUND",    (0, 10), (-1, 10), _GRAY),
        ("BACKGROUND",    (0, 13), (-1, 13), _GRAY),
        ("BACKGROUND",    (0, 5),  (-1, 5),  _AMBER),
        ("BACKGROUND",    (0, 7),  (-1, 9),  _AMBER),
    ]
    for r in span_rows:
        style_cmds.append(("SPAN", (0, r), (3, r)))
    for r in [14, 15, 16]:
        style_cmds.append(("SPAN", (1, r), (3, r)))

    return Table(data, colWidths=_CW, style=TableStyle(style_cmds))


def _build_signature_table(requestors_name: str,
                           approved_by_1: str,
                           approved_by_2: str) -> Table:
    data = [
        [_p("Request by", _S_LABEL),  "", _p("Approved by", _S_LABEL), ""],
        ["", "", "", ""],
        ["", "", "", ""],
        ["", "", "", ""],
        [_p(requestors_name, _style("sn", fontName="Helvetica-Bold", fontSize=8)),
         "",
         _p(f"{approved_by_1}          {approved_by_2}",
            _style("sn2", fontName="Helvetica-Bold", fontSize=8)),
         ""],
        [_p("Checked by (QA)", _S_LABEL), "", "", ""],
        ["", "", "", ""],
        ["", "", "", ""],
        ["", "", "", ""],
    ]
    cw = [4.25*cm, 4.25*cm, 4.25*cm, 4.25*cm]
    style_cmds = [
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("BACKGROUND",    (0, 0), (-1, -1), _LGRAY),
        ("SPAN", (0, 0), (1, 0)), ("SPAN", (0, 1), (1, 1)),
        ("SPAN", (0, 2), (1, 2)), ("SPAN", (0, 3), (1, 3)),
        ("SPAN", (0, 4), (1, 4)),
        ("SPAN", (2, 0), (3, 0)), ("SPAN", (2, 1), (3, 1)),
        ("SPAN", (2, 2), (3, 2)), ("SPAN", (2, 3), (3, 3)),
        ("SPAN", (2, 4), (3, 4)),
        ("SPAN", (0, 5), (3, 5)), ("SPAN", (0, 6), (3, 6)),
        ("SPAN", (0, 7), (3, 7)), ("SPAN", (0, 8), (3, 8)),
        ("LINEBELOW", (0, 3), (1, 3), 0.8, colors.HexColor("#94a3b8")),
        ("LINEBELOW", (2, 3), (3, 3), 0.8, colors.HexColor("#94a3b8")),
        ("LINEBELOW", (0, 8), (3, 8), 0.8, colors.HexColor("#94a3b8")),
    ]
    return Table(data, colWidths=cw, style=TableStyle(style_cmds))


def _build_pam_table_custom(data: dict) -> Table:
    pam_date    = _fmt_date(data.get("pam_date", ""))
    due_date    = _fmt_date(data.get("due_date", ""))
    invoice_amt = _fmt_rp(data.get("total_amount", 0))

    def _cb(val):
        return "☑" if val else "☐"

    def _maybe_terlampir(val):
        s = str(val) if val is not None else ""
        if not s or s.strip().lower() == "terlampir":
            return _terlampir()
        return _p(s, _S_VAL)

    bu_row  = f"{_cb(data.get('bu_upstream'))}  Upstream          "
    bu_row += f"{_cb(data.get('bu_downstream'))}  Downstream          "
    bu_row += f"{_cb(data.get('bu_corporate'))}  Corporate"

    data_rows = [
        [_p("PAM No.", _S_LABEL), _p(data.get("pam_no", ""), _S_VAL_HI),
         _p("Cost Center", _S_LABEL), _p(data.get("cost_center", ""), _S_VAL)],
        [_p("Date", _S_LABEL), _p(pam_date, _S_VAL),
         _p("GL Account", _S_LABEL), _p(str(data.get("gl_account", "")), _S_VAL)],
        [_p("Requestor's Name", _S_LABEL), _p(data.get("requestors_name", ""), _S_VAL),
         _p("SO / SC", _S_LABEL), _p(data.get("so_sc", ""), _S_VAL)],
        [_p("Department", _S_LABEL), _p(data.get("department", "-"), _S_VAL),
         _p("Company", _S_LABEL), _p(data.get("pt", ""), _S_VAL)],
        [_p("Business Unit", _S_SECTION), "", "", ""],
        [_p(bu_row, _S_CB), "", "", ""],
        [_p("Type of Request", _S_SECTION), "", "", ""],
        [_p(f"{_cb(data.get('type_downpayment'))}  Downpayment to vendor", _S_CB),
         "", "", ""],
        [_p(f"{_cb(data.get('type_invoice'))}  Invoice Payment – Non PO Invoice", _S_CB),
         "", "", ""],
        [_p(f"{_cb(data.get('type_advance'))}  Employee Advance / Reimbursement (Fund Transfer)",
            _S_CB), "", "", ""],
        [_p("Invoice Information", _S_SECTION), "", "", ""],
        [_p("Vendor Name", _S_LABEL), _maybe_terlampir(data.get("vendor_name", "")),
         _p("Invoice / Memo No", _S_LABEL), _p(data.get("invoice_memo_no", "-"), _S_VAL)],
        [_p("Invoice Amount", _S_LABEL), _p(invoice_amt, _S_VAL_HI),
         _p("Expected Due Date", _S_LABEL), _p(due_date, _S_VAL)],
        [_p("Vendor Bank Account Details", _S_SECTION), "", "", ""],
        [_p("Bank Account Name", _S_LABEL),
         _maybe_terlampir(data.get("bank_account_name", "")), "", ""],
        [_p("Bank Name", _S_LABEL),
         _maybe_terlampir(data.get("bank_name", "")), "", ""],
        [_p("Bank Account Number", _S_LABEL),
         _maybe_terlampir(data.get("bank_account_no", "")), "", ""],
    ]

    span_rows = [4, 5, 6, 7, 8, 9, 10, 13]
    style_cmds = [
        ("FONTNAME",      (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("BACKGROUND",    (0, 0), (-1, -1), _LGRAY),
        ("BACKGROUND",    (0, 1), (-1, 1),  _WHITE),
        ("BACKGROUND",    (0, 3), (-1, 3),  _WHITE),
        ("BACKGROUND",    (0, 4),  (-1, 4),  _GRAY),
        ("BACKGROUND",    (0, 6),  (-1, 6),  _GRAY),
        ("BACKGROUND",    (0, 10), (-1, 10), _GRAY),
        ("BACKGROUND",    (0, 13), (-1, 13), _GRAY),
        ("BACKGROUND",    (0, 5),  (-1, 5),  _AMBER),
        ("BACKGROUND",    (0, 7),  (-1, 9),  _AMBER),
    ]
    for r in span_rows:
        style_cmds.append(("SPAN", (0, r), (3, r)))
    for r in [14, 15, 16]:
        style_cmds.append(("SPAN", (1, r), (3, r)))
    return Table(data_rows, colWidths=_CW, style=TableStyle(style_cmds))


def _build_detail_pdf_table(detail: list) -> Table:
    _s7h = _style("d7h", fontName="Helvetica-Bold", fontSize=7,
                  textColor=_WHITE, alignment=TA_CENTER)
    _s7  = _style("d7",  fontSize=7, textColor=colors.HexColor("#1e293b"))
    _s7r = _style("d7r", fontSize=7, alignment=2,
                  textColor=colors.HexColor("#1e293b"))
    _s7c = _style("d7c", fontSize=7, alignment=1,
                  textColor=colors.HexColor("#1e293b"))

    col_w = [0.6*cm, 3.8*cm, 3.2*cm, 2.4*cm, 2.4*cm, 2.4*cm,
             2.5*cm, 3.0*cm, 1.8*cm, 1.8*cm, 1.8*cm]

    hdrs = [_p(h, _s7h) for h in [
        "No", "Nama Siswa", "Keterangan",
        "By\nPend", "By\nTunj", "By\nRiset",
        "Total\nPbyran", "No. Rekening",
        "Sisa\nPend", "Sisa\nTunj", "Sisa\nRiset",
    ]]
    rows = [hdrs]
    grand_total = 0.0

    for siswa in detail:
        sis_rows = siswa.get("rows", [])
        grand_total += float(siswa.get("total_pembayaran") or 0)
        for idx, pr in enumerate(sis_rows):
            first = (idx == 0)
            rows.append([
                _p(str(siswa["no"]) if first else "", _s7c),
                _p((siswa.get("nama") or siswa.get("siswa_code") or "") if first else "", _s7),
                _p(pr.get("keterangan") or "", _s7),
                _p(f"{pr['pendidikan']:,.0f}" if pr.get("pendidikan") else "", _s7r),
                _p(f"{pr['tunjangan']:,.0f}"  if pr.get("tunjangan")  else "", _s7r),
                _p(f"{pr['penelitian']:,.0f}" if pr.get("penelitian") else "", _s7r),
                _p(f"{siswa['total_pembayaran']:,.0f}" if first else "", _s7r),
                _p((siswa.get("norek") or "") if first else "", _s7),
                _p(f"{siswa['sisa_pendidikan']:,.0f}" if first else "", _s7r),
                _p(f"{siswa['sisa_tunjangan']:,.0f}"  if first else "", _s7r),
                _p(f"{siswa['sisa_penelitian']:,.0f}" if first else "", _s7r),
            ])

    rows.append([_p("", _s7)] * 6 + [
        _p(f"{grand_total:,.0f}",
           _style("gtv7", fontName="Helvetica-Bold", fontSize=7, alignment=2,
                  textColor=colors.HexColor("#1e293b"))),
    ] + [_p("", _s7)] * 4)

    return Table(rows, colWidths=col_w, repeatRows=1, style=TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  _BLUE),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [_WHITE, _LGRAY]),
        ("BACKGROUND",    (0, -1),(-1, -1), colors.HexColor("#e8f0fe")),
        ("FONTNAME",      (0, -1),(-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",     (0, -1),(-1, -1), 1.2, _BLUE),
        ("ALIGN",         (0, 0), (0, -1),  "CENTER"),
        ("ALIGN",         (3, 0), (6, -1),  "RIGHT"),
        ("ALIGN",         (8, 0), (10, -1), "RIGHT"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
    ]))


def _group_payments_by_siswa(payments: list) -> list:
    """Aggregate individual payment rows into one entry per siswa, summing amounts."""
    from collections import OrderedDict
    grouped: dict = OrderedDict()
    for pb in payments:
        key = pb.get("siswa_code") or pb.get("nama") or ""
        if key not in grouped:
            grouped[key] = {
                "nama":   pb.get("nama") or pb.get("siswa_code", ""),
                "bank":   pb.get("bank") or "",
                "norek":  pb.get("norek") or "",
                "amount": 0.0,
            }
        grouped[key]["amount"] += float(pb.get("amount") or 0)
    return list(grouped.values())


def export_pam_pdf_custom(data: dict, payments: list) -> bytes:
    approved_by_1 = data.get("approved_by_1") or config.PAM_APPROVED_BY_1
    approved_by_2 = data.get("approved_by_2") or config.PAM_APPROVED_BY_2
    pam_no        = data.get("pam_no", "")

    buf = io.BytesIO()
    _lm, _rm, _tm, _bm = 2*cm, 2*cm, 1.5*cm, 1.5*cm
    _A4L = landscape(A4)
    _port_frame = Frame(_lm, _bm, A4[0]-_lm-_rm, A4[1]-_tm-_bm, id='pf')
    _land_frame = Frame(_lm, _bm, _A4L[0]-_lm-_rm, _A4L[1]-_tm-_bm, id='lf')
    doc = BaseDocTemplate(buf,
        pageTemplates=[
            PageTemplate(id='portrait',  frames=[_port_frame],  pagesize=A4),
            PageTemplate(id='landscape', frames=[_land_frame],  pagesize=_A4L),
        ],
        leftMargin=_lm, rightMargin=_rm, topMargin=_tm, bottomMargin=_bm)
    elems = []

    title_tbl = Table([[_p("PAYMENT APPROVAL MEMO", _S_TITLE)]], colWidths=[17*cm])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _BLUE),
        ("LINEBELOW",     (0, 0), (-1, -1), 3, _GOLD),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elems.append(title_tbl)
    elems.append(Spacer(1, 0.15*cm))
    elems.append(_build_pam_table_custom(data))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(_build_signature_table(
        data.get("requestors_name", ""), approved_by_1, approved_by_2
    ))

    elems.append(PageBreak())
    att_hdr = Table(
        [[_p("Rangkuman PAM — Jadwal Pembayaran Beasiswa", _S_ATCH_H)],
         [_p(f"{pam_no}  \xb7  {data.get('pt','')}  \xb7  {_fmt_date(data.get('pam_date',''))}",
             _style("atsub", fontSize=7.5, textColor=colors.HexColor("#cbd5e1")))]],
        colWidths=[17*cm]
    )
    att_hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _BLUE),
        ("LINEBELOW",     (0, -1), (-1, -1), 3, _GOLD),
        ("TOPPADDING",    (0, 0), (0, 0), 7),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    elems.append(att_hdr)
    elems.append(Spacer(1, 0.3*cm))

    hdrs = ["No", "Nama Siswa", "Bank", "No. Rekening", "Amount (Rp)"]
    rows = [[_p(h, _S_TH) for h in hdrs]]
    total = 0.0
    for i, pb in enumerate(_group_payments_by_siswa(payments), 1):
        rows.append([
            _p(str(i), _style("n", alignment=1, fontSize=8)),
            _p(pb["nama"], _S_TD),
            _p(pb["bank"], _S_TD),
            _p(pb["norek"], _S_TD),
            _p(f"{pb['amount']:,.0f}", _S_TD_R),
        ])
        total += pb["amount"]
    rows.append([
        _p("", _S_TD), _p("", _S_TD), _p("", _S_TD),
        _p("TOTAL", _style("tot", fontName="Helvetica-Bold", fontSize=8, alignment=2)),
        _p(f"{total:,.0f}", _style("totv", fontName="Helvetica-Bold", fontSize=8, alignment=2)),
    ])
    col_w = [0.6*cm, 5.0*cm, 2.8*cm, 4.6*cm, 4.0*cm]
    att_tbl = Table(rows, colWidths=col_w, repeatRows=1)
    att_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  _BLUE),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1), 8),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [_WHITE, _LGRAY]),
        ("BACKGROUND",     (0, -1), (-1, -1), colors.HexColor("#e8f0fe")),
        ("FONTNAME",       (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",      (0, -1), (-1, -1), 1.5, _BLUE),
        ("ALIGN",          (0, 0), (0, -1), "CENTER"),
        ("ALIGN",          (4, 0), (4, -1), "RIGHT"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
        ("LEFTPADDING",    (0, 0), (-1, -1), 4),
    ]))
    elems.append(att_tbl)

    # Page 3: Detail PAM
    _pam_no_det = data.get("pam_no", "")
    _cid_det    = int(data.get("company_id") or 0)
    if _pam_no_det and _cid_det:
        _detail = get_pam_payments_detail(_pam_no_det, _cid_det)
        if _detail:
            _land_w = _A4L[0] - _lm - _rm
            elems.append(NextPageTemplate('landscape'))
            elems.append(PageBreak())
            _det_hdr = Table(
                [[_p("Detail PAM", _S_ATCH_H)],
                 [_p(f"{_pam_no_det}  \xb7  {data.get('pt','')}  \xb7  {_fmt_date(data.get('pam_date',''))}",
                     _style("dtsub", fontSize=7.5, textColor=colors.HexColor("#cbd5e1")))]],
                colWidths=[_land_w]
            )
            _det_hdr.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), _BLUE),
                ("LINEBELOW",     (0, -1), (-1, -1), 3, _GOLD),
                ("TOPPADDING",    (0, 0), (0, 0), 7),
                ("BOTTOMPADDING", (0, -1), (-1, -1), 6),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ]))
            elems.append(_det_hdr)
            elems.append(Spacer(1, 0.3*cm))
            elems.append(_build_detail_pdf_table(_detail))

    doc.build(elems)
    buf.seek(0)
    return buf.read()


def export_pam_pdf(pam_id: int, company_id: int,
                   approved_by_1: str, approved_by_2: str) -> bytes:
    pam = get_pam_detail(pam_id, company_id)
    if not pam:
        raise ValueError("PAM record tidak ditemukan.")

    approved_by_1 = approved_by_1 or config.PAM_APPROVED_BY_1
    approved_by_2 = approved_by_2 or config.PAM_APPROVED_BY_2
    pam_no        = pam["pam_no"]
    payments      = get_pam_payments(pam_no, company_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    elems = []

    # Page 1: PAM Form
    title_data = [[_p("PAYMENT APPROVAL MEMO", _S_TITLE)]]
    title_tbl  = Table(title_data, colWidths=[17*cm])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _BLUE),
        ("LINEBELOW",     (0, 0), (-1, -1), 3, _GOLD),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elems.append(title_tbl)
    elems.append(Spacer(1, 0.15*cm))
    elems.append(_build_pam_table(pam, approved_by_1, approved_by_2))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(_build_signature_table(
        pam.get("requestors_name", ""), approved_by_1, approved_by_2
    ))

    # Page 2: Lampiran
    elems.append(PageBreak())

    att_hdr_tbl = Table(
        [[_p(f"Rangkuman PAM — Jadwal Pembayaran Beasiswa", _S_ATCH_H)],
         [_p(f"{pam_no}  ·  {pam.get('pt','')}  ·  {_fmt_date(pam.get('pam_date',''))}",
             _style("atsub", fontSize=7.5, textColor=colors.HexColor("#cbd5e1")))]],
        colWidths=[17*cm]
    )
    att_hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _BLUE),
        ("LINEBELOW",     (0, -1), (-1, -1), 3, _GOLD),
        ("TOPPADDING",    (0, 0), (0, 0), 7),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    elems.append(att_hdr_tbl)
    elems.append(Spacer(1, 0.3*cm))

    headers = ["No", "Nama Siswa", "Bank", "No. Rekening", "Amount (Rp)"]
    rows = [[_p(h, _S_TH) for h in headers]]
    total = 0.0
    for i, pb in enumerate(_group_payments_by_siswa(payments), 1):
        rows.append([
            _p(str(i), _style("n", alignment=1, fontSize=8)),
            _p(pb["nama"], _S_TD),
            _p(pb["bank"], _S_TD),
            _p(pb["norek"], _S_TD),
            _p(f"{pb['amount']:,.0f}", _S_TD_R),
        ])
        total += pb["amount"]
    rows.append([
        _p("", _S_TD), _p("", _S_TD), _p("", _S_TD),
        _p("TOTAL", _style("tot", fontName="Helvetica-Bold", fontSize=8, alignment=2)),
        _p(f"{total:,.0f}", _style("totv", fontName="Helvetica-Bold", fontSize=8, alignment=2)),
    ])

    col_w = [0.6*cm, 5.0*cm, 2.8*cm, 4.6*cm, 4.0*cm]
    att_tbl = Table(rows, colWidths=col_w, repeatRows=1)
    att_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  _BLUE),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1), 8),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [_WHITE, _LGRAY]),
        ("BACKGROUND",     (0, -1), (-1, -1), colors.HexColor("#e8f0fe")),
        ("FONTNAME",       (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",      (0, -1), (-1, -1), 1.5, _BLUE),
        ("ALIGN",          (0, 0), (0, -1), "CENTER"),
        ("ALIGN",          (4, 0), (4, -1), "RIGHT"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
        ("LEFTPADDING",    (0, 0), (-1, -1), 4),
    ]))
    elems.append(att_tbl)

    doc.build(elems)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Excel export — Book9.xlsx Detail PAM (Sheet 3)
# ─────────────────────────────────────────────────────────────────────────────

def _build_detail_sheet(ws, pam: dict, detail: list) -> None:
    from openpyxl.utils import range_boundaries
    from datetime import datetime as _dt

    _thin = Side(style="thin")
    _bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hf   = PatternFill("solid", fgColor="D9D9D9")
    _nfmt = '#,##0'

    def _s(r, c, val=None, bold=False, sz=10, ha="general", va="center",
           fill=None, bdr=None, fmt=None):
        cell = ws.cell(r, c)
        if val is not None:
            cell.value = val
        cell.font      = Font(bold=bold, size=sz)
        cell.alignment = Alignment(horizontal=ha, vertical=va)
        if fill: cell.fill   = fill
        if bdr:  cell.border = bdr
        if fmt:  cell.number_format = fmt
        return cell

    def _outer_border(merge_str):
        min_col, min_row, max_col, max_row = range_boundaries(merge_str)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                prev = ws.cell(row, col).border
                ws.cell(row, col).border = Border(
                    left   = _thin if col == min_col else prev.left,
                    right  = _thin if col == max_col else prev.right,
                    top    = _thin if row == min_row else prev.top,
                    bottom = _thin if row == max_row else prev.bottom,
                )

    def _hdr(r, c, val):
        cell = ws.cell(r, c, val)
        cell.font      = Font(bold=True, size=10)
        cell.fill      = _hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _bdr
        return cell

    def _merge_hdr(merge_str, val):
        ws.merge_cells(merge_str)
        min_col, min_row, max_col, max_row = range_boundaries(merge_str)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).fill = _hf
        c = ws.cell(min_row, min_col, val)
        c.font      = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        _outer_border(merge_str)

    def _data(r, c, val, ha="left", fmt=None, bold=False):
        cell = ws.cell(r, c, val)
        cell.font      = Font(size=10, bold=bold)
        cell.alignment = Alignment(horizontal=ha, vertical="center")
        cell.border    = _bdr
        if fmt:
            cell.number_format = fmt
        return cell

    def _vert_merge(r_start, r_end, col_letter, col_num):
        if r_start == r_end:
            return
        ws.merge_cells(f"{col_letter}{r_start}:{col_letter}{r_end}")
        for r in range(r_start, r_end + 1):
            ws.cell(r, col_num).border = Border(
                left   = _thin,
                right  = _thin,
                top    = _thin if r == r_start else None,
                bottom = _thin if r == r_end   else None,
            )

    # ── Header block (rows 2-3) ──────────────────────────────────────────────
    ws.row_dimensions[2].height = 15.9
    ws.row_dimensions[3].height = 15.9

    _s(2, 2, "NO.",    bold=True, sz=12, ha="left", va="top")
    _s(2, 4, ":",      bold=True, sz=12)
    _s(2, 5, pam.get("pam_no", ""), sz=12, ha="left", va="top")
    _s(2, 9, "COST CENTER",  bold=True, sz=12, ha="left")
    _s(2, 10, ":",     bold=True, sz=12)
    _s(2, 11, pam.get("cost_center", ""), bold=True, sz=12, ha="left")

    _s(3, 2, "TANGGAL", bold=True, sz=12, ha="left")
    _s(3, 4, ":",        bold=True, sz=12)
    try:
        dv = _dt.strptime(pam.get("pam_date", "")[:10], "%Y-%m-%d")
        c3 = _s(3, 5, dv, bold=True, sz=12, ha="left")
        c3.number_format = '[$-421]dd\\ mmmm\\ yyyy;@'
    except Exception:
        _s(3, 5, pam.get("pam_date", ""), bold=True, sz=12, ha="left")
    _s(3, 9, "GL ACCOUNT",  bold=True, sz=12, ha="left")
    _s(3, 10, ":",           bold=True, sz=12)
    _s(3, 11, pam.get("gl_account", ""), bold=True, sz=12, ha="left")

    # ── 2-row column headers (rows 5-6) ──────────────────────────────────────
    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 15

    for merge_str, val in [
        ("B5:B6", "NO"), ("C5:D6", "NAMA SISWA"), ("E5:E6", "ANGKATAN"),
        ("J5:J6", "KETERANGAN"), ("K5:M5", "PEMBAYARAN"),
        ("P5:P6", "BANK"), ("R5:T5", "SISA SALDO"),
    ]:
        _merge_hdr(merge_str, val)

    for r, c, val in [
        (5, 6, "JENJANG"),   (6, 6, "STUDI"),
        (5, 7, "PROGRAM"),   (6, 7, "ETF"),
        (5, 8, "PERGURUAN"), (6, 8, "TINGGI"),
        (5, 9, "PROGRAM"),   (6, 9, "STUDI"),
        (6, 11, "PENDIDIKAN"), (6, 12, "TUNJANGAN"), (6, 13, "PENELITIAN"),
        (5, 14, "TOTAL"),    (6, 14, "PEMBAYARAN"),
        (5, 15, "NAMA"),     (6, 15, "REKENING"),
        (5, 17, "NO"),       (6, 17, "REKENING"),
        (6, 18, "BY PENDIDIKAN"), (6, 19, "BY TUNJANGAN"), (6, 20, "BY PENELITIAN"),
    ]:
        _hdr(r, c, val)

    # ── Data rows ────────────────────────────────────────────────────────────
    r0          = 7
    grand_total = 0.0

    for siswa in detail:
        rows   = siswa["rows"]
        n      = len(rows)
        r_last = r0 + n - 1

        for r in range(r0, r_last + 1):
            ws.row_dimensions[r].height = 18

        # Personal-info columns (vertically merged per siswa)
        _data(r0, 2, siswa["no"], ha="center")
        _vert_merge(r0, r_last, "B", 2)

        # NAMA SISWA: merge C:D across all rows for this siswa
        ws.merge_cells(f"C{r0}:D{r_last}")
        cell = ws.cell(r0, 3, siswa["nama"] or "")
        cell.font      = Font(size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for r in range(r0, r_last + 1):
            ws.cell(r, 3).border = Border(
                left   = _thin,
                top    = _thin if r == r0    else None,
                bottom = _thin if r == r_last else None,
            )
            ws.cell(r, 4).border = Border(
                right  = _thin,
                top    = _thin if r == r0    else None,
                bottom = _thin if r == r_last else None,
            )

        _data(r0, 5, siswa["angkatan"] or "", ha="center")
        _vert_merge(r0, r_last, "E", 5)

        _data(r0, 6, siswa["jenjang"] or "", ha="center")
        _vert_merge(r0, r_last, "F", 6)

        _data(r0, 7, siswa["program"] or "")
        _vert_merge(r0, r_last, "G", 7)

        _data(r0, 8, siswa["universitas"] or "")
        _vert_merge(r0, r_last, "H", 8)

        _data(r0, 9, siswa["fakultas"] or "")
        _vert_merge(r0, r_last, "I", 9)

        # Payment rows (keterangan + amounts) — col J now follows directly after I
        for idx, row_data in enumerate(rows):
            r = r0 + idx
            _data(r, 10, row_data["keterangan"])
            _data(r, 11, row_data["pendidikan"], ha="right", fmt=_nfmt)
            _data(r, 12, row_data["tunjangan"],  ha="right", fmt=_nfmt)
            _data(r, 13, row_data["penelitian"], ha="right", fmt=_nfmt)

        # Total + bank + saldo (vertically merged)
        _data(r0, 14, siswa["total_pembayaran"], ha="right", fmt=_nfmt)
        _vert_merge(r0, r_last, "N", 14)

        _data(r0, 15, siswa["namarek"] or "")
        _vert_merge(r0, r_last, "O", 15)

        _data(r0, 16, siswa["bank"] or "")
        _vert_merge(r0, r_last, "P", 16)

        _data(r0, 17, siswa["norek"] or "")
        _vert_merge(r0, r_last, "Q", 17)

        _data(r0, 18, siswa["sisa_pendidikan"], ha="right", fmt=_nfmt)
        _vert_merge(r0, r_last, "R", 18)

        _data(r0, 19, siswa["sisa_tunjangan"],  ha="right", fmt=_nfmt)
        _vert_merge(r0, r_last, "S", 19)

        _data(r0, 20, siswa["sisa_penelitian"], ha="right", fmt=_nfmt)
        _vert_merge(r0, r_last, "T", 20)

        grand_total += siswa["total_pembayaran"]
        r0 = r_last + 1

    # ── Grand total row ──────────────────────────────────────────────────────
    ws.merge_cells(f"B{r0}:M{r0}")
    _outer_border(f"B{r0}:M{r0}")
    c = ws.cell(r0, 14, grand_total)
    c.font         = Font(bold=True, size=10)
    c.alignment    = Alignment(horizontal="right", vertical="center")
    c.border       = _bdr
    c.number_format = _nfmt
    for col in range(15, 21):
        ws.cell(r0, col).border = _bdr
    ws.row_dimensions[r0].height = 15

    # ── Column widths ────────────────────────────────────────────────────────
    for col, w in [
        ("A",2),("B",5),("C",15),("D",4),("E",8),("F",8),("G",10),
        ("H",22),("I",13),("J",14),("K",14),("L",12),("M",12),
        ("N",14),("O",18),("P",22),("Q",14),("R",14),("S",12),("T",12),
    ]:
        ws.column_dimensions[col].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Excel export — Book6.xlsx standard format
# ─────────────────────────────────────────────────────────────────────────────

def export_pam_excel(pam_id: int, company_id: int,
                     approved_by_1: str, approved_by_2: str) -> bytes:
    from datetime import datetime as _dt

    pam = get_pam_detail(pam_id, company_id)
    if not pam:
        raise ValueError("PAM record tidak ditemukan.")

    approved_by_1 = approved_by_1 or config.PAM_APPROVED_BY_1
    approved_by_2 = approved_by_2 or config.PAM_APPROVED_BY_2
    pam_no        = pam["pam_no"]
    payments      = get_pam_payments(pam_no, company_id)

    wb = openpyxl.Workbook()

    # ── Sheet 1: PAM NEW (Book6 standard format — no fills, exact layout) ────
    ws = wb.active
    ws.title = "PAM NEW"

    # Column widths from Book6
    for col, w in [("A",3.0),("B",11.15),("C",1.84),("D",7.53),("E",6.43),
                   ("F",2.84),("G",9.0),("H",2.0),("J",4.53),("K",5.69),
                   ("L",4.69),("N",5.30),("O",1.15),("P",7.84)]:
        ws.column_dimensions[col].width = w

    # Row heights from Book6
    ws.row_dimensions[9].height  = 8.25
    ws.row_dimensions[11].height = 16.5
    ws.row_dimensions[12].height = 11.25
    ws.row_dimensions[19].height = 17.25
    ws.row_dimensions[25].height = 15.0

    _thin      = Side(style="thin")
    _box       = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _box_lr    = Border(left=_thin, top=_thin, bottom=_thin)  # no right (B42/G42)
    _C         = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _CN        = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _L         = Alignment(horizontal="left",   vertical="center")
    _R         = Alignment(horizontal="right",  vertical="center")

    def _bold(sz=11): return Font(bold=True,  size=sz)
    def _norm(sz=11): return Font(bold=False, size=sz)

    def _set(coord, val, font=None, align=None, border=None):
        c = ws[coord]
        c.value = val
        if font:   c.font      = font
        if align:  c.alignment = align
        if border: c.border    = border

    def _draw_box(r1, c1, r2, c2):
        """Draw a thin outline border on all cells of range (r1,c1):(r2,c2)."""
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                prev = cell.border
                cell.border = Border(
                    top    = _thin if r == r1 else prev.top,
                    bottom = _thin if r == r2 else prev.bottom,
                    left   = _thin if c == c1 else prev.left,
                    right  = _thin if c == c2 else prev.right,
                )

    def _dt_val(s):
        try:    return _dt.strptime(s[:10], "%Y-%m-%d")
        except: return s or ""

    pam_date = _dt_val(pam.get("pam_date", ""))
    due_date = _dt_val(pam.get("due_date", ""))

    # Row 1 — title
    ws.merge_cells("A1:Q1")
    _set("A1", "PAYMENT APPROVAL MEMO", font=_bold(11), align=_C)

    # Row 2 — empty separator
    ws.merge_cells("B2:Q2")

    # Row 4 — PAM No / Cost Center
    ws.merge_cells("L4:N4")
    _set("B4", "PAM No.",                   align=_L)
    _set("E4", ":")
    _set("F4", pam_no,                      align=_L)
    _set("L4", "Cost Center",               align=_R)
    _set("O4", ":",                          align=_R)
    _set("P4", pam.get("cost_center", ""),  align=_L)

    # Row 5 — Date / GL Account
    ws.merge_cells("F5:J5")
    ws.merge_cells("L5:N5")
    ws.merge_cells("P5:Q5")
    _set("B5", "Date",                      align=_L)
    _set("E5", ":")
    _set("F5", pam_date,                    align=_L)
    ws["F5"].number_format = '[$-421]dd\\ mmmm\\ yyyy;@'
    _set("L5", "GL Account",               align=_R)
    _set("O5", ":",                          align=_R)
    _set("P5", pam.get("gl_account", ""),  align=_L)

    # Row 6 — Requestor / SO-SC
    ws.merge_cells("L6:N6")
    _set("B6", "Requestor’s Name   ", align=_L)
    _set("E6", ":")
    _set("F6", pam.get("requestors_name", ""), align=_L)
    _set("L6", "SO / SC",                  align=_R)
    _set("O6", ":",                          align=_R)

    # Row 7 — Department
    _set("B7", "Department",               align=_L)
    _set("E7", ":")
    _set("F7", "-",                         align=_L)

    # Row 8 — Company
    _set("B8", "Company",                  align=_L)
    _set("E8", ":")
    _set("F8", pam.get("pt", ""),          align=_L)

    # Row 10 — Business Unit header
    _set("B10", "Bussiness Unit ",         align=_L)

    # Row 11 — BU checkboxes (unmerged single cells)
    ws["E11"].border = _box                         # empty checkbox (Upstream)
    _set("G11", "  Upstream",              align=_L)
    ws["I11"].border = Border(right=_thin)
    ws["J11"].border = Border(top=_thin, bottom=_thin, right=_thin)
    _set("K11", "  Downstream",            align=_L)
    _set("N11", "V",                        align=_C, border=_box)
    _set("O11", "  Corporate",             align=_L)

    # Row 13 — Type of Request header
    _set("B13", "Type of Request ",        align=_L)

    # Row 14 — Downpayment (unchecked)
    ws["E14"].border = _box                         # empty checkbox
    _set("G14", "  Downpayment to vendor", align=_L)

    # Row 15 — Invoice Payment (checked)
    _set("E15", "V",                        align=_C, border=_box)
    _set("G15", "  Invoice Payment – Non PO Invoice", align=_L)

    # Row 16 — Employee Advance (unchecked)
    ws["E16"].border = _box                         # empty checkbox
    _set("G16", "  Employee Advance/ Reimbursement (Fund Transfer)", align=_L)

    # Row 18 — Invoice Information header
    ws.merge_cells("B18:Q18")
    _set("B18", "Invoice Information",     font=_bold(), align=_C)

    # Row 19 — Vendor Name
    ws.merge_cells("I19:Q19")
    _set("G19", "Vendor Name",             align=_R)
    _set("H19", ":",                        align=_R)
    _set("I19", "Terlampir",               align=_L)

    # Row 20 — Invoice/Memo Number
    ws.merge_cells("I20:Q20")
    _set("G20", "Invoice/ Memorandum Number", align=_R)
    _set("H20", ":",                        align=_R)
    _set("I20", "-",                        align=_L)

    # Row 21 — Invoice Amount
    ws.merge_cells("I21:L21")
    _set("G21", "Invoice Amount",          align=_R)
    _set("H21", ":",                        align=_R)
    _set("I21", pam.get("total_amount", 0), align=_C)
    ws["I21"].number_format = '_("Rp"* #,##0_);_("Rp"* \\(#,##0\\);_("Rp"* "-"_);_(@_)'

    # Row 22 — Expected Due Date
    ws.merge_cells("I22:O22")
    _set("G22", "Expected Due Date",       align=_R)
    _set("H22", ":",                        align=_R)
    _set("I22", due_date,                  align=_L)
    ws["I22"].number_format = '[$-421]dd\\ mmmm\\ yyyy;@'

    # Row 24 — Vendor Bank Account header
    ws.merge_cells("B24:Q24")
    _set("B24", "Vendor Bank Account Details", font=_bold(), align=_C)

    # Rows 25-27 — Bank details
    ws.merge_cells("I25:Q25")
    _set("D25", "Bank Account Name ",      align=_L)
    _set("H25", ":")
    _set("I25", "Terlampir",               align=_L)

    ws.merge_cells("I26:Q26")
    _set("D26", "Bank Name ",              align=_L)
    _set("H26", ":")
    _set("I26", "Terlampir",               align=_L)

    ws.merge_cells("I27:Q27")
    _set("D27", "Bank Account Number",     align=_L)
    _set("H27", ":")
    _set("I27", "Terlampir",               align=_L)

    # Row 28 — separator
    ws.merge_cells("I28:Q28")

    # Rows 29-34 — Request by signature (unmerged, cell-by-cell borders)
    _draw_box(29, 2, 29, 6)
    _set("C29", "Request by",              font=_bold(), align=_CN)
    _draw_box(30, 2, 33, 6)
    _draw_box(34, 2, 34, 6)
    _set("C34", pam.get("requestors_name", ""), align=_CN)

    # Rows 36-42 — Approved by signature (unmerged, cell-by-cell borders)
    _draw_box(36, 2, 36, 11)
    _set("E36", "Approved by",             font=_bold(), align=_CN)
    for _r in range(36, 43):
        ws.cell(_r, 12).border = Border(left=_thin)   # L36:L42 — left edge only
    _draw_box(37, 2, 41, 6)
    _draw_box(37, 7, 41, 11)
    _draw_box(42, 2, 42, 6)
    _draw_box(42, 7, 42, 11)
    _set("C42", approved_by_1,             align=_CN)
    _set("H42", approved_by_2,             align=_CN)

    # Row 43 — Checked by (QA)
    _set("B43", "Checked by (QA)",         font=_bold())

    # Rows 44-48 — QA signature space (unmerged)
    _draw_box(44, 2, 48, 6)

    # ── Sheet 2: Rangkuman PAM (Book8 standard format) ───────────────────────
    ws2 = wb.create_sheet("Rangkuman PAM")

    # Column widths matching Book8
    for _c2, _w2 in [("A",2.77),("B",10.30),("C",1.15),("D",37.84),
                     ("E",37.0), ("F",23.15),("G",1.0), ("H",18.46),("I",2.84)]:
        ws2.column_dimensions[_c2].width = _w2

    # Styles
    _hf2 = PatternFill("solid", fgColor="D9D9D9")  # header fill (theme 0 tint -0.15)
    _t2  = Side(style="thin")
    _k2  = Side(style="thick")

    def _b2(l=False, r=False, t=False, b=False):
        return Border(left=_t2 if l else None, right=_t2 if r else None,
                      top=_t2 if t else None,  bottom=_t2 if b else None)

    def _s2(row, col, val=None, bold=False, sz=11,
            ha=None, va=None, fill=None, bdr=None, fmt=None):
        c2 = ws2.cell(row, col)
        if val is not None:
            c2.value = val
        c2.font = Font(bold=bold, size=sz)
        if ha or va:
            c2.alignment = Alignment(horizontal=ha or "general",
                                     vertical=va or "bottom")
        if fill:
            c2.fill = fill
        if bdr:
            c2.border = bdr
        if fmt:
            c2.number_format = fmt
        return c2

    # Row heights for info rows
    for _ri2 in (2, 3, 4, 5):
        ws2.row_dimensions[_ri2].height = 15.9

    # Rows 2-3: PAM info header
    _s2(2, 2, "NO. ",         bold=True, sz=12, va="top")
    _s2(2, 3, ":",             bold=True, sz=12)
    _s2(2, 4, pam_no,          sz=12,    va="top")
    _s2(2, 6, "COST CENTER ", bold=True, sz=12)
    _s2(2, 7, ":",             bold=True, sz=12)
    _s2(2, 8, pam.get("cost_center", ""), bold=True, sz=12)

    _s2(3, 2, "TANGGAL",      bold=True, sz=12)
    _s2(3, 3, ":",             bold=True, sz=12)
    try:
        _d3v = _dt.strptime(pam.get("pam_date", "")[:10], "%Y-%m-%d")
        _s2(3, 4, _d3v, bold=True, sz=12, ha="left",
            fmt='[$-421]dd\\ mmmm\\ yyyy;@')
    except Exception:
        _s2(3, 4, pam.get("pam_date", ""), bold=True, sz=12, ha="left")
    _s2(3, 6, "GL ACCOUNT     ", bold=True, sz=12)
    _s2(3, 7, ":",               bold=True, sz=12)
    _s2(3, 8, pam.get("gl_account", ""), bold=True, sz=12, ha="left")

    # Split payments: universities (Section 2) vs individuals (Section 1)
    _UNI_KW = ("universitas", "institut ", "akademi", "politeknik",
               "sekolah tinggi", "stie", "stmik", "kelolaan")
    _sec1, _sec2 = [], []
    for _pb2 in payments:
        _check = ((_pb2.get("namarek") or "") + " " +
                  (_pb2.get("nama") or "")).lower()
        if any(_kw2 in _check for _kw2 in _UNI_KW):
            _sec2.append(_pb2)
        else:
            _sec1.append(_pb2)

    def _write_hdr2(r0, sz=11):
        """Write 2-row table header at rows r0 and r0+1."""
        ws2.merge_cells(f"B{r0}:B{r0+1}")
        ch = ws2.cell(r0, 2, "NO")
        ch.font = Font(bold=True, size=sz)
        ch.fill = _hf2
        ch.alignment = Alignment(horizontal="center", vertical="center")
        ch.border = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(r0+1, 2).border = Border(left=_t2, right=_t2, bottom=_t2)
        # C spacer
        ws2.cell(r0,   3).fill = _hf2
        ws2.cell(r0,   3).border = _b2(l=True, t=True)
        ws2.cell(r0+1, 3).fill = _hf2
        ws2.cell(r0+1, 3).border = _b2(l=True, b=True)
        # D: "NAMA " / "REKENING"
        for _rh, _th in [(r0, "NAMA "), (r0+1, "REKENING")]:
            ch = ws2.cell(_rh, 4, _th)
            ch.font = Font(bold=True, size=sz)
            ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center", vertical="center")
            ch.border = Border(right=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))
        # E: BANK
        for _rh, _th in [(r0, "BANK"), (r0+1, None)]:
            ch = ws2.cell(_rh, 5, _th)
            ch.font = Font(bold=True, size=sz)
            ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center")
            ch.border = Border(left=_t2, right=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))
        # F: "NO" / "REKENING"
        for _rh, _th in [(r0, "NO"), (r0+1, "REKENING")]:
            ch = ws2.cell(_rh, 6, _th)
            ch.font = Font(bold=True, size=sz)
            ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center")
            ch.border = Border(left=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))
        # G spacer
        ws2.cell(r0,   7).fill = _hf2
        ws2.cell(r0,   7).border = _b2(r=True, t=True)
        ws2.cell(r0+1, 7).fill = _hf2
        ws2.cell(r0+1, 7).border = _b2(r=True, b=True)
        # H: "TOTAL" / "PEMBAYARAN"
        for _rh, _th in [(r0, "TOTAL"), (r0+1, "PEMBAYARAN")]:
            ch = ws2.cell(_rh, 8, _th)
            ch.font = Font(bold=True, size=sz)
            ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center")
            ch.border = Border(left=_t2, right=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))

    def _write_data2(row, seq, namarek, bank, norek, amt, sz=12):
        ws2.row_dimensions[row].height = 29.25
        fn2 = Font(size=sz)
        ws2.cell(row, 2, str(seq)).font = fn2
        ws2.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 2).border = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 3).border = _b2(l=True, t=True, b=True)
        ws2.cell(row, 4, namarek).font = fn2
        ws2.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center")
        ws2.cell(row, 4).border = Border(right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 5, bank).font = fn2
        ws2.cell(row, 5).alignment = Alignment(horizontal="left", vertical="center")
        ws2.cell(row, 5).border = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 6, norek).font = fn2
        ws2.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 6).border = _b2(l=True, t=True, b=True)
        ws2.cell(row, 7).border = _b2(r=True, t=True, b=True)
        ws2.cell(row, 8, amt).font = fn2
        ws2.cell(row, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 8).border = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 8).number_format = '#,##0'

    def _write_tot2(row, total_val, label="Total", sz=12, thick_top=False):
        ws2.merge_cells(f"B{row}:F{row}")
        fn2 = Font(bold=True, size=sz)
        ws2.cell(row, 2, label).font = fn2
        ws2.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 2).border = Border(left=_t2, top=_t2, bottom=_t2)
        for _col2 in range(3, 7):  # C–F: top+bottom, no sides (within merge)
            ws2.cell(row, _col2).border = Border(top=_t2, bottom=_t2)
        ws2.cell(row, 7).border = Border(
            right=_t2, bottom=_t2,
            top=(_k2 if thick_top else None))
        ws2.cell(row, 8, total_val).font = fn2
        ws2.cell(row, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 8).border = Border(
            left=_t2, right=_t2, bottom=_t2,
            top=(_k2 if thick_top else _t2))
        ws2.cell(row, 8).number_format = '#,##0'

    # ── Section 1: individual / student payments ──────────────────────────────
    _cur2 = 6
    _write_hdr2(_cur2, sz=11)
    _cur2 += 2
    _tot1 = 0.0
    for _i2, _pb2 in enumerate(_sec1, 1):
        _nm2  = _pb2.get("namarek") or _pb2.get("nama") or _pb2.get("siswa_code", "")
        _bk2  = _pb2.get("bank") or ""
        _nr2  = _pb2.get("norek") or ""
        _am2  = float(_pb2.get("amount", 0))
        _write_data2(_cur2, _i2, _nm2, _bk2, _nr2, _am2, sz=12)
        _tot1 += _am2
        _cur2 += 1
    _write_tot2(_cur2, _tot1, "Total", sz=12, thick_top=True)
    _cur2 += 1

    _grand2 = _tot1

    # ── Section 2: university payments (if any) ───────────────────────────────
    if _sec2:
        _cur2 += 1  # spacer row
        ws2.cell(_cur2, 2, "Dibayarkan ke Universitas").font = Font(bold=True, size=11)
        _cur2 += 1
        _write_hdr2(_cur2, sz=10)
        _cur2 += 2
        _tot2 = 0.0
        for _i2, _pb2 in enumerate(_sec2, 1):
            _nm2 = _pb2.get("namarek") or _pb2.get("nama") or _pb2.get("siswa_code", "")
            _bk2 = _pb2.get("bank") or ""
            _nr2 = _pb2.get("norek") or ""
            _am2 = float(_pb2.get("amount", 0))
            _write_data2(_cur2, _i2, _nm2, _bk2, _nr2, _am2, sz=10)
            _tot2 += _am2
            _cur2 += 1
        _write_tot2(_cur2, _tot2, "Total", sz=10, thick_top=False)
        _cur2 += 1
        _grand2 += _tot2

        # Spacer + Grand Total
        _cur2 += 1
        ws2.merge_cells(f"B{_cur2}:F{_cur2}")
        _fg = Font(bold=True, size=10)
        ws2.cell(_cur2, 2, "Grand Total").font = _fg
        ws2.cell(_cur2, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(_cur2, 2).border = Border(left=_t2, top=_t2, bottom=_t2)
        for _col2 in range(3, 7):
            ws2.cell(_cur2, _col2).border = Border(top=_t2, bottom=_t2)
        ws2.cell(_cur2, 7).border = _b2(r=True, t=True, b=True)
        ws2.cell(_cur2, 8, _grand2).font = _fg
        ws2.cell(_cur2, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(_cur2, 8).border = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(_cur2, 8).number_format = '#,##0'

    # ── Sheet 3: Detail PAM (Book9 format) ──────────────────────────────────
    ws3 = wb.create_sheet("Detail PAM")
    _build_detail_sheet(ws3, pam, get_pam_payments_detail(pam_no, company_id))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_pam_excel_custom(data: dict, payments: list) -> bytes:
    """Generate PAM Excel (Book6 format) from a user-supplied data dict."""
    from datetime import datetime as _dt

    approved_by_1 = data.get("approved_by_1") or config.PAM_APPROVED_BY_1
    approved_by_2 = data.get("approved_by_2") or config.PAM_APPROVED_BY_2
    pam_no = data.get("pam_no", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAM NEW"

    for col, w in [("A",3.0),("B",11.15),("C",1.84),("D",7.53),("E",6.43),
                   ("F",2.84),("G",9.0),("H",2.0),("J",4.53),("K",5.69),
                   ("L",4.69),("N",5.30),("O",1.15),("P",7.84)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[9].height  = 8.25
    ws.row_dimensions[11].height = 16.5
    ws.row_dimensions[12].height = 11.25
    ws.row_dimensions[19].height = 17.25
    ws.row_dimensions[25].height = 15.0

    _thin = Side(style="thin")
    _box  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _CN   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _L    = Alignment(horizontal="left",   vertical="center")
    _R    = Alignment(horizontal="right",  vertical="center")

    def _bold(sz=11): return Font(bold=True,  size=sz)

    def _set(coord, val, font=None, align=None, border=None):
        c = ws[coord]
        c.value = val
        if font:   c.font      = font
        if align:  c.alignment = align
        if border: c.border    = border

    def _draw_box(r1, c1, r2, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                prev = cell.border
                cell.border = Border(
                    top    = _thin if r == r1 else prev.top,
                    bottom = _thin if r == r2 else prev.bottom,
                    left   = _thin if c == c1 else prev.left,
                    right  = _thin if c == c2 else prev.right,
                )

    def _dt_val(s):
        try:    return _dt.strptime(s[:10], "%Y-%m-%d")
        except: return s or ""

    pam_date = _dt_val(data.get("pam_date", ""))
    due_date = _dt_val(data.get("due_date", ""))

    ws.merge_cells("A1:Q1")
    _set("A1", "PAYMENT APPROVAL MEMO", font=_bold(11), align=_C)
    ws.merge_cells("B2:Q2")

    ws.merge_cells("L4:N4")
    _set("B4", "PAM No.",                              align=_L)
    _set("E4", ":")
    _set("F4", pam_no,                                 align=_L)
    _set("L4", "Cost Center",                          align=_R)
    _set("O4", ":",                                    align=_R)
    _set("P4", data.get("cost_center", ""),            align=_L)

    ws.merge_cells("F5:J5")
    ws.merge_cells("L5:N5")
    ws.merge_cells("P5:Q5")
    _set("B5", "Date",                                 align=_L)
    _set("E5", ":")
    _set("F5", pam_date,                               align=_L)
    ws["F5"].number_format = '[$-421]dd\\ mmmm\\ yyyy;@'
    _set("L5", "GL Account",                          align=_R)
    _set("O5", ":",                                    align=_R)
    _set("P5", data.get("gl_account", ""),            align=_L)

    ws.merge_cells("L6:N6")
    _set("B6", "Requestor's Name   ",                  align=_L)
    _set("E6", ":")
    _set("F6", data.get("requestors_name", ""),       align=_L)
    _set("L6", "SO / SC",                              align=_R)
    _set("O6", ":",                                    align=_R)
    _set("P6", data.get("so_sc", ""),                 align=_L)

    _set("B7", "Department",                           align=_L)
    _set("E7", ":")
    _set("F7", data.get("department", "-"),           align=_L)

    _set("B8", "Company",                              align=_L)
    _set("E8", ":")
    _set("F8", data.get("pt", ""),                    align=_L)

    _set("B10", "Bussiness Unit ",                     align=_L)

    ws["E11"].border = _box
    if data.get("bu_upstream"):
        _set("E11", "V", align=_CN, border=_box)
    _set("G11", "  Upstream",                          align=_L)
    ws["I11"].border = Border(right=_thin)
    ws["J11"].border = Border(top=_thin, bottom=_thin, right=_thin)
    if data.get("bu_downstream"):
        ws["J11"].value     = "V"
        ws["J11"].alignment = _CN
    _set("K11", "  Downstream",                        align=_L)
    _set("N11", "V" if data.get("bu_corporate") else "", align=_C, border=_box)
    _set("O11", "  Corporate",                         align=_L)

    _set("B13", "Type of Request ",                    align=_L)

    ws["E14"].border = _box
    if data.get("type_downpayment"):
        _set("E14", "V", align=_CN, border=_box)
    _set("G14", "  Downpayment to vendor",             align=_L)

    ws["E15"].border = _box
    if data.get("type_invoice"):
        _set("E15", "V", align=_CN, border=_box)
    _set("G15", "  Invoice Payment – Non PO Invoice", align=_L)

    ws["E16"].border = _box
    if data.get("type_advance"):
        _set("E16", "V", align=_CN, border=_box)
    _set("G16", "  Employee Advance/ Reimbursement (Fund Transfer)", align=_L)

    ws.merge_cells("B18:Q18")
    _set("B18", "Invoice Information",                 font=_bold(), align=_C)

    ws.merge_cells("I19:Q19")
    _set("G19", "Vendor Name",                         align=_R)
    _set("H19", ":",                                    align=_R)
    _set("I19", data.get("vendor_name", "Terlampir"),  align=_L)

    ws.merge_cells("I20:Q20")
    _set("G20", "Invoice/ Memorandum Number",          align=_R)
    _set("H20", ":",                                    align=_R)
    _set("I20", data.get("invoice_memo_no", "-"),      align=_L)

    ws.merge_cells("I21:L21")
    _set("G21", "Invoice Amount",                      align=_R)
    _set("H21", ":",                                    align=_R)
    _set("I21", float(data.get("total_amount", 0) or 0), align=_C)
    ws["I21"].number_format = '_("Rp"* #,##0_);_("Rp"* \\(#,##0\\);_("Rp"* "-"_);_(@_)'

    ws.merge_cells("I22:O22")
    _set("G22", "Expected Due Date",                   align=_R)
    _set("H22", ":",                                    align=_R)
    _set("I22", due_date,                              align=_L)
    ws["I22"].number_format = '[$-421]dd\\ mmmm\\ yyyy;@'

    ws.merge_cells("B24:Q24")
    _set("B24", "Vendor Bank Account Details",         font=_bold(), align=_C)

    ws.merge_cells("I25:Q25")
    _set("D25", "Bank Account Name ",                  align=_L)
    _set("H25", ":")
    _set("I25", data.get("bank_account_name", "Terlampir"), align=_L)

    ws.merge_cells("I26:Q26")
    _set("D26", "Bank Name ",                          align=_L)
    _set("H26", ":")
    _set("I26", data.get("bank_name", "Terlampir"),   align=_L)

    ws.merge_cells("I27:Q27")
    _set("D27", "Bank Account Number",                 align=_L)
    _set("H27", ":")
    _set("I27", data.get("bank_account_no", "Terlampir"), align=_L)

    ws.merge_cells("I28:Q28")

    _draw_box(29, 2, 29, 6)
    _set("C29", "Request by",                          font=_bold(), align=_CN)
    _draw_box(30, 2, 33, 6)
    _draw_box(34, 2, 34, 6)
    _set("C34", data.get("requestors_name", ""),       align=_CN)

    _draw_box(36, 2, 36, 11)
    _set("E36", "Approved by",                         font=_bold(), align=_CN)
    for _r in range(36, 43):
        ws.cell(_r, 12).border = Border(left=_thin)
    _draw_box(37, 2, 41, 6)
    _draw_box(37, 7, 41, 11)
    _draw_box(42, 2, 42, 6)
    _draw_box(42, 7, 42, 11)
    _set("C42", approved_by_1,                         align=_CN)
    _set("H42", approved_by_2,                         align=_CN)

    _set("B43", "Checked by (QA)",                     font=_bold())
    _draw_box(44, 2, 48, 6)

    # ── Sheet 2: Rangkuman PAM (Book8 format) ─────────────────────────────────
    ws2 = wb.create_sheet("Rangkuman PAM")

    # Column widths matching Book8
    for _c2, _w2 in [("A", 2.77), ("B", 10.30), ("C", 1.15), ("D", 37.84),
                     ("E", 37.0),  ("F", 23.15), ("G", 1.0),  ("H", 18.46), ("I", 2.84)]:
        ws2.column_dimensions[_c2].width = _w2

    _hf2 = PatternFill("solid", fgColor="D9D9D9")
    _t2  = Side(style="thin")
    _k2  = Side(style="thick")

    def _b2(l=False, r=False, t=False, b=False):
        return Border(left=_t2 if l else None, right=_t2 if r else None,
                      top=_t2 if t else None,  bottom=_t2 if b else None)

    def _s2(row, col, val=None, bold=False, sz=11,
            ha=None, va=None, fill=None, bdr=None, fmt=None):
        c2 = ws2.cell(row, col)
        if val is not None:
            c2.value = val
        c2.font = Font(bold=bold, size=sz)
        if ha or va:
            c2.alignment = Alignment(horizontal=ha or "general",
                                     vertical=va or "bottom")
        if fill:  c2.fill   = fill
        if bdr:   c2.border = bdr
        if fmt:   c2.number_format = fmt
        return c2

    # Row heights for info rows
    for _ri2 in (2, 3, 4, 5):
        ws2.row_dimensions[_ri2].height = 15.9

    # Rows 2-3: PAM info header
    _s2(2, 2, "NO. ",          bold=True, sz=12, va="top")
    _s2(2, 3, ":",              bold=True, sz=12)
    _s2(2, 4, pam_no,           sz=12,    va="top")
    _s2(2, 6, "COST CENTER ",  bold=True, sz=12)
    _s2(2, 7, ":",              bold=True, sz=12)
    _s2(2, 8, data.get("cost_center", ""), bold=True, sz=12)

    _s2(3, 2, "TANGGAL",       bold=True, sz=12)
    _s2(3, 3, ":",              bold=True, sz=12)
    try:
        _d3v = _dt.strptime(data.get("pam_date", "")[:10], "%Y-%m-%d")
        _s2(3, 4, _d3v, bold=True, sz=12, ha="left",
            fmt='[$-421]dd\\ mmmm\\ yyyy;@')
    except Exception:
        _s2(3, 4, data.get("pam_date", ""), bold=True, sz=12, ha="left")
    _s2(3, 6, "GL ACCOUNT     ", bold=True, sz=12)
    _s2(3, 7, ":",               bold=True, sz=12)
    _s2(3, 8, data.get("gl_account", ""), bold=True, sz=12, ha="left")

    # Split payments: universities (Section 2) vs individuals (Section 1)
    _UNI_KW = ("universitas", "institut ", "akademi", "politeknik",
               "sekolah tinggi", "stie", "stmik", "kelolaan")
    grouped_pay = _group_payments_by_siswa(payments)
    _sec1, _sec2 = [], []
    for _pb2 in grouped_pay:
        if any(_kw2 in _pb2["nama"].lower() for _kw2 in _UNI_KW):
            _sec2.append(_pb2)
        else:
            _sec1.append(_pb2)

    def _write_hdr2(r0, sz=11):
        ws2.merge_cells(f"B{r0}:B{r0+1}")
        ch = ws2.cell(r0, 2, "NO")
        ch.font = Font(bold=True, size=sz)
        ch.fill = _hf2
        ch.alignment = Alignment(horizontal="center", vertical="center")
        ch.border = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(r0+1, 2).border = Border(left=_t2, right=_t2, bottom=_t2)
        # C spacer
        ws2.cell(r0,   3).fill = _hf2;  ws2.cell(r0,   3).border = _b2(l=True, t=True)
        ws2.cell(r0+1, 3).fill = _hf2;  ws2.cell(r0+1, 3).border = _b2(l=True, b=True)
        # D: "NAMA " / "REKENING"
        for _rh, _th in [(r0, "NAMA "), (r0+1, "REKENING")]:
            ch = ws2.cell(_rh, 4, _th)
            ch.font = Font(bold=True, size=sz); ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center", vertical="center")
            ch.border = Border(right=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))
        # E: "BANK"
        for _rh, _th in [(r0, "BANK"), (r0+1, None)]:
            ch = ws2.cell(_rh, 5, _th)
            ch.font = Font(bold=True, size=sz); ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center")
            ch.border = Border(left=_t2, right=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))
        # F: "NO" / "REKENING"
        for _rh, _th in [(r0, "NO"), (r0+1, "REKENING")]:
            ch = ws2.cell(_rh, 6, _th)
            ch.font = Font(bold=True, size=sz); ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center")
            ch.border = Border(left=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))
        # G spacer
        ws2.cell(r0,   7).fill = _hf2;  ws2.cell(r0,   7).border = _b2(r=True, t=True)
        ws2.cell(r0+1, 7).fill = _hf2;  ws2.cell(r0+1, 7).border = _b2(r=True, b=True)
        # H: "TOTAL" / "PEMBAYARAN"
        for _rh, _th in [(r0, "TOTAL"), (r0+1, "PEMBAYARAN")]:
            ch = ws2.cell(_rh, 8, _th)
            ch.font = Font(bold=True, size=sz); ch.fill = _hf2
            ch.alignment = Alignment(horizontal="center")
            ch.border = Border(left=_t2, right=_t2,
                               top=(_t2 if _rh == r0   else None),
                               bottom=(_t2 if _rh == r0+1 else None))

    def _write_data2(row, seq, nama, bank, norek, amt, sz=12):
        ws2.row_dimensions[row].height = 29.25
        fn2 = Font(size=sz)
        ws2.cell(row, 2, str(seq)).font = fn2
        ws2.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 2).border    = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 3).border    = _b2(l=True, t=True, b=True)
        ws2.cell(row, 4, nama).font      = fn2
        ws2.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center")
        ws2.cell(row, 4).border    = Border(right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 5, bank).font      = fn2
        ws2.cell(row, 5).alignment = Alignment(horizontal="left", vertical="center")
        ws2.cell(row, 5).border    = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 6, norek).font     = fn2
        ws2.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 6).border    = _b2(l=True, t=True, b=True)
        ws2.cell(row, 7).border    = _b2(r=True, t=True, b=True)
        ws2.cell(row, 8, amt).font       = fn2
        ws2.cell(row, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 8).border    = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(row, 8).number_format = '#,##0'

    def _write_tot2(row, total_val, label="Total", sz=12, thick_top=False):
        ws2.merge_cells(f"B{row}:F{row}")
        fn2 = Font(bold=True, size=sz)
        ws2.cell(row, 2, label).font      = fn2
        ws2.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 2).border    = Border(left=_t2, top=_t2, bottom=_t2)
        for _col2 in range(3, 7):
            ws2.cell(row, _col2).border = Border(top=_t2, bottom=_t2)
        ws2.cell(row, 7).border = Border(right=_t2, bottom=_t2,
                                          top=(_k2 if thick_top else None))
        ws2.cell(row, 8, total_val).font      = fn2
        ws2.cell(row, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row, 8).border    = Border(left=_t2, right=_t2, bottom=_t2,
                                             top=(_k2 if thick_top else _t2))
        ws2.cell(row, 8).number_format = '#,##0'

    # ── Section 1: individual / student payments ──────────────────────────────
    _cur2 = 6
    _write_hdr2(_cur2, sz=11)
    _cur2 += 2
    _tot1 = 0.0
    for _i2, _pb2 in enumerate(_sec1, 1):
        _write_data2(_cur2, _i2, _pb2["nama"], _pb2["bank"], _pb2["norek"],
                     _pb2["amount"], sz=12)
        _tot1 += _pb2["amount"]
        _cur2 += 1
    _write_tot2(_cur2, _tot1, "Total", sz=12, thick_top=True)
    _cur2 += 1

    _grand2 = _tot1

    # ── Section 2: university payments (if any) ───────────────────────────────
    if _sec2:
        _cur2 += 1
        ws2.cell(_cur2, 2, "Dibayarkan ke Universitas").font = Font(bold=True, size=11)
        _cur2 += 1
        _write_hdr2(_cur2, sz=10)
        _cur2 += 2
        _tot2 = 0.0
        for _i2, _pb2 in enumerate(_sec2, 1):
            _write_data2(_cur2, _i2, _pb2["nama"], _pb2["bank"], _pb2["norek"],
                         _pb2["amount"], sz=10)
            _tot2 += _pb2["amount"]
            _cur2 += 1
        _write_tot2(_cur2, _tot2, "Total", sz=10, thick_top=False)
        _cur2 += 1
        _grand2 += _tot2

        # Grand Total row
        _cur2 += 1
        ws2.merge_cells(f"B{_cur2}:F{_cur2}")
        _fg = Font(bold=True, size=10)
        ws2.cell(_cur2, 2, "Grand Total").font      = _fg
        ws2.cell(_cur2, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(_cur2, 2).border    = Border(left=_t2, top=_t2, bottom=_t2)
        for _col2 in range(3, 7):
            ws2.cell(_cur2, _col2).border = Border(top=_t2, bottom=_t2)
        ws2.cell(_cur2, 7).border = _b2(r=True, t=True, b=True)
        ws2.cell(_cur2, 8, _grand2).font      = _fg
        ws2.cell(_cur2, 8).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(_cur2, 8).border    = Border(left=_t2, right=_t2, top=_t2, bottom=_t2)
        ws2.cell(_cur2, 8).number_format = '#,##0'

    # ── Sheet 3: Detail PAM (Book9 format) ──────────────────────────────────
    company_id = int(data.get("company_id") or 0)
    if pam_no and company_id:
        ws3 = wb.create_sheet("Detail PAM")
        _build_detail_sheet(ws3, data, get_pam_payments_detail(pam_no, company_id))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Generic xlsx builder ────────────────────────────────────────────────────
def _make_xlsx(sheet_title: str, col_headers: list, fields: list,
               rows: list, col_widths: list) -> bytes:
    """Buat file xlsx dengan header navy, freeze row 1, auto-filter."""
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin     = Side(style="thin", color="D1D5DB")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 32

    for ri, r in enumerate(rows, 2):
        for ci, f in enumerate(fields, 1):
            val = r.get(f)
            if val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(size=9)
            cell.border    = border
            cell.alignment = Alignment(vertical="top")

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tab export functions ────────────────────────────────────────────────────
def export_open_pam_excel(company_id: int) -> bytes:
    from modules.payment_memo.service import get_draft_payments
    rows    = get_draft_payments(company_id)
    headers = ["Code", "Nama Siswa", "Kategori 1", "Kategori 2", "Tanggal",
               "PAM No", "Perusahaan", "Amount (Rp)", "Status"]
    fields  = ["siswa_code", "nama", "cat1", "cat2", "tanggal",
               "pam", "perusahaan", "amount", "status"]
    widths  = [14, 24, 18, 18, 12, 22, 22, 16, 12]
    return _make_xlsx("Open PAM", headers, fields, rows, widths)
