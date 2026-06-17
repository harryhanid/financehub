"""
Import script: sync perubahan dari Excel → SQLite (PAM AGRI + Open PAM).

Usage:
    python tmp_import_excel.py           # dry-run — lihat diff, tidak ada yang berubah

File Excel dikonfigurasi di konstanta PAM_AGRI_FILE dan OPEN_PAM_FILE di bawah.
"""
import sys
import os
import math
import shutil
import argparse
from datetime import datetime, date

# ── Path setup ──────────────────────────────────────────────────────────────────
_ROOT = os.path.dirname(os.path.abspath(__file__))
_APP  = os.path.join(_ROOT, "app")
sys.path.insert(0, _APP)

from database import get_conn  # noqa: E402

# ── Config ──────────────────────────────────────────────────────────────────────
PAM_AGRI_FILE = r"C:\Users\25010160\Downloads\PAM_AGRI_20260617_1111.xlsx"
OPEN_PAM_FILE = r"C:\Users\25010160\Downloads\Open_PAM_20260617_1111.xlsx"
DB_PATH       = os.path.join(_APP, "finance_hub.db")
COMPANY_ID    = 2   # ETF company_id di DB


# ── Normalize helpers ───────────────────────────────────────────────────────────

def normalize_date(val) -> str | None:
    """Return YYYY-MM-DD string atau None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s[:10] if len(s) >= 10 else None


def normalize_amount(val) -> float:
    """Return float, 0.0 jika tidak valid atau NaN."""
    if val is None:
        return 0.0
    try:
        result = float(val)
        return 0.0 if math.isnan(result) else result
    except (TypeError, ValueError):
        return 0.0


# ── Match logic: pam_records ────────────────────────────────────────────────────

def match_pam_agri(excel_rows: list[dict], db_rows: list[dict]) -> dict:
    """
    Bandingkan Excel rows vs DB rows untuk pam_records.

    Returns:
        {
            "updates": [(db_row, excel_row)],   # same PAM No, field berubah
            "renames": [(db_row, excel_row)],   # PAM No beda, auto-detect via date+total
            "deletes": [db_row],                # di DB tapi tidak di Excel
            "skips":   [excel_row],             # di Excel tapi tidak bisa di-match
        }
    """
    db_by_pam    = {r["pam_no"]: r for r in db_rows}
    matched_ids  = set()
    updates, tentative_skips = [], []

    for ex in excel_rows:
        pno = (ex.get("PAM No") or "").strip()
        if not pno:
            continue
        if pno in db_by_pam:
            db = db_by_pam[pno]
            matched_ids.add(db["id"])
            ex_status = (ex.get("Status") or "").strip()
            ex_tgl    = normalize_date(ex.get("Tgl Paid"))
            if ex_status != (db.get("status") or "").strip() or \
               ex_tgl != db.get("tanggal_bayar"):
                updates.append((db, ex))
        else:
            tentative_skips.append(ex)

    # Auto-detect renames: skip row yang tidak match PAM No,
    # cek apakah ada DB row "orphan" (tidak di Excel) dengan pam_date + total_amount sama.
    unmatched_db = [r for r in db_rows if r["id"] not in matched_ids]
    renames, real_skips = [], []

    for ex in tentative_skips:
        ex_date = normalize_date(ex.get("PAM Date"))
        ex_amt  = normalize_amount(ex.get("Total (Rp)"))
        candidates = [
            r for r in unmatched_db
            if normalize_date(r.get("pam_date")) == ex_date
            and abs(normalize_amount(r.get("total_amount")) - ex_amt) < 0.01
        ]
        if len(candidates) == 1:
            db = candidates[0]
            matched_ids.add(db["id"])
            unmatched_db.remove(db)
            renames.append((db, ex))
        else:
            real_skips.append(ex)

    deletes = [r for r in db_rows if r["id"] not in matched_ids]
    return {"updates": updates, "renames": renames,
            "deletes": deletes, "skips": real_skips}


# ── Match logic: payment_beasiswa ───────────────────────────────────────────────

def _pb_key(code, cat1, cat2, tanggal, amount) -> tuple:
    """Composite key untuk matching payment_beasiswa."""
    return (
        (code    or "").strip(),
        (cat1    or "").strip(),
        (cat2    or "").strip(),
        normalize_date(tanggal) or "",
        round(normalize_amount(amount), 2),
    )


def match_open_pam(excel_rows: list[dict], db_rows: list[dict]) -> dict:
    """
    Bandingkan Excel rows vs DB rows untuk payment_beasiswa.

    Returns:
        {
            "updates": [(db_row, excel_row)],   # composite key cocok, field berubah
            "deletes": [db_row],                # di DB tapi tidak di Excel
            "skips":   [excel_row],             # di Excel tapi tidak ada match di DB
        }
    """
    db_by_key = {}
    for r in db_rows:
        k = _pb_key(r["siswa_code"], r["cat1"], r["cat2"], r["tanggal"], r["amount"])
        if k in db_by_key:
            raise ValueError(
                f"Duplicate composite key in DB: siswa_code={r['siswa_code']}, "
                f"tanggal={r['tanggal']}, amount={r['amount']}"
            )
        db_by_key[k] = r

    updates, skips = [], []
    matched_ids    = set()

    for ex in excel_rows:
        k = _pb_key(
            ex.get("Code"), ex.get("Kategori 1"), ex.get("Kategori 2"),
            ex.get("Tanggal"), ex.get("Amount (Rp)")
        )
        if k in db_by_key:
            db = db_by_key[k]
            matched_ids.add(db["id"])
            ex_pam = (ex.get("PAM No")     or "").strip()
            ex_per = (ex.get("Perusahaan") or "").strip()
            ex_st  = (ex.get("Status")     or "").strip()
            if ex_pam != (db.get("pam")        or "").strip() or \
               ex_per != (db.get("perusahaan") or "").strip() or \
               ex_st  != (db.get("status")     or "").strip():
                updates.append((db, ex))
        else:
            skips.append(ex)

    deletes = [r for r in db_rows if r["id"] not in matched_ids]
    return {"updates": updates, "deletes": deletes, "skips": skips}


# ── Load functions ──────────────────────────────────────────────────────────────

def _load_excel(filepath: str) -> list[dict]:
    """Baca sheet pertama Excel, return list of dicts. Skip fully blank rows."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def load_pam_agri_excel(filepath: str) -> list[dict]:
    """Baca PAM AGRI Excel file."""
    return _load_excel(filepath)


def load_open_pam_excel(filepath: str) -> list[dict]:
    """Baca Open PAM Excel file."""
    return _load_excel(filepath)


def fetch_pam_agri_db(company_id: int) -> list[dict]:
    """Ambil semua pam_records dari DB untuk company_id ini."""
    conn = get_conn()
    try:
        return [dict(r) for r in conn.execute(
            "SELECT id, pam_no, pam_date, total_amount, status, tanggal_bayar "
            "FROM pam_records WHERE company_id=?", (company_id,)
        ).fetchall()]
    finally:
        conn.close()


def fetch_open_pam_db(company_id: int) -> list[dict]:
    """Ambil payment_beasiswa status='open' dari DB.
    Filter status='open' karena Open PAM export hanya mengeksport record open.
    """
    conn = get_conn()
    try:
        return [dict(r) for r in conn.execute(
            "SELECT id, siswa_code, cat1, cat2, tanggal, amount, pam, perusahaan, status "
            "FROM payment_beasiswa WHERE company_id=? AND status='open'", (company_id,)
        ).fetchall()]
    finally:
        conn.close()


# ── Diff output ─────────────────────────────────────────────────────────────────

def print_diff(pam_result: dict, open_result: dict, dry_run: bool = True) -> None:
    sep = "=" * 66
    mode = "DRY-RUN MODE — tidak ada perubahan yang disimpan" if dry_run \
           else "APPLY MODE — perubahan akan diterapkan"
    print(f"\n{sep}\n{mode}\n{sep}\n")

    # PAM AGRI
    print("=== PAM AGRI (pam_records) ===")
    for db, ex in pam_result["updates"]:
        tgl = normalize_date(ex.get("Tgl Paid")) or "-"
        print(f"[UPDATE]  {db['pam_no']:<30} "
              f"status: {db['status']} → {ex['Status']} | tgl_paid: {tgl}")
    for db, ex in pam_result["renames"]:
        print(f"[RENAME]  {db['pam_no']:<30} → {ex['PAM No']}  (cascade 4 tabel)")
    for db in pam_result["deletes"]:
        print(f"[DELETE]  {db['pam_no']:<30} (cascade cancel_pam_record)")
    for ex in pam_result["skips"]:
        print(f"[SKIP ⚠]  {ex.get('PAM No','?'):<30} → tidak ada match di DB")

    print()

    # Open PAM
    print("=== OPEN PAM (payment_beasiswa) ===")
    for db, ex in open_result["updates"]:
        print(f"[UPDATE]  {db['siswa_code']} | {db['cat2']} | {db['tanggal']} : "
              f"pam={ex.get('PAM No','-')}, "
              f"perusahaan={ex.get('Perusahaan','-')}, "
              f"status={ex.get('Status','-')}")
    for db in open_result["deletes"]:
        print(f"[DELETE]  {db['siswa_code']} | {db['cat1']} | {db['cat2']} | {db['tanggal']}")
    for ex in open_result["skips"]:
        print(f"[SKIP ⚠]  {ex.get('Code','?')} | {ex.get('Kategori 2','?')} "
              f"→ tidak ada match di DB")

    print()
    print("-" * 66)
    pu = len(pam_result["updates"])
    pr = len(pam_result["renames"])
    pd = len(pam_result["deletes"])
    ps = len(pam_result["skips"])
    ou = len(open_result["updates"])
    od = len(open_result["deletes"])
    os_ = len(open_result["skips"])
    print(f"  pam_records      : {pu} update, {pr} rename, {pd} delete, {ps} skip")
    print(f"  payment_beasiswa : {ou} update, {od} delete, {os_} skip")
    if not dry_run:
        print("\nPerubahan diterapkan.")
    else:
        print("\nJalankan dengan --apply untuk menerapkan perubahan.")
    print(sep)


# ── Temp main for smoke test (will be replaced in Task 4) ──────────────────────

def main() -> None:
    print("Membaca Excel files...")
    ex_pam  = load_pam_agri_excel(PAM_AGRI_FILE)
    ex_open = load_open_pam_excel(OPEN_PAM_FILE)
    print(f"  PAM AGRI : {len(ex_pam)} baris")
    print(f"  Open PAM : {len(ex_open)} baris")

    print("Membaca data DB...")
    db_pam  = fetch_pam_agri_db(COMPANY_ID)
    db_open = fetch_open_pam_db(COMPANY_ID)
    print(f"  pam_records      : {len(db_pam)} rows")
    print(f"  payment_beasiswa : {len(db_open)} rows")

    print("Menghitung perubahan...")
    pam_result  = match_pam_agri(ex_pam, db_pam)
    open_result = match_open_pam(ex_open, db_open)

    print_diff(pam_result, open_result, dry_run=True)


if __name__ == "__main__":
    main()
