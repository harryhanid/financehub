import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC  = r'C:\Users\25010160\Downloads\UNIQLO_Week7_StepByStep.xlsx'
DEST = r'C:\Users\25010160\Downloads\UNIQLO_Week7_FINAL.xlsx'

thin = Side(style='thin', color="CCCCCC")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, bg, fg="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr

print("Loading workbook...")
wb = load_workbook(SRC)
ws = wb["Data + Risk Score"]
max_row = ws.max_row  # 49801

# ── Helper column definitions (start at col 30 = AD) ─────────────────────────
helpers = [
    # (header_text,  bg_color,  formula_template)
    ("LANGKAH 1\nJual Rugi?\n(cost > price)",   "C0392B",
     lambda r: f'=IF(N{r}>O{r},"YA  +40","Tidak")'),

    ("LANGKAH 2\nDiskon Besar?\n(≥20% atau 10%)", "E67E22",
     lambda r: f'=IF(OR(G{r}="20%",G{r}="30%"),"YA  +20",IF(G{r}="10%","YA  +10","Tidak"))'),

    ("LANGKAH 3\nToko Medan?",                   "8E44AD",
     lambda r: f'=IF(T{r}="Medan Center","YA  +15","Tidak")'),

    ("LANGKAH 4\nAccessories?",                  "2980B9",
     lambda r: f'=IF(I{r}="Accessories","YA  +15","Tidak")'),

    ("LANGKAH 5\nBulan Rawan?\n(2,5,8,10)",      "1F8A4C",
     lambda r: f'=IF(OR(MONTH(DATEVALUE(B{r}))=2,MONTH(DATEVALUE(B{r}))=5,MONTH(DATEVALUE(B{r}))=8,MONTH(DATEVALUE(B{r}))=10),"YA  +10","Tidak")'),
]

HELPER_START = 30  # AD

# ── Write headers ─────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 48
for i, (label, color, _) in enumerate(helpers):
    col = HELPER_START + i
    cell = ws.cell(row=1, column=col, value=label)
    hdr(cell, color)
    ws.column_dimensions[get_column_letter(col)].width = 13

# ── Write formula rows ────────────────────────────────────────────────────────
print("Adding helper columns...")

# Pre-read risk_score column (col 29) for conditional fill color
# We'll just use alternating row colors for helper cols; color by "YA" in cell text

for r in range(2, max_row + 1):
    alt_bg = "F9F9F9" if r % 2 == 0 else "FFFFFF"

    for i, (_, color, formula_fn) in enumerate(helpers):
        col = HELPER_START + i
        cell = ws.cell(row=r, column=col, value=formula_fn(r))
        cell.font = Font(size=8, color="222222")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
        cell.fill = PatternFill("solid", fgColor=alt_bg)

    if r % 5000 == 0:
        print(f"  {r:,}/{max_row:,}...")

# ── Also add a Summary sheet ──────────────────────────────────────────────────
if "Summary Analysis" not in wb.sheetnames:
    ws3 = wb.create_sheet("Summary Analysis")
    ws3.sheet_view.showGridLines = False

    for c in ('A','B','C','D','E'):
        ws3.column_dimensions[c].width = {'A':28,'B':36,'C':16,'D':26,'E':18}[c]

    rows_s = [
        ("UNIQLO Indonesia — Week 7 Analytics Summary", "", "", "", ""),
        ("49,800 Transactions | 2020–2024 | Team 2 Group F4", "", "", "", ""),
        ("", "", "", "", ""),
        ("TOP 5 PATTERNS", "", "", "", ""),
        ("Pattern", "Kondisi", "% Terdampak", "Dampak", "Verifikasi Excel"),
        ("P1 — Below-Cost SKU",       "cost_price > list_price",      "17.3%",  "IDR 8.59B loss (5yr)", "Filter col N > col O"),
        ("P2 — Heavy Discount",        "discount >= 20%",              "15.0%",  "GP margin -26pts, zero volume gain", "PivotTable G vs Gross_Profit"),
        ("P3 — Medan Store",           "store = Medan Center",         "19.9%",  "Rev/m² 4.2× di bawah Jakarta", "Filter col T = Medan Center"),
        ("P4 — Accessories Returns",   "category = Accessories",       "20.1%",  "Return rate 10.38% (tertinggi)", "PivotTable I vs returned"),
        ("P5 — Seasonal Spike",        "Bulan 2,5,8,10",               "32.7%",  "Return rate 10.24–10.80%", "MONTH(DATEVALUE(B2)) helper col"),
        ("", "", "", "", ""),
        ("WHAT-IF SCENARIOS", "", "", "", ""),
        ("Skenario", "Aksi", "Transaksi", "Gain 5 Tahun", "Biaya"),
        ("A — Hapus SKU Below-Cost",   "Stop jual jika cost > price",  "8,628",  "IDR 8.59B GP recovery", "Tidak ada"),
        ("B — Cap Diskon 10%",         "Hapus tier 20% dan 30%",       "7,490",  "IDR 1.09B GP gain",     "Tidak ada"),
        ("C — Perbaiki Medan",         "Naikkan produktivitas toko",   "9,929",  "IDR 26.92B revenue gain","Ada (investasi)"),
        ("", "", "", "", ""),
        ("REKOMENDASI: Scenario A dulu — stop kerugian terjamin, zero cost, break-even 11 hari.", "", "", "", ""),
    ]

    NAVY  = "1F3864"; WHITE = "FFFFFF"; LIGHT = "DCE6F1"
    for r_i, row_data in enumerate(rows_s, 1):
        ws3.row_dimensions[r_i].height = 20
        for c_i, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r_i, column=c_i, value=val)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(size=9)

            if r_i == 1:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(bold=True, color=WHITE, size=13)
            elif r_i == 2:
                cell.fill = PatternFill("solid", fgColor="2E4A8B")
                cell.font = Font(color=WHITE, size=8)
            elif val in ("TOP 5 PATTERNS", "WHAT-IF SCENARIOS"):
                cell.fill = PatternFill("solid", fgColor="2E4A8B")
                cell.font = Font(bold=True, color=WHITE, size=10)
            elif row_data[0] in ("Pattern", "Skenario"):
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(bold=True, color=WHITE, size=9)
            elif row_data[0].startswith("P") and len(row_data[0]) > 1 and row_data[0][1].isdigit():
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            elif row_data[0] in ("A — Hapus SKU Below-Cost","B — Cap Diskon 10%","C — Perbaiki Medan"):
                fills = {"A — Hapus SKU Below-Cost":"E2EFDA","B — Cap Diskon 10%":"FFF2CC","C — Perbaiki Medan":"DDEEFF"}
                cell.fill = PatternFill("solid", fgColor=fills.get(row_data[0], WHITE))
            elif row_data[0].startswith("REKOMENDASI"):
                cell.fill = PatternFill("solid", fgColor="FFF9E6")
                cell.font = Font(bold=True, size=9, color="8B4513")

    if r_i == 1 or rows_s[0][0]:
        try: ws3.merge_cells("A1:E1")
        except: pass
        try: ws3.merge_cells("A2:E2")
        except: pass
        try: ws3.merge_cells("A18:E18")
        except: pass

    print("  Summary sheet added.")

print("Saving...")
wb.save(DEST)
print(f"Done! → {DEST}")
