import re
import calendar
import config
from datetime import datetime
from database import get_conn


def _ts():
    return datetime.now().isoformat(timespec="seconds")


def generate_memo_number(company_id: int, company_code: str, year: str) -> str:
    prefix  = f"PAM/{company_code}/{year}/"
    pattern = re.compile(rf"PAM/{re.escape(company_code)}/{re.escape(year)}/(\d+)")
    conn = get_conn()
    rows = conn.execute(
        "SELECT memo_number FROM payment_memo WHERE company_id = ? AND memo_number LIKE ?",
        (company_id, prefix + "%")
    ).fetchall()
    conn.close()
    max_seq = 0
    for row in rows:
        m = pattern.match(row["memo_number"])
        if m:
            seq = int(m.group(1))
            if seq > max_seq:
                max_seq = seq
    return f"{prefix}{max_seq + 1:03d}"


def get_draft_payments(company_id: int) -> list:
    conn = get_conn()
    # Rows dari payment_beasiswa — JOIN ke pam_records untuk pillar & due_date
    rows = [dict(r) for r in conn.execute(
        """SELECT pb.*, s.nama, s.bank, s.norek, s.namarek, 'beasiswa' AS _type,
                  pr.pillar, pr.due_date
           FROM payment_beasiswa pb
           LEFT JOIN siswa s  ON s.company_id  = pb.company_id AND s.code   = pb.siswa_code
           LEFT JOIN pam_records pr ON pr.pam_no = pb.pam AND pr.company_id = pb.company_id
           WHERE pb.company_id = ? AND pb.status = 'open'
           ORDER BY pb.tanggal DESC""",
        (company_id,)
    ).fetchall()]
    # Rows dari pam_records untuk source others/tagihan/sponsor
    pr_rows = [dict(r) for r in conn.execute(
        """SELECT
               id, company_id, pam_no AS pam, pam_date AS tanggal,
               total_amount AS amount, pillar, source, keterangan, due_date,
               NULL AS siswa_code, NULL AS cat1, NULL AS cat2,
               NULL AS nama, NULL AS bank, NULL AS norek, NULL AS namarek,
               NULL AS etf_pa_line_id, 'pam_record' AS _type
           FROM pam_records
           WHERE company_id = ? AND source IN ('others', 'tagihan', 'sponsor')
           AND status = 'open'
           ORDER BY pam_date DESC""",
        (company_id,)
    ).fetchall()]
    conn.close()
    return rows + pr_rows


def create_memo(company_id: int, company_code: str, tanggal: str,
                notes: str, created_by: str, items: list) -> dict:
    if not items:
        return {"ok": False, "pesan": "Pilih minimal 1 payment untuk dimasukkan ke memo."}
    year        = (tanggal or datetime.now().strftime("%Y"))[:4]
    memo_number = generate_memo_number(company_id, company_code, year)
    total       = sum(float(item.get("amount", 0)) for item in items)
    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO payment_memo
               (company_id, memo_number, tanggal, total_amount, status, notes, created_by, created_at)
               VALUES (?,?,?,?,'open',?,?,?)""",
            (company_id, memo_number, tanggal, total, notes, created_by, _ts())
        )
        memo_id = cur.lastrowid
        for item in items:
            conn.execute(
                """INSERT INTO payment_memo_items
                   (memo_id, source_module, source_id, description, amount, vendor, bank_account)
                   VALUES (?,?,?,?,?,?,?)""",
                (memo_id,
                 item.get("source_module", "beasiswa"),
                 item["source_id"],
                 item.get("description", ""),
                 float(item.get("amount", 0)),
                 item.get("vendor", ""),
                 item.get("bank_account", ""))
            )
            if item.get("source_module", "beasiswa") == "beasiswa":
                conn.execute(
                    "UPDATE payment_beasiswa SET status='on_process', memo_id=? WHERE id=?",
                    (memo_id, item["source_id"])
                )
        conn.commit()
        return {"ok": True, "memo_id": memo_id, "memo_number": memo_number,
                "pesan": f"Memo {memo_number} berhasil dibuat."}
    except Exception as e:
        conn.rollback()
        return {"ok": False, "pesan": f"Gagal membuat memo: {e}"}
    finally:
        conn.close()


def get_memo_list(company_id: int, status: str = "") -> list:
    sql    = "SELECT * FROM payment_memo WHERE company_id = ?"
    params = [company_id]
    if status:
        sql    += " AND status = ?"
        params += [status]
    sql += " ORDER BY created_at DESC"
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def get_memo_detail(memo_id: int, company_id: int) -> dict | None:
    conn = get_conn()
    memo = conn.execute(
        "SELECT * FROM payment_memo WHERE id = ? AND company_id = ?",
        (memo_id, company_id)
    ).fetchone()
    if not memo:
        conn.close()
        return None
    items = [dict(r) for r in conn.execute(
        "SELECT * FROM payment_memo_items WHERE memo_id = ?", (memo_id,)
    ).fetchall()]
    conn.close()
    return {**dict(memo), "items": items}


def update_memo_status(memo_id: int, new_status: str, by_user: str, company_id: int = 0) -> dict:
    allowed = {"open", "on_process", "complete"}
    if new_status not in allowed:
        return {"ok": False, "pesan": f"Status '{new_status}' tidak valid."}

    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM payment_memo WHERE id=? AND (? = 0 OR company_id=?)",
        (memo_id, company_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "Memo tidak ditemukan."}

    now = _ts()
    conn.execute(
        "UPDATE payment_memo SET status=?, updated_at=? WHERE id=?",
        (new_status, now, memo_id)
    )

    if new_status == "complete":
        conn.execute(
            "UPDATE payment_beasiswa SET status='complete' WHERE memo_id=?",
            (memo_id,)
        )
    elif new_status in ("open", "on_process"):
        # Revert payment_beasiswa ke open saat memo diturunkan statusnya
        conn.execute(
            "UPDATE payment_beasiswa SET status='open' WHERE memo_id=? AND status='complete'",
            (memo_id,)
        )

    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Status memo diubah ke '{new_status}'."}


def set_memo_tanggal_bayar(memo_id: int, tanggal_bayar: str, company_id: int) -> dict:
    """
    Set tanggal_bayar di payment_memo → status complete,
    cascade ke etf_pa: tanggal_bayar + status=complete.
    """
    if not tanggal_bayar:
        return {"ok": False, "pesan": "Tanggal bayar wajib diisi."}

    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM payment_memo WHERE id=? AND company_id=?",
        (memo_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "Memo tidak ditemukan."}

    now = _ts()

    # 1. Update payment_memo
    conn.execute(
        "UPDATE payment_memo SET tanggal_bayar=?, status='complete', updated_at=? WHERE id=?",
        (tanggal_bayar, now, memo_id)
    )

    # 2. Update payment_beasiswa status
    conn.execute(
        "UPDATE payment_beasiswa SET status='complete' WHERE memo_id=?",
        (memo_id,)
    )

    # 3. Cascade ke semua PA tables yang di-referensi
    lines = conn.execute(
        "SELECT DISTINCT etf_pa_line_id FROM payment_beasiswa WHERE memo_id=? AND etf_pa_line_id IS NOT NULL",
        (memo_id,)
    ).fetchall()
    line_ids = [r[0] for r in lines]

    if line_ids:
        ph = ",".join("?" * len(line_ids))
        for lines_tbl, pa_tbl in [
            ("etf_pa_lines",    "etf_pa"),
            ("app_pa_lines",    "app_pa"),
            ("sml_pa_lines",    "sml_pa"),
            ("energy_pa_lines", "energy_pa"),
            ("setf_pa_lines",   "setf_pa"),
        ]:
            conn.execute(
                f"""UPDATE {pa_tbl} SET tanggal_bayar=?, status='complete', updated_at=?
                    WHERE id IN (
                        SELECT DISTINCT pa_id FROM {lines_tbl} WHERE id IN ({ph})
                    ) AND company_id=?""",
                [tanggal_bayar, now] + line_ids + [company_id]
            )

    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Tanggal bayar berhasil disimpan."}


# ── PAM helpers ─────────────────────────────────────────────────────────────

def _add_one_month(date_str: str) -> str:
    try:
        dt    = datetime.strptime(date_str, "%Y-%m-%d")
        month = dt.month % 12 + 1
        year  = dt.year + (1 if dt.month == 12 else 0)
        day   = min(dt.day, calendar.monthrange(year, month)[1])
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return date_str


_PILLAR_LINES_TABLE = {
    "AGRI":     "agri_pam_lines",
    "APP":      "app_pam_lines",
    "LAND":     "land_pam_lines",
    "ENERGY":   "energy_pam_lines",
    "SETF":     "setf_pam_lines",
    "SMT":      "smt_pam_lines",
    "ADVANCE":  "advance_pam_lines",
}
_VALID_PILLARS = set(_PILLAR_LINES_TABLE)

_STANDARD_LINE_FIELDS = [
    "no_vendor", "nama_vendor", "tgl_terima_doc", "tgl_proses",
    "tgl_verifikasi_tax", "tgl_approval_1", "tgl_approval_2",
    "tgl_approval_3", "tgl_kirim",
]
_ADVANCE_LINE_FIELDS = [
    "no_vendor", "nama_vendor", "tgl_received", "tgl_a0", "tgl_a1",
    "tgl_a2", "tgl_a3", "tgl_a4", "tgl_paid",
]
# Columns SELECTed per pillar in get_pam_by_pillar (SMT adds tgl_realisasi,
# which is system-set only — see _PILLAR_ALLOWED_FIELDS below, it is
# deliberately NOT user-editable via upsert_pam_lines).
_PILLAR_SELECT_FIELDS = {
    "ADVANCE": _ADVANCE_LINE_FIELDS,
    "SMT":     _STANDARD_LINE_FIELDS + ["tgl_realisasi"],
}
# Columns upsert_pam_lines is allowed to write per pillar.
_PILLAR_ALLOWED_FIELDS = {
    "ADVANCE": set(_ADVANCE_LINE_FIELDS),
}
_DEFAULT_ALLOWED_FIELDS = set(_STANDARD_LINE_FIELDS)

_BEASISWA_PAID_COL = {
    "APP":    "tgl_Paid_APP",
    "LAND":   "tgl_Paid_LAND",
    "ENERGY": "tgl_Paid_ENERGY",
    "SETF":   "tgl_Paid_SETF",
}
_JENJANG_SORT = {"S3": 0, "S2": 1, "S1": 2}


# Tab → pam_prefix mapping (mirrors etf_payment_application._TAB_CFG)
_IPAY_PAM_PREFIX = {
    "agri":     "ETF",
    "app":      "APP",
    "sml":      "LAND",
    "setf":     "SETF",
    "smt":      "SMT",
    "advance":  "SMT",
}


def get_pam_by_pillar(company_id: int, pillar: str,
                      search: str = "", bulan: str = "", tahun: str = "",
                      status: str = "", source: str = "") -> list:
    """Return pam_records LEFT JOIN {pillar}_pam_lines filtered by pillar."""
    if pillar not in _VALID_PILLARS:
        return []
    tbl         = _PILLAR_LINES_TABLE[pillar]
    line_fields = _PILLAR_SELECT_FIELDS.get(pillar, _STANDARD_LINE_FIELDS)
    line_select = ", ".join(f"pl.{f}" for f in line_fields)
    sql = f"""
        SELECT pr.*,
               pl.id         AS lines_id,
               {line_select},
               sla.sub_total,
               sla.cnt_tgl_pengajuan, sla.cnt_tgl_receive,
               sla.cnt_tgl_pa,        sla.cnt_tgl_final,
               sla.cnt_sla1, sla.cnt_sla2, sla.cnt_sla3,
               sla.cnt_sla4, sla.cnt_sla5, sla.cnt_sla6, sla.cnt_sla7
        FROM pam_records pr
        LEFT JOIN {tbl} pl ON pl.pam_id = pr.id
        LEFT JOIN (
            SELECT pam,
                   COUNT(*)                 AS sub_total,
                   COUNT(tgl_pengajuan)     AS cnt_tgl_pengajuan,
                   COUNT(tgl_receive)       AS cnt_tgl_receive,
                   COUNT(tgl_pa)            AS cnt_tgl_pa,
                   COUNT(tgl_final)         AS cnt_tgl_final,
                   COUNT("SLA_Date_1_LL")   AS cnt_sla1,
                   COUNT("SLA_Date_2_HT")   AS cnt_sla2,
                   COUNT("SLA_Date_3_YK")   AS cnt_sla3,
                   COUNT("SLA_Date_4_AK")   AS cnt_sla4,
                   COUNT("SLA_Date_5_PD")   AS cnt_sla5,
                   COUNT("SLA_Date_6_C2")   AS cnt_sla6,
                   COUNT("SLA_Date_7_MSIG") AS cnt_sla7
            FROM payment_beasiswa WHERE company_id = ?
            GROUP BY pam
        ) sla ON sla.pam = pr.pam_no
        WHERE pr.company_id = ? AND pr.pillar = ?
    """
    params = [company_id, company_id, pillar]
    if search:
        q       = f"%{search}%"
        sql    += " AND (pr.pam_no LIKE ? OR pr.pt LIKE ? OR pr.keterangan LIKE ?)"
        params += [q, q, q]
    if bulan:
        sql    += " AND strftime('%m', pr.pam_date) = ?"
        params += [bulan.zfill(2)]
    if tahun:
        sql    += " AND strftime('%Y', pr.pam_date) = ?"
        params += [tahun]
    if status:
        sql    += " AND pr.status = ?"
        params += [status]
    if source:
        sql    += " AND pr.source = ?"
        params += [source]
    sql += " ORDER BY pr.pam_date DESC"
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def get_advance_payments(company_id: int, status: str = "", search: str = "",
                         bulan: str = "", tahun: str = "") -> list:
    """Per-line view of payment_beasiswa rows quarantined under pillar ADVANCE."""
    sql = """
        SELECT pb.*, s.nama, pr.pam_date, pr.gl_account, pr.cost_center,
               pr.requestors_name, pr.keterangan, pr.mata_uang, pr.dpp, pr.ppn,
               pr.due_date, pr.tanggal_bayar
        FROM payment_beasiswa pb
        JOIN pam_records pr ON pr.pam_no = pb.pam AND pr.company_id = pb.company_id
        LEFT JOIN siswa s ON s.company_id = pb.company_id AND s.code = pb.siswa_code
        WHERE pb.company_id = ? AND pr.pillar = 'ADVANCE'
    """
    params = [company_id]
    if status:
        sql    += " AND pb.status = ?"
        params += [status]
    if search:
        q       = f"%{search}%"
        sql    += " AND (pb.pam LIKE ? OR s.nama LIKE ?)"
        params += [q, q]
    if bulan:
        sql    += " AND strftime('%m', pb.tanggal) = ?"
        params += [bulan.zfill(2)]
    if tahun:
        sql    += " AND strftime('%Y', pb.tanggal) = ?"
        params += [tahun]
    sql += " ORDER BY pb.tanggal DESC"
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def upsert_pam_lines(pam_id: int, pillar: str, data: dict, company_id: int) -> dict:
    """Insert or update one lines row for the given pam_id and pillar."""
    if pillar not in _VALID_PILLARS:
        return {"ok": False, "pesan": f"Pillar tidak valid: {pillar}"}
    tbl = _PILLAR_LINES_TABLE[pillar]
    ALLOWED = _PILLAR_ALLOWED_FIELDS.get(pillar, _DEFAULT_ALLOWED_FIELDS)
    fields = {k: v for k, v in data.items() if k in ALLOWED}
    if not fields:
        return {"ok": False, "pesan": "Tidak ada field yang valid."}
    conn = get_conn()
    pam = conn.execute(
        "SELECT id FROM pam_records WHERE id=? AND company_id=? AND pillar=?",
        (pam_id, company_id, pillar)
    ).fetchone()
    if not pam:
        conn.close()
        return {"ok": False, "pesan": "PAM tidak ditemukan."}
    existing = conn.execute(
        f"SELECT id FROM {tbl} WHERE pam_id=?", (pam_id,)
    ).fetchone()
    now = _ts()
    if existing:
        set_clause = ", ".join(f"{k}=?" for k in fields)
        vals       = list(fields.values()) + [now, pam_id]
        conn.execute(
            f"UPDATE {tbl} SET {set_clause}, updated_at=? WHERE pam_id=?", vals
        )
    else:
        cols = ", ".join(["pam_id"] + list(fields.keys()) + ["created_at"])
        ph   = ", ".join(["?"] * (len(fields) + 2))
        vals = [pam_id] + list(fields.values()) + [now]
        conn.execute(f"INSERT INTO {tbl} ({cols}) VALUES ({ph})", vals)

    if pillar == "ADVANCE" and fields.get("tgl_paid"):
        _convert_advance_to_smt(conn, pam_id, fields["tgl_paid"], now)

    conn.commit()
    conn.close()
    return {"ok": True}


def _convert_advance_to_smt(conn, pam_id: int, tgl_realisasi: str, now: str) -> None:
    """Advance realized (tgl_paid filled) -> flip pillar to SMT.

    Carries vendor into a fresh smt_pam_lines row (7 standard stage dates
    start empty, tgl_realisasi records when the advance was realized).
    The old advance_pam_lines row is left in place as an archive — it is
    simply no longer reachable via get_pam_by_pillar('ADVANCE') once the
    pillar flips, since that query filters by pam_records.pillar.
    """
    adv = conn.execute(
        "SELECT no_vendor, nama_vendor FROM advance_pam_lines WHERE pam_id=?",
        (pam_id,)
    ).fetchone()
    no_vendor   = adv["no_vendor"]   if adv else None
    nama_vendor = adv["nama_vendor"] if adv else None
    conn.execute(
        "UPDATE pam_records SET pillar='SMT', status='complete', updated_at=? WHERE id=?",
        (now, pam_id)
    )
    conn.execute(
        """INSERT INTO smt_pam_lines (pam_id, no_vendor, nama_vendor, tgl_realisasi, created_at)
           VALUES (?,?,?,?,?)""",
        (pam_id, no_vendor, nama_vendor, tgl_realisasi, now)
    )


def bulk_update_pam_lines_dates(pillar: str, ids: list, field: str,
                                 value, company_id: int) -> dict:
    """Bulk-set one date field in {pillar}_pam_lines across multiple pam_records ids.

    Upserts per id (via upsert_pam_lines) rather than a single UPDATE, since a
    lines row may not exist yet for a given pam_id.
    """
    ALLOWED = {"tgl_terima_doc", "tgl_proses", "tgl_verifikasi_tax",
               "tgl_approval_1", "tgl_approval_2", "tgl_approval_3", "tgl_kirim"}
    if pillar not in _VALID_PILLARS:
        return {"ok": False, "pesan": f"Pillar tidak valid: {pillar}"}
    if field not in ALLOWED:
        return {"ok": False, "pesan": f"Kolom tidak valid: {field}"}
    if not ids:
        return {"ok": False, "pesan": "Tidak ada baris yang dipilih."}
    updated = 0
    for pam_id in ids:
        result = upsert_pam_lines(pam_id, pillar, {field: value or None}, company_id)
        if result.get("ok"):
            updated += 1
    return {"ok": True, "updated": updated,
            "pesan": f"{updated} dari {len(ids)} baris berhasil diperbarui."}


def get_next_pam_no(company_id: int, company_code: str,
                    tab: str, date_str: str) -> str:
    """Return next PAM number for selected type, e.g. 'PAM-054-SETF-06-2026'."""
    prefix = _IPAY_PAM_PREFIX.get(tab, company_code)
    year   = date_str[:4]
    month  = date_str[5:7]
    return generate_pam_number(company_id, prefix, year, month)


def get_siswa_medical(company_id: int, search: str = "") -> list:
    sql = """
        SELECT s.code, s.nama,
               b.pillar AS pillar,
               SUM(b.amount) AS medical_budget,
               SUM(CASE WHEN b.cat2='Rawat Inap'  THEN b.amount ELSE 0 END) AS budget_inap,
               SUM(CASE WHEN b.cat2='Rawat Jalan' THEN b.amount ELSE 0 END) AS budget_jalan,
               COALESCE((
                   SELECT SUM(pb.amount)
                   FROM payment_beasiswa pb
                   WHERE pb.siswa_code = s.code
                     AND pb.company_id = s.company_id
                     AND pb.cat1 = 'By Medical'
               ), 0) AS spent_amount,
               COALESCE((
                   SELECT SUM(pb.amount)
                   FROM payment_beasiswa pb
                   WHERE pb.siswa_code = s.code
                     AND pb.company_id = s.company_id
                     AND pb.cat1 = 'By Medical'
                     AND pb.cat2 = 'Rawat Inap'
               ), 0) AS spent_inap,
               COALESCE((
                   SELECT SUM(pb.amount)
                   FROM payment_beasiswa pb
                   WHERE pb.siswa_code = s.code
                     AND pb.company_id = s.company_id
                     AND pb.cat1 = 'By Medical'
                     AND pb.cat2 = 'Rawat Jalan'
               ), 0) AS spent_jalan
        FROM siswa s
        JOIN budget_beasiswa b ON b.siswa_code = s.code
                               AND b.company_id = s.company_id
                               AND b.cat1 = 'By Medical'
        WHERE s.company_id = ?
    """
    params: list = [company_id]
    if search:
        sql    += " AND (s.nama LIKE ? OR s.code LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]
    sql += " GROUP BY s.code, s.nama, b.pillar ORDER BY s.nama"
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def save_klaim_payment(company_id: int, company_code: str, data: dict) -> dict:
    tanggal    = data.get("tanggal") or _ts()[:10]
    pam_no     = (data.get("pam_no") or "").strip()
    keterangan = data.get("keterangan") or ""
    perusahaan = data.get("perusahaan") or ""
    pillar     = data.get("pillar") or ""
    rows       = data.get("rows") or []

    if not pam_no:
        return {"ok": False, "pesan": "No. PAM wajib diisi."}
    if not rows:
        return {"ok": False, "pesan": "Minimal 1 baris siswa."}

    for row in rows:
        if not (row.get("cat3_items") and
                any(float(i.get("amount", 0)) > 0 for i in row["cat3_items"])):
            return {"ok": False,
                    "pesan": f"Siswa {row.get('siswa_code', '?')} harus memiliki minimal 1 cat3 dengan amount > 0."}

    conn = get_conn()
    try:
        grand_total = 0.0
        for row in rows:
            siswa_code   = (row.get("siswa_code") or "").strip()
            cat2         = row.get("cat2") or ""
            kelas        = row.get("kelas") or ""
            rumah_sakit  = row.get("rumah_sakit") or ""
            diagnosa     = row.get("diagnosa") or ""
            spesialisasi = row.get("spesialisasi") or ""
            cat3_items   = [i for i in row.get("cat3_items", [])
                            if float(i.get("amount", 0)) > 0]
            row_total    = sum(float(i["amount"]) for i in cat3_items)

            cur = conn.execute(
                """INSERT INTO payment_beasiswa
                   (company_id, siswa_code, cat1, cat2, tanggal, amount,
                    pillar, perusahaan, pam, status)
                   VALUES (?,?,?,?,?,?,?,?,?,'open')""",
                (company_id, siswa_code, "By Medical", cat2, tanggal,
                 row_total, pillar, perusahaan, pam_no)
            )
            pb_id = cur.lastrowid

            for item in cat3_items:
                conn.execute(
                    """INSERT INTO klaim_medical
                       (company_id, siswa_code, pam, tanggal, amount, perawatan,
                        kelas, rumah_sakit, diagnosa, spesialisasi,
                        pillar, perusahaan, payment_id, created_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (company_id, siswa_code, pam_no,
                     item.get("tanggal") or tanggal,
                     float(item["amount"]),
                     item.get("cat3") or "",
                     kelas, rumah_sakit, diagnosa, spesialisasi,
                     pillar, perusahaan, pb_id, _ts())
                )
            grand_total += row_total

        due_date = _add_one_month(tanggal)
        conn.execute(
            """INSERT INTO pam_records
               (company_id, pam_no, pam_date, requestors_name, keterangan,
                total_amount, due_date, pillar, source, status, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, pam_no, tanggal,
             config.PAM_DEFAULT_REQUESTOR, keterangan,
             grand_total, due_date, pillar, "klaim_medis", "open", _ts())
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return {"ok": False, "pesan": f"Gagal menyimpan: {e}"}

    conn.close()
    return {"ok": True, "pesan": f"PAM Klaim Medis {pam_no} berhasil dibuat.", "pam_no": pam_no}


def save_others_payment(company_id: int, company_code: str, data: dict) -> dict:
    pam_no     = (data.get("pam_no") or "").strip()
    keterangan = (data.get("keterangan") or "").strip()
    tanggal    = data.get("tanggal") or _ts()[:10]
    perusahaan = data.get("perusahaan") or ""
    pillar     = data.get("pillar") or ""
    transaksi  = (data.get("transaksi") or "others").lower()
    mata_uang  = data.get("mata_uang") or "IDR"

    try:
        dpp = float(data.get("dpp") or 0)
        ppn = float(data.get("ppn") or 0)
    except (ValueError, TypeError):
        dpp, ppn = 0.0, 0.0

    if not pam_no:
        return {"ok": False, "pesan": "No. PAM wajib diisi."}
    if not keterangan:
        return {"ok": False, "pesan": "Keterangan wajib diisi."}
    if dpp <= 0:
        return {"ok": False, "pesan": "DPP harus lebih dari 0."}

    total    = dpp + ppn
    due_date = _add_one_month(tanggal)
    conn     = get_conn()
    try:
        conn.execute(
            """INSERT INTO pam_records
               (company_id, pam_no, pam_date, requestors_name, keterangan,
                total_amount, due_date, pillar, source, pt,
                mata_uang, dpp, ppn, status, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, pam_no, tanggal,
             config.PAM_DEFAULT_REQUESTOR, keterangan,
             total, due_date, pillar, transaksi, perusahaan,
             mata_uang, dpp, ppn, "open", _ts())
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return {"ok": False, "pesan": f"Gagal menyimpan: {e}"}

    conn.close()
    return {"ok": True, "pesan": f"PAM {pam_no} berhasil dibuat.", "pam_no": pam_no}


def save_smt_pam_transaction(company_id: int, company_code: str, data: dict) -> dict:
    """Itemized GL/Advance PAM save for SMT — replaces save_others_payment for
    company SMT only. ETF keeps using save_others_payment unchanged."""
    tanggal    = data.get("tanggal") or _ts()[:10]
    pam_no     = (data.get("pam_no") or "").strip()
    perusahaan = data.get("perusahaan") or ""
    pillar     = (data.get("pillar") or "").upper()
    transaksi  = (data.get("transaksi") or "gl").lower()
    rows       = data.get("rows") or []

    if not pam_no:
        return {"ok": False, "pesan": "No. PAM wajib diisi."}
    if not rows:
        return {"ok": False, "pesan": "Minimal 1 baris transaksi."}

    for row in rows:
        if not (row.get("klasifikasi_sr") or "").strip():
            return {"ok": False, "pesan": "Setiap baris wajib memilih Jenis Biaya (SR)."}
        try:
            dpp = float(row.get("dpp") or 0)
        except (TypeError, ValueError):
            dpp = 0
        if dpp <= 0:
            return {"ok": False, "pesan": "Setiap baris wajib DPP lebih dari 0."}
        if not (row.get("keterangan") or "").strip():
            return {"ok": False, "pesan": "Setiap baris wajib diisi Keterangan."}

    conn = get_conn()
    try:
        grand_dpp = 0.0
        grand_ppn = 0.0
        line_data = []
        for row in rows:
            dpp = float(row.get("dpp") or 0)
            ppn = float(row.get("ppn") or 0)
            grand_dpp += dpp
            grand_ppn += ppn
            line_data.append((row, dpp, ppn, dpp + ppn))
        grand_total = grand_dpp + grand_ppn

        due_date = _add_one_month(tanggal)
        cur = conn.execute(
            """INSERT INTO pam_records
               (company_id, pam_no, pam_date, requestors_name, keterangan,
                total_amount, due_date, pillar, source, pt,
                mata_uang, dpp, ppn, status, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, pam_no, tanggal,
             config.PAM_DEFAULT_REQUESTOR, "Lihat rincian baris",
             grand_total, due_date, pillar, transaksi, perusahaan,
             "IDR", grand_dpp, grand_ppn, "open", _ts())
        )
        pam_id = cur.lastrowid

        for row, dpp, ppn, total in line_data:
            conn.execute(
                """INSERT INTO pam_transaction_lines
                   (pam_id, coa_pam_id, klasifikasi_sr, klasifikasi_mr, gl_account,
                    tipe_dokumen, no_invoice, dpp, ppn, total_amount,
                    cost_center, budget_activity, keterangan, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (pam_id, row.get("coa_pam_id"), row.get("klasifikasi_sr"),
                 row.get("klasifikasi_mr"), row.get("gl_account"),
                 row.get("tipe_dokumen"), row.get("no_invoice"),
                 dpp, ppn, total, row.get("cost_center"),
                 row.get("budget_activity"), row.get("keterangan"), _ts())
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return {"ok": False, "pesan": f"Gagal menyimpan: {e}"}

    conn.close()
    return {"ok": True, "pesan": f"PAM {pam_no} berhasil dibuat.", "pam_no": pam_no}


def save_pa_payment(company_id: int, company_code: str, data: dict) -> dict:
    """
    Unified save for Input PA (AGRI/APP/SML/SETF):
    1. Insert payment_beasiswa rows via insert_payment_rows (no pam_record)
    2. Link rows to user-provided pam_no
    3. Create exactly one pam_records entry with correct total
    4. Update PA header: nomor_pam + status='on_process'
    """
    from modules.beasiswa.service import insert_payment_rows
    from modules.etf_payment_application.service import _TAB_CFG

    tab        = (data.get("tab") or "agri").lower()
    tanggal    = data.get("tanggal") or _ts()[:10]
    pam_no     = (data.get("pam_no") or "").strip()
    keterangan = data.get("keterangan") or ""
    perusahaan = data.get("perusahaan") or ""
    pillar     = data.get("pillar") or ""
    rows       = data.get("rows") or []

    if not pam_no:
        return {"ok": False, "pesan": "No. PAM wajib diisi."}
    if not rows:
        return {"ok": False, "pesan": "Minimal 1 baris siswa."}

    pa_tbl, lines_tbl, _, _ = _TAB_CFG.get(tab, _TAB_CFG["agri"])

    conn = get_conn()
    try:
        # 0. Derive route from the PA header of the lines being pulled — route is a
        #    decision made at create_pa() time, not here. Reject if the selected rows
        #    span PA headers with different routes.
        line_ids_for_route = [r.get("etf_pa_line_id") for r in rows if r.get("etf_pa_line_id")]
        route = "gl"
        if line_ids_for_route:
            ph0 = ",".join("?" * len(line_ids_for_route))
            route_rows = conn.execute(
                f"SELECT DISTINCT route FROM {lines_tbl} WHERE id IN ({ph0})",
                line_ids_for_route
            ).fetchall()
            distinct_routes = {(r[0] or "gl") for r in route_rows}
            if len(distinct_routes) > 1:
                conn.close()
                return {"ok": False, "pesan": "Baris yang dipilih berasal dari PA dengan route berbeda (GL dan Advance tidak bisa digabung dalam satu PAM). Pisahkan submission-nya."}
            route = distinct_routes.pop() if distinct_routes else "gl"

        # 1. Insert payment rows — does NOT create pam_record
        ins = insert_payment_rows(conn, company_id, company_code,
                                  tanggal, pillar, perusahaan, rows, route=route)
        if not ins.get("ok"):
            conn.close()
            return ins

        payment_ids = ins["payment_ids"]
        total       = ins["total"]

        # 2. Create pam_records entry FIRST (FK must exist before UPDATE payment_beasiswa)
        due_date    = _add_one_month(tanggal)
        cost_center = config.COST_CENTER_MAP.get(perusahaan, "")
        pillar_for_pam = "ADVANCE" if route == "advance" else pillar
        conn.execute(
            """INSERT INTO pam_records
               (company_id, pam_no, pam_date, requestors_name, keterangan,
                total_amount, dpp, ppn, due_date, pillar, source,
                pt, cost_center, status, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'open',?)""",
            (company_id, pam_no, tanggal,
             config.PAM_DEFAULT_REQUESTOR, keterangan,
             total, total, 0.0,
             due_date, pillar_for_pam, "beasiswa",
             perusahaan, cost_center, _ts())
        )

        # 3. Link payment_beasiswa rows to user pam_no (pam_records now exists)
        if payment_ids:
            ph = ",".join("?" * len(payment_ids))
            conn.execute(
                f"UPDATE payment_beasiswa SET pam=? WHERE id IN ({ph})",
                [pam_no] + list(payment_ids)
            )

        # 4. Update PA: nomor_pam + status='on_process'
        line_ids = [r.get("etf_pa_line_id") for r in rows
                    if r.get("etf_pa_line_id")]
        if line_ids:
            ph = ",".join("?" * len(line_ids))
            pa_rows = conn.execute(
                f"SELECT DISTINCT pa_id FROM {lines_tbl} WHERE id IN ({ph})",
                line_ids
            ).fetchall()
            pa_ids = [row[0] for row in pa_rows]
            if pa_ids:
                ph2 = ",".join("?" * len(pa_ids))
                conn.execute(
                    f"UPDATE {pa_tbl} SET nomor_pam=?, status='on_process'"
                    f" WHERE id IN ({ph2}) AND company_id=?",
                    [pam_no] + list(pa_ids) + [company_id]
                )

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return {"ok": False, "pesan": f"Gagal menyimpan: {e}"}

    conn.close()
    return {"ok": True, "pesan": f"PAM {pam_no} berhasil dibuat.", "pam_no": pam_no}


def generate_pam_number(company_id: int, company_code: str, year: str,
                        month: str, conn=None) -> str:
    like_pat  = f"PAM-%-{company_code}-{month}-{year}"
    pattern   = re.compile(rf"PAM-(\d+)-{re.escape(company_code)}-{re.escape(month)}-{re.escape(year)}")
    owns_conn = conn is None
    if owns_conn:
        conn = get_conn()
    rows = conn.execute(
        "SELECT pam_no FROM pam_records WHERE company_id=? AND pam_no LIKE ?",
        (company_id, like_pat)
    ).fetchall()
    if owns_conn:
        conn.close()
    max_seq = 0
    for row in rows:
        m = pattern.match(row["pam_no"])
        if m:
            seq = int(m.group(1))
            if seq > max_seq:
                max_seq = seq
    return f"PAM-{max_seq + 1:03d}-{company_code}-{month}-{year}"


def create_pam_record(conn, company_id: int, company_code: str,
                      data: dict) -> str:
    pam_date    = data.get("pam_date") or _ts()[:10]
    year        = pam_date[:4]
    month       = pam_date[5:7]
    pam_no      = generate_pam_number(company_id, company_code, year, month, conn)
    due_date    = _add_one_month(pam_date)
    cost_center = config.COST_CENTER_MAP.get(data.get("pt", ""), "")
    total       = float(data.get("total_amount", 0))
    dpp         = float(data.get("dpp", total))
    conn.execute(
        """INSERT INTO pam_records
           (company_id, pam_no, pam_date, gl_account, cost_center, pt,
            requestors_name, keterangan, total_amount, dpp, ppn,
            due_date, pillar, source, status, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'open',?)""",
        (company_id, pam_no, pam_date,
         data.get("gl_account", config.PAM_DEFAULT_GL),
         cost_center, data.get("pt", ""),
         data.get("requestors_name", config.PAM_DEFAULT_REQUESTOR),
         data.get("keterangan", ""),
         total, dpp, 0.0,
         due_date, data.get("pillar", ""), data.get("source", "beasiswa"),
         _ts())
    )
    for pid in data.get("payment_ids", []):
        conn.execute(
            "UPDATE payment_beasiswa SET pam=? WHERE id=?", (pam_no, pid)
        )
    return pam_no


def get_pam_list(company_id: int, search: str = "", bulan: str = "",
                 tahun: str = "", source: str = "", pillar: str = "",
                 status: str = "") -> list:
    sql    = "SELECT * FROM pam_records WHERE company_id=?"
    params = [company_id]
    if pillar:
        sql    += " AND pillar=?"
        params += [pillar.upper()]
    if search:
        q       = f"%{search}%"
        sql    += " AND (pam_no LIKE ? OR pt LIKE ? OR keterangan LIKE ?)"
        params += [q, q, q]
    if bulan:
        sql    += " AND strftime('%m', pam_date)=?"
        params += [bulan.zfill(2)]
    if tahun:
        sql    += " AND strftime('%Y', pam_date)=?"
        params += [tahun]
    if source:
        sql    += " AND source LIKE ?"
        params += [f"etf_{source}%"]
    if status:
        sql    += " AND status=?"
        params += [status]
    sql += " ORDER BY created_at DESC"
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def get_coa_list() -> list:
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        "SELECT gl_code, gl_name FROM coa WHERE is_active=1 ORDER BY gl_code"
    ).fetchall()]
    conn.close()
    return rows


def get_coa_pam_list(search: str = "") -> list:
    conn = get_conn()
    if search:
        rows = conn.execute(
            """SELECT * FROM coa_pam
               WHERE klasifikasi_sr LIKE ? OR klasifikasi_mr LIKE ?
               ORDER BY klasifikasi_sr""",
            (f"%{search}%", f"%{search}%")
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM coa_pam ORDER BY klasifikasi_sr").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pam_transaction_lines(pam_id: int) -> list:
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM pam_transaction_lines WHERE pam_id=? ORDER BY id", (pam_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def update_pam_gl_account(pam_id: int, gl_account: str,
                           company_id: int) -> dict:
    conn = get_conn()
    coa  = conn.execute(
        "SELECT gl_code FROM coa WHERE gl_code=? AND is_active=1", (gl_account,)
    ).fetchone()
    if not coa:
        conn.close()
        return {"ok": False, "pesan": f"GL Account '{gl_account}' tidak ditemukan di COA."}
    row = conn.execute(
        "SELECT id FROM pam_records WHERE id=? AND company_id=?",
        (pam_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "PAM record tidak ditemukan."}
    conn.execute(
        "UPDATE pam_records SET gl_account=?, updated_at=? WHERE id=? AND company_id=?",
        (gl_account, _ts(), pam_id, company_id)
    )
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"GL Account diubah ke {gl_account}."}


def update_pam_status(pam_id: int, new_status: str, company_id: int) -> dict:
    allowed = {"open", "on_process", "complete"}
    if new_status not in allowed:
        return {"ok": False, "pesan": f"Status '{new_status}' tidak valid."}
    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM pam_records WHERE id=? AND company_id=?",
        (pam_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "PAM record tidak ditemukan."}
    conn.execute(
        "UPDATE pam_records SET status=?, updated_at=? WHERE id=? AND company_id=?",
        (new_status, _ts(), pam_id, company_id)
    )
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Status PAM diubah ke '{new_status}'."}


def update_pam_record(pam_id: int, data: dict, company_id: int) -> dict:
    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM pam_records WHERE id=? AND company_id=?",
        (pam_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "PAM record tidak ditemukan."}
    allowed_fields = ("keterangan", "requestors_name", "total_amount", "due_date", "pam_date", "tanggal_bayar")
    fields, params = [], []
    for key in allowed_fields:
        if key in data and data[key] is not None:
            fields.append(f"{key}=?")
            params.append(data[key])
    if not fields:
        conn.close()
        return {"ok": False, "pesan": "Tidak ada data yang diubah."}
    params += [_ts(), pam_id, company_id]
    conn.execute(
        f"UPDATE pam_records SET {', '.join(fields)}, updated_at=? WHERE id=? AND company_id=?",
        params
    )
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "PAM record berhasil diperbarui."}


def delete_payment_beasiswa(payment_id: int, company_id: int) -> dict:
    conn = get_conn()
    row = conn.execute(
        "SELECT status FROM payment_beasiswa WHERE id=? AND company_id=?",
        (payment_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "Payment tidak ditemukan."}
    if row["status"] != "open":
        conn.close()
        return {"ok": False, "pesan": "Hanya payment berstatus open yang bisa dihapus."}
    conn.execute("DELETE FROM klaim_medical WHERE payment_id=? AND company_id=?", (payment_id, company_id))
    conn.execute("DELETE FROM payment_beasiswa WHERE id=? AND company_id=?", (payment_id, company_id))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Payment berhasil dihapus."}


def cancel_pam_record(pam_id: int, company_id: int) -> dict:
    conn = get_conn()
    pam = conn.execute(
        "SELECT pam_no, source FROM pam_records WHERE id=? AND company_id=?", (pam_id, company_id)
    ).fetchone()
    if not pam:
        conn.close()
        return {"ok": False, "pesan": "PAM record tidak ditemukan."}
    pam_no = pam["pam_no"]
    source = pam["source"] or "beasiswa"

    now = _ts()

    if source == "etf_agri":
        # AGRI flow: revert etf_pa yang linked via nomor_pam → back to 'open', clear nomor_pam
        conn.execute(
            """UPDATE etf_pa SET status='open', nomor_pam=NULL, updated_at=?
               WHERE nomor_pam=? AND company_id=?""",
            (now, pam_no, company_id)
        )
        conn.execute("DELETE FROM pam_records WHERE id=? AND company_id=?", (pam_id, company_id))
    else:
        # Beasiswa flow
        lines = conn.execute(
            "SELECT DISTINCT etf_pa_line_id FROM payment_beasiswa WHERE pam=? AND company_id=? AND etf_pa_line_id IS NOT NULL",
            (pam_no, company_id)
        ).fetchall()
        line_ids = [r[0] for r in lines]

        conn.execute("DELETE FROM klaim_medical WHERE pam=? AND company_id=?", (pam_no, company_id))
        conn.execute("DELETE FROM payment_beasiswa WHERE pam=? AND company_id=?", (pam_no, company_id))
        conn.execute("DELETE FROM pam_records WHERE id=? AND company_id=?", (pam_id, company_id))

        # Revert semua PA tables ke open jika tidak ada payment aktif lagi
        if line_ids:
            ph = ",".join("?" * len(line_ids))
            for lines_tbl, pa_tbl in [
                ("etf_pa_lines",    "etf_pa"),
                ("app_pa_lines",    "app_pa"),
                ("sml_pa_lines",    "sml_pa"),
                ("energy_pa_lines", "energy_pa"),
                ("setf_pa_lines",   "setf_pa"),
            ]:
                pa_rows = conn.execute(
                    f"SELECT DISTINCT pa_id FROM {lines_tbl} WHERE id IN ({ph})",
                    line_ids
                ).fetchall()
                for row in pa_rows:
                    pa_id_inner = row[0]
                    remaining = conn.execute(
                        f"""SELECT COUNT(*) FROM payment_beasiswa pb
                               JOIN {lines_tbl} el ON el.id = pb.etf_pa_line_id
                               WHERE el.pa_id=? AND pb.company_id=?""",
                        (pa_id_inner, company_id)
                    ).fetchone()[0]
                    if remaining == 0:
                        conn.execute(
                            f"UPDATE {pa_tbl} SET status='open', nomor_pam=NULL, updated_at=? "
                            f"WHERE id=? AND company_id=?",
                            (now, pa_id_inner, company_id)
                        )

    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"PAM {pam_no} berhasil dihapus, PA terkait dikembalikan ke Open."}


def remove_student_from_pam(payment_beasiswa_id: int, company_id: int) -> dict:
    """Hapus satu siswa dari Open PAM dengan cascade:
    1. DELETE payment_beasiswa row
    2. Recalculate pam_records.total_amount (atau DELETE jika tidak ada sisa)
    3. Revert etf_pa (dan pillar PA lain) jika tidak ada payment aktif tersisa
    """
    conn = get_conn()
    now = _ts()

    # 1. Ambil payment_beasiswa row
    pb = conn.execute(
        "SELECT id, pam, etf_pa_line_id, amount FROM payment_beasiswa WHERE id=? AND company_id=?",
        (payment_beasiswa_id, company_id)
    ).fetchone()
    if not pb:
        conn.close()
        return {"ok": False, "pesan": "Payment tidak ditemukan."}

    pam_no       = pb["pam"]
    line_id      = pb["etf_pa_line_id"]

    # 2. Validasi pam_records.status == 'open'
    pam_rec = conn.execute(
        "SELECT id, status FROM pam_records WHERE pam_no=? AND company_id=?",
        (pam_no, company_id)
    ).fetchone()
    if not pam_rec:
        conn.close()
        return {"ok": False, "pesan": "PAM tidak ditemukan."}
    if pam_rec["status"] != "open":
        conn.close()
        return {"ok": False, "pesan": "PAM sudah diproses, tidak dapat mengubah siswa."}

    # 3. Hapus payment_beasiswa (dan klaim_medical jika ada)
    conn.execute("DELETE FROM klaim_medical WHERE payment_id=? AND company_id=?",
                 (payment_beasiswa_id, company_id))
    conn.execute("DELETE FROM payment_beasiswa WHERE id=? AND company_id=?",
                 (payment_beasiswa_id, company_id))

    # 4. Hitung sisa rows untuk PAM ini
    remaining_count = conn.execute(
        "SELECT COUNT(*) FROM payment_beasiswa WHERE pam=? AND company_id=?",
        (pam_no, company_id)
    ).fetchone()[0]

    if remaining_count == 0:
        # Hapus pam_records jika sudah kosong
        conn.execute("DELETE FROM pam_records WHERE pam_no=? AND company_id=?",
                     (pam_no, company_id))
    else:
        # Recalculate total
        new_total = conn.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM payment_beasiswa WHERE pam=? AND company_id=?",
            (pam_no, company_id)
        ).fetchone()[0]
        conn.execute(
            "UPDATE pam_records SET total_amount=?, updated_at=? WHERE pam_no=? AND company_id=?",
            (new_total, now, pam_no, company_id)
        )

    # 5. Revert PA entry jika tidak ada payment aktif tersisa
    if line_id:
        for lines_tbl, pa_tbl in [
            ("etf_pa_lines",    "etf_pa"),
            ("app_pa_lines",    "app_pa"),
            ("sml_pa_lines",    "sml_pa"),
            ("energy_pa_lines", "energy_pa"),
            ("setf_pa_lines",   "setf_pa"),
        ]:
            pa_row = conn.execute(
                f"SELECT pa_id FROM {lines_tbl} WHERE id=?", (line_id,)
            ).fetchone()
            if not pa_row:
                continue
            pa_id = pa_row[0]
            # etf_pa_line_id is the single shared FK column across all pillars in payment_beasiswa
            still_has_payment = conn.execute(
                f"""SELECT COUNT(*) FROM payment_beasiswa pb
                    JOIN {lines_tbl} el ON el.id = pb.etf_pa_line_id
                    WHERE el.pa_id=? AND pb.company_id=?""",
                (pa_id, company_id)
            ).fetchone()[0]
            if still_has_payment == 0:
                conn.execute(
                    f"UPDATE {pa_tbl} SET status='open', nomor_pam=NULL, updated_at=? "
                    f"WHERE id=? AND company_id=?",
                    (now, pa_id, company_id)
                )
            break  # line_id hanya bisa ada di satu PA table

    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Siswa berhasil dikeluarkan dari PAM."}


def get_pam_detail(pam_id: int, company_id: int) -> dict | None:
    conn = get_conn()
    pam = conn.execute(
        "SELECT * FROM pam_records WHERE id=? AND company_id=?", (pam_id, company_id)
    ).fetchone()
    if not pam:
        conn.close()
        return None
    result = dict(pam)
    app = conn.execute(
        """SELECT pa.* FROM payment_application pa
           JOIN payment_beasiswa pb ON pb.memo_id = pa.memo_id AND pb.company_id = pa.company_id
           WHERE pb.pam=? AND pa.company_id=? LIMIT 1""",
        (result["pam_no"], company_id)
    ).fetchone()
    result["payment_application"] = dict(app) if app else None
    conn.close()
    return result


def get_pam_payments(pam_no: str, company_id: int) -> list:
    from collections import defaultdict
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        """SELECT pb.id, pb.siswa_code, pb.cat1, pb.cat2,
                  pb.amount, pb.tanggal,
                  s.nama, s.bank, s.norek, s.namarek, s.jenjang
           FROM payment_beasiswa pb
           LEFT JOIN siswa s
             ON s.company_id = pb.company_id AND s.code = pb.siswa_code
           WHERE pb.pam = ? AND pb.company_id = ?""",
        (pam_no, company_id)
    ).fetchall()]
    conn.close()
    if not rows:
        return rows
    totals: dict = defaultdict(float)
    jenjang_of: dict = {}
    for r in rows:
        code = r.get("siswa_code") or ""
        totals[code] += float(r.get("amount") or 0)
        if code not in jenjang_of:
            jenjang_of[code] = (r.get("jenjang") or "").upper()
    rows.sort(key=lambda r: (
        _JENJANG_SORT.get(jenjang_of.get(r.get("siswa_code") or "", ""), 99),
        -totals.get(r.get("siswa_code") or "", 0.0),
    ))
    return rows


def get_pam_payments_detail(pam_no: str, company_id: int) -> list:
    conn = get_conn()
    pay_rows = [dict(r) for r in conn.execute(
        """SELECT pb.siswa_code, pb.cat1, pb.cat2, pb.amount,
                  s.nama, s.angkatan, s.jenjang, s.program, s.universitas, s.fakultas,
                  s.bank, s.norek, s.namarek
           FROM payment_beasiswa pb
           LEFT JOIN siswa s ON s.company_id = pb.company_id AND s.code = pb.siswa_code
           WHERE pb.pam = ? AND pb.company_id = ?
           ORDER BY CASE s.jenjang
               WHEN 'S3' THEN 1
               WHEN 'S2' THEN 2
               WHEN 'S1' THEN 3
               ELSE 99
           END, pb.siswa_code, pb.id
           -- approximate pre-sort only; Python sort below is authoritative""",
        (pam_no, company_id)
    ).fetchall()]
    if not pay_rows:
        conn.close()
        return []

    codes = list(dict.fromkeys(r["siswa_code"] for r in pay_rows))
    ph    = ",".join("?" * len(codes))

    bud = {r["siswa_code"]: dict(r) for r in conn.execute(
        f"""SELECT siswa_code,
                   SUM(CASE WHEN cat1='By Pendidikan' THEN amount ELSE 0 END) AS bud_p,
                   SUM(CASE WHEN cat1='By Tunjangan'  THEN amount ELSE 0 END) AS bud_t,
                   SUM(CASE WHEN cat1='By Penelitian' THEN amount ELSE 0 END) AS bud_r
            FROM budget_beasiswa
            WHERE company_id=? AND siswa_code IN ({ph})
            GROUP BY siswa_code""",
        [company_id] + codes
    ).fetchall()}

    paid = {r["siswa_code"]: dict(r) for r in conn.execute(
        f"""SELECT siswa_code,
                   SUM(CASE WHEN cat1='By Pendidikan' THEN amount ELSE 0 END) AS paid_p,
                   SUM(CASE WHEN cat1='By Tunjangan'  THEN amount ELSE 0 END) AS paid_t,
                   SUM(CASE WHEN cat1='By Penelitian' THEN amount ELSE 0 END) AS paid_r
            FROM payment_beasiswa
            WHERE company_id=? AND siswa_code IN ({ph})
            GROUP BY siswa_code""",
        [company_id] + codes
    ).fetchall()}

    conn.close()

    siswa_info = {}
    ket_map    = {}
    ket_order  = {}

    for r in pay_rows:
        code = r["siswa_code"]
        if code not in siswa_info:
            siswa_info[code] = {k: r[k] for k in
                ("nama","angkatan","jenjang","program","universitas","fakultas",
                 "bank","norek","namarek")}
            ket_map[code]   = {}
            ket_order[code] = []
        ket = r["cat2"] or ""
        if ket not in ket_map[code]:
            ket_map[code][ket] = {"pendidikan": 0.0, "tunjangan": 0.0, "penelitian": 0.0}
            ket_order[code].append(ket)
        cat1 = (r["cat1"] or "").lower()
        amt  = float(r["amount"] or 0)
        if   cat1 == "by pendidikan": ket_map[code][ket]["pendidikan"] += amt
        elif cat1 == "by tunjangan":  ket_map[code][ket]["tunjangan"]  += amt
        elif cat1 == "by penelitian": ket_map[code][ket]["penelitian"] += amt

    result = []
    for i, code in enumerate(codes, 1):
        b    = bud.get(code, {})
        p    = paid.get(code, {})
        rows = [{"keterangan": k, **ket_map[code][k]} for k in ket_order[code]]
        total = sum(r["pendidikan"] + r["tunjangan"] + r["penelitian"] for r in rows)
        result.append({
            "no": i,
            "siswa_code": code,
            **siswa_info[code],
            "total_pembayaran": total,
            "sisa_pendidikan": float(b.get("bud_p") or 0) - float(p.get("paid_p") or 0),
            "sisa_tunjangan":  float(b.get("bud_t") or 0) - float(p.get("paid_t") or 0),
            "sisa_penelitian": float(b.get("bud_r") or 0) - float(p.get("paid_r") or 0),
            "rows": rows,
        })
    result.sort(key=lambda x: (
        _JENJANG_SORT.get((x.get("jenjang") or "").upper(), 99),
        -float(x.get("total_pembayaran") or 0),
    ))
    for i, item in enumerate(result, 1):
        item["no"] = i
    return result


_SOURCE_MAP = {
    "AGRI": {"pr_source": "etf_agri"},
    "APP":  {"pr_source": "etf_app",  "paid_col": "tgl_Paid_APP"},
}


def get_days_of_pam(
    company_id: int,
    source: str = "AGRI",
    paid_only: bool = True,
    pam: str = None,
    nama: str = None,
    limit: int = 100,
    offset: int = 0,
) -> dict:
    src       = _SOURCE_MAP.get(source.upper(), _SOURCE_MAP["AGRI"])
    pr_source = src["pr_source"]
    paid_col  = src.get("paid_col")

    conditions = [
        "pb.company_id = ?",
        "pb.pam IS NOT NULL",
        "pb.pam != ''",
        "pr.source = ?",
    ]
    params = [company_id, pr_source]

    if paid_only:
        if paid_col:
            conditions.append(f'pb."{paid_col}" IS NULL')
        else:
            conditions.append("pr.tanggal_bayar IS NULL")
    if pam:
        conditions.append("pb.pam LIKE ?")
        params.append(f"%{pam}%")
    if nama:
        conditions.append("LOWER(s.nama) LIKE ?")
        params.append(f"%{nama.lower()}%")

    where = " AND ".join(conditions)
    base_from = f"""
        FROM payment_beasiswa pb
        LEFT JOIN siswa s
               ON s.company_id = pb.company_id AND s.code = pb.siswa_code
        JOIN pam_records pr
               ON pr.pam_no = pb.pam AND pr.company_id = pb.company_id
        WHERE {where}
    """

    conn  = get_conn()
    total = conn.execute(f"SELECT COUNT(*) {base_from}", params).fetchone()[0]
    rows  = [dict(r) for r in conn.execute(
        f"""SELECT pb.id, pb.siswa_code, s.nama,
                  pb.pam AS pam_no,
                  pb.cat1, pb.cat2, pb.perusahaan, pb.pillar,
                  pb.amount, pb.tanggal,
                  pb.tgl_pengajuan, pb.tgl_receive,
                  pb.tgl_pa, pb.tgl_final,
                  pb.tgl_retur, pb.tgl_final6, pb.tgl_proses,
                  pb.SLA_Date_1_LL, pb.SLA_Date_2_HT, pb.SLA_Date_3_YK,
                  pb.SLA_Date_4_AK, pb.SLA_Date_5_PD, pb.SLA_Date_6_C2,
                  pb.SLA_Date_7_MSIG,
                  pb."tgl_A-GS_APP", pb."tgl_A-HJK_APP",
                  pb.tgl_ASPIRO_APP, pb.tgl_Paid_APP
           {base_from}
           ORDER BY pb.tanggal DESC
           LIMIT ? OFFSET ?""",
        params + [limit, offset]
    ).fetchall()]
    conn.close()
    return {"rows": rows, "total": total}


def get_days_of_pam_candidates(company_id: int) -> list:
    """Lightweight SELECT for autocomplete — only 3 fields, DISTINCT rows."""
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        """SELECT DISTINCT
               pb.pam        AS pam_no,
               pb.siswa_code,
               s.nama
           FROM payment_beasiswa pb
           LEFT JOIN siswa s
                  ON s.company_id = pb.company_id AND s.code = pb.siswa_code
           WHERE pb.company_id = ?
             AND pb.pam IS NOT NULL AND pb.pam != ''
           ORDER BY pb.pam""",
        (company_id,)
    ).fetchall()]
    conn.close()
    return rows


def bulk_update_dates(ids: list, dates: dict, company_id: int) -> dict:
    _ALLOWED = {
        "tanggal", "tgl_pengajuan", "tgl_receive", "tgl_pa", "tgl_final",
        "tgl_retur", "tgl_final6", "tgl_proses",
        "SLA_Date_1_LL", "SLA_Date_2_HT", "SLA_Date_3_YK",
        "SLA_Date_4_AK", "SLA_Date_5_PD", "SLA_Date_6_C2", "SLA_Date_7_MSIG",
        "tgl_A-GS_APP", "tgl_A-HJK_APP", "tgl_ASPIRO_APP", "tgl_Paid_APP",
    }
    fields = [(k, v if v else None) for k, v in dates.items() if k in _ALLOWED]
    if not fields:
        return {"ok": False, "pesan": "Tidak ada field tanggal yang valid."}
    if not ids:
        return {"ok": False, "pesan": "Tidak ada baris yang dipilih."}
    # Quote column names to handle names containing hyphens (e.g. tgl_A-GS_APP)
    set_clause   = ", ".join(f'"{k}"=?' for k, _ in fields)
    vals         = [v for _, v in fields]
    placeholders = ",".join("?" * len(ids))
    conn = get_conn()
    cur  = conn.execute(
        f"UPDATE payment_beasiswa SET {set_clause}"
        f" WHERE id IN ({placeholders}) AND company_id=?",
        vals + list(ids) + [company_id]
    )
    updated = cur.rowcount
    conn.commit()
    conn.close()
    return {"ok": True, "updated": updated,
            "pesan": f"{updated} baris berhasil diperbarui."}


def update_pam_and_application(pam_id: int, pam_data: dict,
                                app_data: dict, company_id: int) -> dict:
    conn = get_conn()
    pam = conn.execute(
        "SELECT pam_no FROM pam_records WHERE id=? AND company_id=?", (pam_id, company_id)
    ).fetchone()
    if not pam:
        conn.close()
        return {"ok": False, "pesan": "PAM record tidak ditemukan."}

    old_pam_no = pam["pam_no"]
    new_pam_no = (pam_data.get("pam_no") or "").strip() or None
    pam_no_changed = new_pam_no and new_pam_no != old_pam_no

    # Validate new pam_no not already used by another record
    if pam_no_changed:
        conflict = conn.execute(
            "SELECT id FROM pam_records WHERE pam_no=? AND company_id=? AND id!=?",
            (new_pam_no, company_id, pam_id)
        ).fetchone()
        if conflict:
            conn.close()
            return {"ok": False, "pesan": f"No. PAM '{new_pam_no}' sudah digunakan oleh PAM lain."}

    _PAM_FIELDS = ("keterangan", "requestors_name", "total_amount", "due_date",
                   "pam_date", "tanggal_bayar")
    f, p = [], []
    if pam_no_changed:
        f.append("pam_no=?"); p.append(new_pam_no)
    for k in _PAM_FIELDS:
        if k in pam_data and pam_data[k] is not None:
            f.append(f"{k}=?"); p.append(pam_data[k])
    if f:
        conn.execute(
            f"UPDATE pam_records SET {','.join(f)}, updated_at=? WHERE id=? AND company_id=?",
            p + [_ts(), pam_id, company_id]
        )

    # Cascade pam_no rename to all referencing tables
    if pam_no_changed:
        conn.execute(
            "UPDATE payment_beasiswa SET pam=? WHERE pam=? AND company_id=?",
            (new_pam_no, old_pam_no, company_id)
        )
        for tbl in ("etf_pa", "app_pa", "sml_pa", "setf_pa"):
            conn.execute(
                f"UPDATE {tbl} SET nomor_pam=? WHERE nomor_pam=? AND company_id=?",
                (new_pam_no, old_pam_no, company_id)
            )

    # Update payment_application notes only
    effective_pam = new_pam_no if pam_no_changed else old_pam_no
    if app_data.get("notes") is not None:
        conn.execute(
            "UPDATE payment_application SET notes=? "
            "WHERE memo_id IN (SELECT memo_id FROM payment_beasiswa WHERE pam=? AND memo_id IS NOT NULL) "
            "AND company_id=?",
            (app_data["notes"], effective_pam, company_id)
        )

    conn.commit()
    conn.close()
    msg = "PAM berhasil diperbarui."
    if pam_no_changed:
        msg = f"No. PAM diubah dari '{old_pam_no}' → '{new_pam_no}' dan semua referensi sudah diperbarui."
    return {"ok": True, "pesan": msg}


def get_draft_payment_detail(payment_id: int, company_id: int) -> dict | None:
    conn = get_conn()
    pb = conn.execute(
        """SELECT pb.*, s.nama, s.bank, s.norek, s.namarek
           FROM payment_beasiswa pb
           LEFT JOIN siswa s ON s.company_id = pb.company_id AND s.code = pb.siswa_code
           WHERE pb.id=? AND pb.company_id=?""",
        (payment_id, company_id)
    ).fetchone()
    if not pb:
        conn.close()
        return None
    result = dict(pb)
    if result.get("pam"):
        pam = conn.execute(
            "SELECT * FROM pam_records WHERE pam_no=? AND company_id=?",
            (result["pam"], company_id)
        ).fetchone()
        result["pam_record"] = dict(pam) if pam else None
    else:
        result["pam_record"] = None
    if result.get("memo_id"):
        app = conn.execute(
            "SELECT * FROM payment_application WHERE memo_id=? AND company_id=?",
            (result["memo_id"], company_id)
        ).fetchone()
        result["payment_application"] = dict(app) if app else None
    else:
        result["payment_application"] = None
    conn.close()
    return result


def update_draft_and_linked(payment_id: int, pb_data: dict,
                             pam_data: dict, app_data: dict,
                             company_id: int) -> dict:
    conn = get_conn()
    pb = conn.execute(
        "SELECT pam, memo_id FROM payment_beasiswa WHERE id=? AND company_id=?",
        (payment_id, company_id)
    ).fetchone()
    if not pb:
        conn.close()
        return {"ok": False, "pesan": "Payment tidak ditemukan."}
    _PB = ("cat1", "cat2", "tanggal", "amount", "pillar", "perusahaan",
           "tgl_pengajuan", "tgl_receive", "tgl_pa", "tgl_final")
    f, p = [], []
    for k in _PB:
        if k in pb_data and pb_data[k] is not None:
            f.append(f"{k}=?"); p.append(pb_data[k])
    if f:
        conn.execute(
            f"UPDATE payment_beasiswa SET {','.join(f)} WHERE id=? AND company_id=?",
            p + [payment_id, company_id]
        )
    if pb["pam"] and pam_data:
        _PAM = ("keterangan", "requestors_name", "total_amount", "due_date", "pam_date")
        f2, p2 = [], []
        for k in _PAM:
            if k in pam_data and pam_data[k] is not None:
                f2.append(f"{k}=?"); p2.append(pam_data[k])
        if f2:
            conn.execute(
                f"UPDATE pam_records SET {','.join(f2)}, updated_at=? WHERE pam_no=? AND company_id=?",
                p2 + [_ts(), pb["pam"], company_id]
            )
    if pb["memo_id"] and app_data:
        _APP = ("submitted_at", "target_payment_date", "actual_payment_date", "notes", "status")
        fa, pa = [], []
        for k in _APP:
            if k in app_data and app_data[k] is not None:
                fa.append(f"{k}=?"); pa.append(app_data[k])
        if fa:
            conn.execute(
                f"UPDATE payment_application SET {','.join(fa)} WHERE memo_id=? AND company_id=?",
                pa + [pb["memo_id"], company_id]
            )
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Data payment berhasil diperbarui."}


def export_pam_excel(pam_id: int, company_id: int) -> bytes:
    import io, openpyxl
    from openpyxl.styles import Font, Alignment
    from datetime import datetime as dt

    conn = get_conn()
    r = conn.execute(
        "SELECT * FROM pam_records WHERE id=? AND company_id=?", (pam_id, company_id)
    ).fetchone()
    conn.close()
    if not r:
        raise ValueError("PAM tidak ditemukan.")

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "PAM NEW"

    bold = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLMNOPQ", [1,18,4,4,4,22,22,4,8,4,14,14,4,4,6,18,4]):
        ws.column_dimensions[col].width = w

    # Row 1 — Title
    ws.merge_cells("A1:Q1")
    ws["A1"] = "PAYMENT APPROVAL MEMO"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Rows 4-8 — Header info (left side)
    for row, label in [(4,"PAM No."),(5,"Date"),(6,"Requestor’s Name"),(7,"Department"),(8,"Company")]:
        ws.cell(row, 2, label).font = bold
        ws.cell(row, 5, ":")
    ws.cell(4, 6, r["pam_no"])
    # format pam_date
    try:    ws.cell(5, 6, dt.strptime(r["pam_date"], "%Y-%m-%d")); ws.cell(5,6).number_format = "DD-MMM-YY"
    except: ws.cell(5, 6, r["pam_date"] or "")
    ws.cell(6, 6, r["requestors_name"] or "")
    ws.cell(7, 6, "-")
    ws.cell(8, 6, r["pt"] or "")

    # Rows 4-6 — Header info (right side)
    for row, label in [(4,"Cost Center"),(5,"GL Account"),(6,"SO / SC")]:
        ws.cell(row, 12, label).font = bold
        ws.cell(row, 15, ":")
    ws.cell(4, 16, r["cost_center"] or "")
    ws.cell(5, 16, r["gl_account"] or "")

    # Row 10-11 — Business Unit
    ws.cell(10, 2, "Bussiness Unit").font = bold
    ws.cell(11, 7, "  Upstream")
    ws.cell(11, 11, "  Downstream")
    ws.cell(11, 14, "V")          # Corporate selected
    ws.cell(11, 15, "  Corporate")

    # Rows 13-16 — Type of Request
    ws.cell(13, 2, "Type of Request").font = bold
    ws.cell(14, 7, "  Downpayment to vendor")
    ws.cell(15, 5, "V")           # Invoice Payment selected
    ws.cell(15, 7, "  Invoice Payment – Non PO Invoice")
    ws.cell(16, 7, "  Employee Advance/ Reimbursement (Fund Transfer)")

    # Rows 18-22 — Invoice Information
    ws.cell(18, 2, "Invoice Information").font = bold
    for row, label, val in [
        (19, "Vendor Name", "Terlampir"),
        (20, "Invoice/ Memorandum Number", "-"),
        (22, "Expected Due Date", None),
    ]:
        ws.cell(row, 7, label)
        ws.cell(row, 8, ":")
        if val is not None:
            ws.cell(row, 9, val)
    ws.cell(21, 7, "Invoice Amount")
    ws.cell(21, 8, ":")
    ws.cell(21, 9, float(r["total_amount"] or 0))
    ws.cell(21, 9).number_format = "#,##0"
    try:    ws.cell(22, 9, dt.strptime(r["due_date"], "%Y-%m-%d")); ws.cell(22,9).number_format = "DD-MMM-YY"
    except: ws.cell(22, 9, r["due_date"] or "")

    # Rows 24-27 — Vendor Bank Account Details
    ws.cell(24, 2, "Vendor Bank Account Details").font = bold
    for row, label in [(25,"Bank Account Name"),(26,"Bank Name"),(27,"Bank Account Number")]:
        ws.cell(row, 4, label)
        ws.cell(row, 8, ":")
        ws.cell(row, 9, "Terlampir")

    # Rows 29, 35 — Request by
    ws.cell(29, 2, "Request by").font = bold
    ws.cell(35, 2, r["requestors_name"] or "")

    # Rows 37, 42 — Approved by
    ws.cell(37, 2, "Approved by").font = bold
    approvers = [a.strip() for a in (r["keterangan"] or "").split(",")]
    if approvers:      ws.cell(42, 2, approvers[0])
    if len(approvers) > 1: ws.cell(42, 7, approvers[1])

    # Row 43 — Checked by QA
    ws.cell(43, 2, "Checked by (QA)").font = bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_memo_pdf(memo_id: int, company_id: int, company_name: str) -> bytes:
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    memo = get_memo_detail(memo_id, company_id)
    if not memo:
        raise ValueError("Memo tidak ditemukan.")

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles   = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(company_name.upper(), styles["Title"]))
    elements.append(Paragraph("PAYMENT APPROVAL MEMO", styles["Heading2"]))
    elements.append(Spacer(1, 0.3*cm))

    info = [
        ["Nomor Memo:", memo["memo_number"]],
        ["Tanggal:",    memo["tanggal"] or ""],
        ["Status:",     memo["status"].upper()],
        ["Dibuat oleh:", memo["created_by"] or ""],
    ]
    if memo.get("approved_by"):
        info.append(["Disetujui oleh:", memo["approved_by"]])
    info_tbl = Table(info, colWidths=[4*cm, 12*cm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 0.5*cm))

    if memo.get("notes"):
        elements.append(Paragraph(f"Keterangan: {memo['notes']}", styles["Normal"]))
        elements.append(Spacer(1, 0.3*cm))

    header = [["No", "Penerima / Keterangan", "Vendor", "Rekening", "Amount (Rp)"]]
    rows   = header
    total  = 0
    for i, item in enumerate(memo["items"], 1):
        rows.append([
            i,
            item.get("description", ""),
            item.get("vendor", ""),
            item.get("bank_account", ""),
            f"{item['amount']:,.0f}",
        ])
        total += item["amount"]
    rows.append(["", "", "", "TOTAL", f"{total:,.0f}"])

    col_widths = [1*cm, 6*cm, 4*cm, 4*cm, 3*cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1a56db")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ALIGN",       (4,0), (4,-1), "RIGHT"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
        ("BACKGROUND",  (0,-1), (-1,-1), colors.HexColor("#f3f4f6")),
        ("FONTNAME",    (0,-1), (-1,-1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 1.5*cm))

    ttd = [["Dibuat oleh", "", "Disetujui oleh", "", "Diketahui oleh"],
           ["", "", "", "", ""],
           ["", "", "", "", ""],
           ["", "", "", "", ""],
           ["(________________)", "", "(________________)", "", "(_______________)"],
           ["Finance Staff", "", "Finance Manager", "", "Direktur"]]
    ttd_tbl = Table(ttd, colWidths=[3.5*cm, 0.5*cm, 3.5*cm, 0.5*cm, 3.5*cm])
    ttd_tbl.setStyle(TableStyle([
        ("FONTSIZE",  (0,0), (-1,-1), 8),
        ("ALIGN",     (0,0), (-1,-1), "CENTER"),
        ("FONTNAME",  (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME",  (0,-1), (-1,-1), "Helvetica"),
    ]))
    elements.append(ttd_tbl)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def get_fiori_list(search: str = "", bulan: str = "", tahun: str = "") -> list:
    conn = get_conn()
    sql  = "SELECT * FROM fiori_pa WHERE 1=1"
    params = []
    if search:
        sql += " AND (no_pa LIKE ? OR nama_vendor LIKE ? OR keterangan LIKE ?)"
        like = f"%{search}%"
        params += [like, like, like]
    if bulan:
        sql += " AND (terima_document LIKE ? OR approval_1 LIKE ?)"
        params += [f"%-{bulan}-%", f"%-{bulan}-%"]
    if tahun:
        sql += " AND (terima_document LIKE ? OR approval_1 LIKE ?)"
        params += [f"{tahun}-%", f"{tahun}-%"]
    sql += " ORDER BY terima_document DESC, id DESC"
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def get_sml_list(search: str = "", bulan: str = "", tahun: str = "") -> list:
    conn = get_conn()
    sql  = "SELECT * FROM sml_pa WHERE 1=1"
    params = []
    if search:
        sql += " AND (no_pa LIKE ? OR nama_vendor LIKE ? OR keterangan LIKE ?)"
        like = f"%{search}%"
        params += [like, like, like]
    if bulan:
        sql += " AND (terima_document LIKE ? OR approval_1 LIKE ?)"
        params += [f"%-{bulan}-%", f"%-{bulan}-%"]
    if tahun:
        sql += " AND (terima_document LIKE ? OR approval_1 LIKE ?)"
        params += [f"{tahun}-%", f"{tahun}-%"]
    sql += " ORDER BY terima_document DESC, id DESC"
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def bulk_update_fiori_dates(ids: list, dates: dict) -> dict:
    _ALLOWED = {"terima_document", "input_aspiro", "verifikasi_tax",
                "approval_1", "approval_2", "kirim_aspiro", "paid"}
    fields = [(k, v) for k, v in dates.items() if k in _ALLOWED and v]
    if not fields:
        return {"ok": False, "pesan": "Tidak ada tanggal yang diisi."}
    if not ids:
        return {"ok": False, "pesan": "Tidak ada baris yang dipilih."}
    set_clause   = ", ".join(f"{k}=?" for k, _ in fields)
    vals         = [v for _, v in fields]
    placeholders = ",".join("?" * len(ids))
    conn = get_conn()
    cur  = conn.execute(
        f"UPDATE fiori_pa SET {set_clause} WHERE id IN ({placeholders})",
        vals + list(ids)
    )
    updated = cur.rowcount
    conn.commit()
    conn.close()
    return {"ok": True, "updated": updated,
            "pesan": f"{updated} baris berhasil diperbarui."}


def update_fiori_status(record_id: int, new_status: str) -> dict:
    _ALLOWED = {"open", "on_process", "complete"}
    if new_status not in _ALLOWED:
        return {"ok": False, "pesan": "Status tidak valid."}
    conn = get_conn()
    r = conn.execute("SELECT id FROM fiori_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute("UPDATE fiori_pa SET status=? WHERE id=?", (new_status, record_id))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Status diubah ke '{new_status}'."}


def cancel_fiori_record(record_id: int) -> dict:
    conn = get_conn()
    r = conn.execute("SELECT no_pa FROM fiori_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute("DELETE FROM fiori_pa WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Record {r['no_pa']} dihapus."}


def update_sml_status(record_id: int, new_status: str) -> dict:
    _ALLOWED = {"open", "on_process", "complete"}
    if new_status not in _ALLOWED:
        return {"ok": False, "pesan": "Status tidak valid."}
    conn = get_conn()
    r = conn.execute("SELECT id FROM sml_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute("UPDATE sml_pa SET status=? WHERE id=?", (new_status, record_id))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Status diubah ke '{new_status}'."}


def cancel_sml_record(record_id: int) -> dict:
    conn = get_conn()
    r = conn.execute("SELECT no_pa FROM sml_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute("DELETE FROM sml_pa WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Record {r['no_pa']} dihapus."}


def bulk_update_sml_dates(ids: list, dates: dict) -> dict:
    _ALLOWED = {"terima_document", "input_aspiro", "verifikasi_tax",
                "approval_1", "approval_2", "kirim_aspiro", "paid"}
    fields = [(k, v) for k, v in dates.items() if k in _ALLOWED and v]
    if not fields:
        return {"ok": False, "pesan": "Tidak ada tanggal yang diisi."}
    if not ids:
        return {"ok": False, "pesan": "Tidak ada baris yang dipilih."}
    set_clause   = ", ".join(f"{k}=?" for k, _ in fields)
    vals         = [v for _, v in fields]
    placeholders = ",".join("?" * len(ids))
    conn = get_conn()
    cur  = conn.execute(
        f"UPDATE sml_pa SET {set_clause} WHERE id IN ({placeholders})",
        vals + list(ids)
    )
    updated = cur.rowcount
    conn.commit()
    conn.close()
    return {"ok": True, "updated": updated,
            "pesan": f"{updated} baris berhasil diperbarui."}


def get_energy_list(search: str = "", bulan: str = "", tahun: str = "",
                    status: str = "", source: str = "") -> list:
    conn = get_conn()
    sql  = "SELECT * FROM energy_pa WHERE 1=1"
    params = []
    if search:
        sql += " AND (no_pa LIKE ? OR nama_vendor LIKE ? OR keterangan LIKE ?)"
        like = f"%{search}%"
        params += [like, like, like]
    if bulan:
        sql += " AND (terima_document LIKE ? OR approval_1 LIKE ?)"
        params += [f"%-{bulan}-%", f"%-{bulan}-%"]
    if tahun:
        sql += " AND (terima_document LIKE ? OR approval_1 LIKE ?)"
        params += [f"{tahun}-%", f"{tahun}-%"]
    if status:
        sql += " AND status = ?"
        params += [status]
    sql += " ORDER BY terima_document DESC, id DESC"
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def update_energy_status(record_id: int, new_status: str) -> dict:
    _ALLOWED = {"open", "on_process", "complete"}
    if new_status not in _ALLOWED:
        return {"ok": False, "pesan": "Status tidak valid."}
    conn = get_conn()
    r = conn.execute("SELECT id FROM energy_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute("UPDATE energy_pa SET status=? WHERE id=?", (new_status, record_id))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Status diubah ke '{new_status}'."}


def cancel_energy_record(record_id: int) -> dict:
    conn = get_conn()
    r = conn.execute("SELECT no_pa FROM energy_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute("DELETE FROM energy_pa WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"Record {r['no_pa']} dihapus."}


def bulk_update_energy_dates(ids: list, dates: dict) -> dict:
    _ALLOWED = {"terima_document", "input_aspiro", "verifikasi_tax",
                "approval_1", "approval_2", "approval_3", "kirim_aspiro", "paid"}
    fields = [(k, v) for k, v in dates.items() if k in _ALLOWED and v]
    if not fields:
        return {"ok": False, "pesan": "Tidak ada tanggal yang diisi."}
    if not ids:
        return {"ok": False, "pesan": "Tidak ada baris yang dipilih."}
    set_clause   = ", ".join(f"{k}=?" for k, _ in fields)
    vals         = [v for _, v in fields]
    placeholders = ",".join("?" * len(ids))
    conn = get_conn()
    cur  = conn.execute(
        f"UPDATE energy_pa SET {set_clause} WHERE id IN ({placeholders})",
        vals + list(ids)
    )
    updated = cur.rowcount
    conn.commit()
    conn.close()
    return {"ok": True, "updated": updated,
            "pesan": f"{updated} baris berhasil diperbarui."}


def get_fiori_detail(record_id: int) -> dict | None:
    conn = get_conn()
    r = conn.execute("SELECT * FROM fiori_pa WHERE id=?", (record_id,)).fetchone()
    conn.close()
    return dict(r) if r else None


def update_fiori_record(record_id: int, data: dict) -> dict:
    _FIELDS = ("no_pa", "category", "keterangan", "categori_1",
               "nomor_vendor", "nama_vendor", "mata_uang", "dpp", "ppn", "total",
               "terima_document", "input_aspiro", "verifikasi_tax",
               "approval_1", "approval_2", "approval_3", "kirim_aspiro", "paid")
    fields = [(k, data[k]) for k in _FIELDS if k in data and data[k] is not None]
    # include null/empty string for date fields so user can clear them
    date_fields = ("terima_document", "input_aspiro", "verifikasi_tax",
                   "approval_1", "approval_2", "approval_3", "kirim_aspiro", "paid")
    for k in date_fields:
        if k in data and data[k] is None and k not in [f for f, _ in fields]:
            fields.append((k, None))
    if not fields:
        return {"ok": False, "pesan": "Tidak ada data untuk diupdate."}
    set_clause = ", ".join(f"{k}=?" for k, _ in fields)
    vals = [v for _, v in fields]
    conn = get_conn()
    r = conn.execute("SELECT id FROM fiori_pa WHERE id=?", (record_id,)).fetchone()
    if not r:
        conn.close()
        return {"ok": False, "pesan": "Record tidak ditemukan."}
    conn.execute(f"UPDATE fiori_pa SET {set_clause} WHERE id=?", vals + [record_id])
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Record berhasil diperbarui."}


# ── AGRI ETF-PA → PAM workflow ──────────────────────────────────────────────

def get_open_etf_pa_for_pam(company_id: int) -> list:
    """Return flat rows etf_pa status='open' dengan info siswa, untuk dipilih di Input AGRI."""
    conn = get_conn()
    rows = conn.execute(
        """SELECT
              p.id                AS pa_id,
              p.pa_number,
              p.tgl_payment_application,
              p.keterangan        AS pa_keterangan,
              p.status,
              s.code              AS siswa_code,
              s.nama,
              s.universitas       AS instansi,
              s.jenjang,
              s.angkatan,
              l.jenis_pembayaran,
              l.semester,
              l.jumlah_pembayaran,
              l.id                AS line_id
           FROM etf_pa p
           JOIN etf_pa_lines l ON l.pa_id = p.id
           JOIN siswa s ON s.id = l.student_id
           WHERE p.company_id = ? AND p.status = 'open'
           ORDER BY p.pa_number, l.id""",
        (company_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_pam_from_etf_pa(company_id: int, company_code: str,
                            pam_date: str, pa_ids: list, keterangan: str = "") -> dict:
    """
    Buat pam_records dari etf_pa yang dipilih.
    - Create pam_records (source='etf_agri')
    - Update etf_pa.nomor_pam = pam_no, status='on_process'
    """
    if not pa_ids:
        return {"ok": False, "pesan": "Pilih minimal 1 PA."}

    conn = get_conn()

    # Validate semua pa_ids milik company dan status='open'
    ph   = ",".join("?" * len(pa_ids))
    rows = conn.execute(
        f"SELECT id, status FROM etf_pa WHERE id IN ({ph}) AND company_id=?",
        pa_ids + [company_id]
    ).fetchall()
    if len(rows) != len(pa_ids):
        conn.close()
        return {"ok": False, "pesan": "Beberapa PA tidak ditemukan."}
    not_open = [dict(r)["id"] for r in rows if dict(r)["status"] != "open"]
    if not_open:
        conn.close()
        return {"ok": False, "pesan": f"PA {not_open} bukan status open."}

    # Hitung total
    total = conn.execute(
        f"""SELECT COALESCE(SUM(l.jumlah_pembayaran),0)
            FROM etf_pa_lines l WHERE l.pa_id IN ({ph})""",
        pa_ids
    ).fetchone()[0]

    # Generate PAM number
    now   = datetime.now()
    year  = now.strftime("%Y")
    month = now.strftime("%m")
    pam_no = generate_pam_number(company_id, company_code, year, month, conn)

    ts = _ts()
    conn.execute(
        """INSERT INTO pam_records
           (company_id, pam_no, pam_date, gl_account, cost_center, pt,
            requestors_name, keterangan, total_amount, due_date, pillar,
            status, source, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,'AGRI','open','etf_agri',?)""",
        (company_id, pam_no, pam_date,
         config.PAM_DEFAULT_GL if hasattr(config, 'PAM_DEFAULT_GL') else "",
         "", company_code,
         config.PAM_DEFAULT_REQUESTOR if hasattr(config, 'PAM_DEFAULT_REQUESTOR') else "",
         keterangan, float(total),
         _add_one_month(pam_date), ts)
    )

    # Update etf_pa
    conn.execute(
        f"UPDATE etf_pa SET nomor_pam=?, status='on_process', updated_at=? WHERE id IN ({ph}) AND company_id=?",
        [pam_no, ts] + pa_ids + [company_id]
    )

    conn.commit()
    conn.close()
    return {"ok": True, "pam_no": pam_no, "pesan": f"PAM {pam_no} berhasil dibuat dari {len(pa_ids)} PA."}


def bulk_complete_pams(company_id: int, pams: list, tanggal_bayar: str) -> dict:
    conn = get_conn()
    success_count = 0
    for pam_no in set(pams):
        row = conn.execute("SELECT id FROM pam_records WHERE pam_no=? AND company_id=?", (pam_no, company_id)).fetchone()
        if row:
            res = set_pam_complete_cascade(row["id"], tanggal_bayar, company_id)
            if res.get("ok"):
                success_count += 1
    conn.close()
    if success_count > 0:
        return {"ok": True, "pesan": f"Berhasil memproses {success_count} PAM menjadi complete."}
    return {"ok": False, "pesan": "Tidak ada PAM yang berhasil di-update."}


def set_pam_complete_cascade(pam_id: int, tanggal_bayar: str, company_id: int) -> dict:
    """
    Set tanggal_bayar + status='complete' di pam_records, cascade ke tabel PA dan
    payment_beasiswa sesuai source PAM (etf_agri → etf_pa via nomor_pam;
    beasiswa/klaim/others → payment_beasiswa + semua PA tables via etf_pa_line_id).
    """
    if not tanggal_bayar:
        return {"ok": False, "pesan": "Tanggal bayar wajib diisi."}
    conn = get_conn()
    pam = conn.execute(
        "SELECT id, pam_no, source, pillar FROM pam_records WHERE id=? AND company_id=?",
        (pam_id, company_id)
    ).fetchone()
    if not pam:
        conn.close()
        return {"ok": False, "pesan": "PAM tidak ditemukan."}

    pam = dict(pam)
    ts  = _ts()

    # 1. Update pam_records
    conn.execute(
        "UPDATE pam_records SET tanggal_bayar=?, status='complete', updated_at=? WHERE id=?",
        (tanggal_bayar, ts, pam_id)
    )

    pam_no = pam["pam_no"]
    source = pam.get("source") or ""

    # 2. Cascade ke etf_pa jika source='etf_agri' (PA-only AGRI flow, no payment_beasiswa)
    if source == "etf_agri":
        conn.execute(
            """UPDATE etf_pa SET tanggal_bayar=?, status='complete', updated_at=?
               WHERE nomor_pam=? AND company_id=?""",
            (tanggal_bayar, ts, pam_no, company_id)
        )

    # 3. Cascade ke payment_beasiswa + PA tables untuk beasiswa iPay flow
    else:
        pillar = pam.get("pillar") or ""
        if pillar == "ADVANCE":
            conn.execute(
                "UPDATE payment_beasiswa SET status='paid' WHERE pam=? AND company_id=?",
                (pam_no, company_id)
            )
            pa_header_status = "paid"
        else:
            paid_col = _BEASISWA_PAID_COL.get(pillar)
            if paid_col:
                conn.execute(
                    f'UPDATE payment_beasiswa SET status=\'complete\', "{paid_col}"=? '
                    f'WHERE pam=? AND company_id=?',
                    (tanggal_bayar, pam_no, company_id)
                )
            else:
                conn.execute(
                    "UPDATE payment_beasiswa SET status='complete' WHERE pam=? AND company_id=?",
                    (pam_no, company_id)
                )
            pa_header_status = "complete"
        line_ids = [
            r[0] for r in conn.execute(
                "SELECT DISTINCT etf_pa_line_id FROM payment_beasiswa "
                "WHERE pam=? AND company_id=? AND etf_pa_line_id IS NOT NULL",
                (pam_no, company_id)
            ).fetchall()
        ]
        if line_ids:
            ph = ",".join("?" * len(line_ids))
            for lines_tbl, pa_tbl in [
                ("etf_pa_lines",    "etf_pa"),
                ("app_pa_lines",    "app_pa"),
                ("sml_pa_lines",    "sml_pa"),
                ("energy_pa_lines", "energy_pa"),
                ("setf_pa_lines",   "setf_pa"),
            ]:
                conn.execute(
                    f"""UPDATE {pa_tbl} SET tanggal_bayar=?, status=?, updated_at=?
                        WHERE id IN (
                            SELECT DISTINCT pa_id FROM {lines_tbl} WHERE id IN ({ph})
                        ) AND company_id=?""",
                    [tanggal_bayar, pa_header_status, ts] + line_ids + [company_id]
                )

    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Tgl Paid disimpan, PAM selesai."}


def realize_advance_payment(payment_id: int, realized_amount, tgl_realisasi: str,
                            company_id: int) -> dict:
    """Realization of Advance payment: correct amount and close pillar cascade.

    Updates the target row's `amount`, `realized_amount`, `tgl_realisasi`, `status='complete'`;
    once every `payment_beasiswa` row sharing that `pam` is `complete`, updates
    `pam_records.pillar` from `"ADVANCE"` to the row's own `payment_beasiswa.pillar`.

    Returns {"ok": True, "pesan": "...", "selisih": <float>} on success, or
            {"ok": False, "pesan": "..."} on validation failure.
    """
    if not tgl_realisasi:
        return {"ok": False, "pesan": "Tanggal realisasi wajib diisi."}
    try:
        realized_amount = float(realized_amount)
    except (TypeError, ValueError):
        return {"ok": False, "pesan": "Realized amount tidak valid."}
    if realized_amount <= 0:
        return {"ok": False, "pesan": "Realized amount harus > 0."}

    conn = get_conn()
    row = conn.execute(
        "SELECT id, pam, pillar, status, advance_amount, etf_pa_line_id FROM payment_beasiswa "
        "WHERE id=? AND company_id=?",
        (payment_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "Payment tidak ditemukan."}
    if row["advance_amount"] is None:
        conn.close()
        return {"ok": False, "pesan": "Baris ini bukan payment Advance."}
    if row["status"] != "paid":
        conn.close()
        return {"ok": False, "pesan": "Payment belum berstatus 'paid', tidak bisa direalisasi."}

    pam_no        = row["pam"]
    target_pillar = row["pillar"]
    selisih       = row["advance_amount"] - realized_amount
    ts            = _ts()

    conn.execute(
        """UPDATE payment_beasiswa
           SET realized_amount=?, tgl_realisasi=?, amount=?, status='complete'
           WHERE id=?""",
        (realized_amount, tgl_realisasi, realized_amount, payment_id)
    )

    remaining = conn.execute(
        "SELECT COUNT(*) FROM payment_beasiswa WHERE pam=? AND company_id=? AND status != 'complete'",
        (pam_no, company_id)
    ).fetchone()[0]
    if remaining == 0:
        conn.execute(
            "UPDATE pam_records SET pillar=?, updated_at=? WHERE pam_no=? AND company_id=?",
            (target_pillar, ts, pam_no, company_id)
        )

    # Update the originating PA line's amount, and close the PA header once every
    # payment_beasiswa row tracing back to that PA header is 'complete'.
    pa_line_id = row["etf_pa_line_id"]
    if pa_line_id:
        for lines_tbl, pa_tbl in [
            ("etf_pa_lines",    "etf_pa"),
            ("app_pa_lines",    "app_pa"),
            ("sml_pa_lines",    "sml_pa"),
            ("energy_pa_lines", "energy_pa"),
            ("setf_pa_lines",   "setf_pa"),
        ]:
            updated = conn.execute(
                f"UPDATE {lines_tbl} SET jumlah_pembayaran=? WHERE id=?",
                (realized_amount, pa_line_id)
            )
            if updated.rowcount:
                pa_id_row = conn.execute(
                    f"SELECT pa_id FROM {lines_tbl} WHERE id=?", (pa_line_id,)
                ).fetchone()
                pa_id = pa_id_row[0]
                sibling_line_ids = [
                    r[0] for r in conn.execute(
                        f"SELECT id FROM {lines_tbl} WHERE pa_id=?", (pa_id,)
                    ).fetchall()
                ]
                ph2 = ",".join("?" * len(sibling_line_ids))
                still_open = conn.execute(
                    f"""SELECT COUNT(*) FROM payment_beasiswa
                        WHERE etf_pa_line_id IN ({ph2}) AND status != 'complete'""",
                    sibling_line_ids
                ).fetchone()[0]
                if still_open == 0:
                    conn.execute(
                        f"UPDATE {pa_tbl} SET status='complete', updated_at=? WHERE id=?",
                        (ts, pa_id)
                    )
                break  # found the table this line belongs to, no need to try the rest

    conn.commit()
    conn.close()
    return {"ok": True, "pesan": "Realisasi tersimpan.", "selisih": selisih}


def check_pam_no_exists(company_id: int, pam_no: str) -> dict:
    conn = get_conn()
    row = conn.execute(
        "SELECT 1 FROM pam_records WHERE pam_no=? AND company_id=?",
        (pam_no, company_id)
    ).fetchone()
    conn.close()
    return {"ok": True, "exists": row is not None}


def get_pam_beasiswa_lines(pam_id: int, company_id: int) -> list | None:
    conn = get_conn()
    pam_no_row = conn.execute(
        "SELECT pam_no FROM pam_records WHERE id = ? AND company_id = ?",
        (pam_id, company_id),
    ).fetchone()
    if not pam_no_row:
        conn.close()
        return None
    pam_no = pam_no_row["pam_no"]
    rows = [dict(r) for r in conn.execute(
        """SELECT
               pb.id,
               pb.siswa_code,
               s.nama,
               pb.cat1,
               pb.cat2,
               pb.amount,
               pb.tgl_pengajuan,
               pb.tgl_receive,
               pb.tgl_pa,
               pb.tgl_final,
               pb.SLA_Date_1_LL,
               pb.SLA_Date_2_HT,
               pb.SLA_Date_3_YK,
               pb.SLA_Date_4_AK,
               pb.SLA_Date_5_PD,
               pb.SLA_Date_6_C2,
               pb.SLA_Date_7_MSIG
           FROM payment_beasiswa pb
           LEFT JOIN siswa s
                  ON s.company_id = pb.company_id AND s.code = pb.siswa_code
           WHERE pb.pam = ? AND pb.company_id = ?
           ORDER BY s.nama ASC""",
        (pam_no, company_id),
    ).fetchall()]
    conn.close()
    return rows
