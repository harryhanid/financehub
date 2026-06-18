# modules/etf_payment_application/service.py
from datetime import datetime
from database import get_conn

VALID_TABS = {"agri", "app", "sml", "setf"}

# Maps tab → (pa_table, lines_table, pa_number_prefix, pam_prefix)
_TAB_CFG = {
    "agri":  ("etf_pa",  "etf_pa_lines",  "ETF",  "ETF"),
    "app":   ("app_pa",  "app_pa_lines",  "APP",  "APP"),
    "sml":   ("sml_pa",  "sml_pa_lines",  "SML",  "SML"),
    "setf":  ("setf_pa", "setf_pa_lines", "SETF", "SETF"),
}


def _tbls(tab: str) -> tuple:
    return _TAB_CFG.get(tab, _TAB_CFG["agri"])


def _ts():
    return datetime.now().isoformat(timespec="seconds")


def _latest_ipk(siswa_row: dict) -> float:
    for i in range(10, 0, -1):
        val = siswa_row.get(f"ipk_sem{i}") or 0
        if val:
            return float(val)
    return 0.0


def get_siswa_autocomplete(company_id: int, q: str) -> list:
    conn = get_conn()
    rows = conn.execute(
        """SELECT id, code, nama, jenjang, angkatan, program, fakultas,
                  universitas, status,
                  ipk_sem1, ipk_sem2, ipk_sem3, ipk_sem4, ipk_sem5,
                  ipk_sem6, ipk_sem7, ipk_sem8, ipk_sem9, ipk_sem10
           FROM siswa
           WHERE company_id=? AND (nama LIKE ? OR code LIKE ?)
           ORDER BY nama LIMIT 20""",
        (company_id, f"%{q}%", f"%{q}%")
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["ipk_terakhir"] = _latest_ipk(d)
        result.append(d)
    return result


def _gen_pa_number(company_id: int, conn, pa_tbl: str, prefix: str) -> str:
    year  = datetime.now().strftime("%Y")
    count = conn.execute(
        f"SELECT COUNT(*) FROM {pa_tbl} WHERE company_id=?", (company_id,)
    ).fetchone()[0]
    return f"PA/{prefix}/{count + 1:03d}/{year}"


def _gen_nomor_pam(company_id: int, conn, pa_tbl: str, prefix: str) -> str:
    now  = datetime.now()
    mm   = now.strftime("%m")
    yyyy = now.strftime("%Y")
    count = conn.execute(
        f"""SELECT COUNT(*) FROM {pa_tbl}
            WHERE company_id=? AND nomor_pam IS NOT NULL
            AND strftime('%Y-%m', created_at)=?""",
        (company_id, f"{yyyy}-{mm}")
    ).fetchone()[0]
    return f"{count + 1:03d}-{prefix}-{mm}-{yyyy}"


def get_pa_list(company_id: int, tab: str = "agri") -> list:
    pa_tbl, lines_tbl, *_ = _tbls(tab)
    conn = get_conn()
    rows = conn.execute(
        f"""SELECT p.*,
                  COUNT(l.id)                            AS jml_siswa,
                  COALESCE(SUM(l.jumlah_pembayaran), 0)  AS total_bayar
           FROM {pa_tbl} p
           LEFT JOIN {lines_tbl} l ON l.pa_id = p.id
           WHERE p.company_id=?
           GROUP BY p.id
           ORDER BY p.created_at DESC""",
        (company_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pa_flat(company_id: int, tab: str = "agri", status_filter: str = "",
                nama: str = "", jenjang: str = "", program: str = "",
                angkatan: str = "", jenis: str = "", pam: str = "",
                bulan_pa: str = "", tahun_pa: str = "") -> list:
    pa_tbl, lines_tbl, *_ = _tbls(tab)
    conn = get_conn()
    extra_where = ""
    params: list = [company_id]
    if status_filter == "active":
        extra_where = " AND LOWER(p.status) IN ('open', 'on_process')"
    elif status_filter:
        extra_where = " AND LOWER(p.status)=?"
        params.append(status_filter.lower())
    if nama:
        extra_where += " AND LOWER(s.nama) LIKE ?"
        params.append(f"%{nama.lower()}%")
    if jenjang:
        extra_where += " AND s.jenjang=?"
        params.append(jenjang)
    if program:
        extra_where += " AND s.program=?"
        params.append(program)
    if angkatan:
        extra_where += " AND s.angkatan=?"
        params.append(angkatan)
    if jenis:
        extra_where += " AND l.jenis_pembayaran=?"
        params.append(jenis)
    if pam:
        extra_where += " AND p.nomor_pam LIKE ?"
        params.append(f"%{pam}%")
    if bulan_pa:
        extra_where += " AND strftime('%m', p.tgl_payment_application)=?"
        params.append(bulan_pa.zfill(2))
    if tahun_pa:
        extra_where += " AND strftime('%Y', p.tgl_payment_application)=?"
        params.append(tahun_pa)
    rows = conn.execute(
        f"""SELECT
                  s.code             AS student_code,
                  p.id               AS pa_id,
                  s.nama,
                  s.status           AS status_pb,
                  s.universitas      AS instansi_pendidikan,
                  s.angkatan         AS angkatan_etf,
                  s.angkatan_kuliah,
                  s.jenjang          AS jenjang_pendidikan,
                  s.program          AS program_beasiswa,
                  s.fakultas,
                  s.prodi            AS program_studi,
                  p.tgl_payment_application,
                  p.tgl_surat_pengajuan,
                  l.jenis_pembayaran,
                  l.semester,
                  l.tahun_ajaran,
                  l.ipk_sem_sebelumnya,
                  l.jumlah_pembayaran,
                  p.doc_received_by_educ,
                  p.received_pa_from_educ,
                  p.checked_by_fincon,
                  p.approved_by_htj_1,
                  p.send_pa_back_to_educ,
                  p.pa_received_by_po_fin,
                  p.approval_by_htj_2,
                  p.nomor_pam,
                  p.tanggal_bayar,
                  p.keterangan,
                  p.status,
                  p.pa_number,
                  l.id               AS line_id
           FROM {pa_tbl} p
           JOIN {lines_tbl} l ON l.pa_id = p.id
           JOIN siswa s ON s.id = l.student_id
           WHERE p.company_id=?{extra_where}
           ORDER BY p.created_at DESC, l.id ASC""",
        params
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def bulk_update_pa(pa_ids: list, field: str, value: str, company_id: int, tab: str = "agri") -> dict:
    ALLOWED_FIELDS = {
        "tgl_payment_application", "tgl_surat_pengajuan",
        "doc_received_by_educ", "received_pa_from_educ", "checked_by_fincon",
        "approved_by_htj_1", "send_pa_back_to_educ", "pa_received_by_po_fin",
        "approval_by_htj_2", "tanggal_bayar", "nomor_pam", "status",
    }
    if field not in ALLOWED_FIELDS:
        return {"ok": False, "pesan": f"Field '{field}' tidak diizinkan."}
    if not pa_ids:
        return {"ok": False, "pesan": "Tidak ada PA yang dipilih."}

    pa_tbl, *_ = _tbls(tab)
    conn = get_conn()
    placeholders = ",".join("?" * len(pa_ids))
    extra_set = ", status='complete'" if field == "tanggal_bayar" and value else ""
    conn.execute(
        f"UPDATE {pa_tbl} SET {field}=?{extra_set}, updated_at=? WHERE id IN ({placeholders}) AND company_id=?",
        [value, _ts()] + list(pa_ids) + [company_id]
    )
    count = conn.execute(
        f"SELECT COUNT(*) FROM {pa_tbl} WHERE id IN ({placeholders}) AND company_id=?",
        list(pa_ids) + [company_id]
    ).fetchone()[0]
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"{count} PA berhasil diupdate ({field} = '{value}')."}


def get_draft_siswa(company_id: int, q: str, tab: str = "agri") -> list:
    pa_tbl, lines_tbl, *_ = _tbls(tab)
    conn = get_conn()
    rows = conn.execute(
        f"""SELECT DISTINCT s.id, s.code, s.nama, s.jenjang, s.universitas
            FROM siswa s
            JOIN {lines_tbl} l ON l.student_id = s.id
            JOIN {pa_tbl} p ON p.id = l.pa_id
            WHERE p.company_id = ? AND LOWER(p.status) = 'open'
              AND (s.nama LIKE ? OR s.code LIKE ?)
            ORDER BY s.nama
            LIMIT 20""",
        (company_id, f"%{q}%", f"%{q}%")
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_draft_lines_for_siswa(company_id: int, siswa_id: int, tab: str = "agri") -> list:
    pa_tbl, lines_tbl, *_ = _tbls(tab)
    conn = get_conn()
    rows = conn.execute(
        f"""SELECT l.id AS line_id, l.pa_id, p.pa_number,
                   l.jenis_pembayaran, l.semester, l.tahun_ajaran, l.jumlah_pembayaran,
                   p.tgl_surat_pengajuan,
                   p.doc_received_by_educ,
                   p.tgl_payment_application
            FROM {lines_tbl} l
            JOIN {pa_tbl} p ON p.id = l.pa_id
            WHERE p.company_id = ? AND LOWER(p.status) = 'open'
              AND l.student_id = ?
            ORDER BY p.created_at DESC""",
        (company_id, siswa_id)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def export_pa_excel(company_id: int, tab: str = "agri",
                    sf: str = "", nama: str = "", jenjang: str = "",
                    program: str = "", angkatan: str = "", jenis: str = "",
                    pam: str = "", bulan_pa: str = "", tahun_pa: str = "") -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = get_pa_flat(company_id, tab, sf, nama, jenjang, program, angkatan, jenis, pam, bulan_pa, tahun_pa)
    tab_label = tab.upper()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{tab_label} Payment Application"

    headers = [
        "No. PA", "ID Siswa", "Nama PB", "Status PB", "Instansi", "Angkatan ETF",
        "Angkatan Kuliah", "Jenjang", "Program", "Fakultas", "Program Studi",
        "Tgl PA", "Tgl Surat Pengajuan", "Jenis Bayar", "Semester", "Tahun Ajaran",
        "IPK Sblmnya", "Jumlah (Rp)", "Doc Recv Educ", "Recv PA Educ",
        "Checked Fincon", "Approved HTj", "Send Back Educ", "PA Recv PO Fin",
        "Approval HTj", "Nomor PAM", "Tgl Bayar", "Keterangan", "Status",
    ]
    fields = [
        "pa_number", "student_code", "nama", "status_pb", "instansi_pendidikan",
        "angkatan_etf", "angkatan_kuliah", "jenjang_pendidikan", "program_beasiswa",
        "fakultas", "program_studi", "tgl_payment_application", "tgl_surat_pengajuan",
        "jenis_pembayaran", "semester", "tahun_ajaran", "ipk_sem_sebelumnya",
        "jumlah_pembayaran", "doc_received_by_educ", "received_pa_from_educ",
        "checked_by_fincon", "approved_by_htj_1", "send_pa_back_to_educ",
        "pa_received_by_po_fin", "approval_by_htj_2", "nomor_pam", "tanggal_bayar",
        "keterangan", "status",
    ]

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    op_fill  = PatternFill("solid", fgColor="FFFBEB")
    thin     = Side(style="thin", color="D1D5DB")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 32

    for ri, r in enumerate(rows, 2):
        is_op = r.get("status") == "on_process"
        for ci, f in enumerate(fields, 1):
            val = r.get(f, "")
            if val is None: val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(f == "keterangan"))
            if is_op:
                cell.fill = op_fill

    col_widths = [16, 12, 22, 9, 22, 11, 12, 9, 12, 18, 20,
                  12, 14, 14, 12, 12, 10, 14, 12, 12, 12, 12, 12, 12, 12, 16, 12, 40, 10]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_pa_header(pa_id: int, company_id: int, tab: str = "agri") -> dict | None:
    pa_tbl, *_ = _tbls(tab)
    conn = get_conn()
    row = conn.execute(
        f"SELECT * FROM {pa_tbl} WHERE id=? AND company_id=?", (pa_id, company_id)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_pa_lines(pa_id: int, company_id: int, tab: str = "agri") -> list:
    pa_tbl, lines_tbl, *_ = _tbls(tab)
    conn = get_conn()
    rows = conn.execute(
        f"""SELECT l.*,
                   s.nama, s.code AS siswa_code, s.status AS status_pb,
                   s.universitas AS instansi_pendidikan,
                   s.angkatan AS angkatan_etf,
                   s.jenjang AS jenjang_pendidikan,
                   s.program AS program_beasiswa,
                   s.fakultas
            FROM {lines_tbl} l
            JOIN siswa s ON s.id = l.student_id
            JOIN {pa_tbl} p ON p.id = l.pa_id
            WHERE l.pa_id=? AND p.company_id=?
            ORDER BY l.id""",
        (pa_id, company_id)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_pa(company_id: int, header: dict, lines: list, tab: str = "agri") -> dict:
    if not lines:
        return {"ok": False, "pesan": "Minimal 1 siswa harus diisi."}

    pa_tbl, lines_tbl, pa_prefix, _ = _tbls(tab)
    conn = get_conn()

    sids = [line.get("student_id") for line in lines]
    ph = ",".join("?" * len(sids))
    found = {r[0] for r in conn.execute(
        f"SELECT id FROM siswa WHERE id IN ({ph}) AND company_id=?",
        sids + [company_id]
    ).fetchall()}
    missing = [sid for sid in sids if sid not in found]
    if missing:
        conn.close()
        return {"ok": False, "pesan": f"Siswa ID {missing[0]} tidak ditemukan."}

    pa_number = _gen_pa_number(company_id, conn, pa_tbl, pa_prefix)
    ts = _ts()
    cur = conn.execute(
        f"""INSERT INTO {pa_tbl}
            (company_id, pa_number, tgl_payment_application, tgl_surat_pengajuan,
             keterangan, doc_received_by_educ, received_pa_from_educ, status, created_at)
            VALUES (?,?,?,?,?,?,?,'open',?)""",
        (company_id, pa_number,
         header.get("tgl_payment_application", ""),
         header.get("tgl_surat_pengajuan", ""),
         header.get("keterangan", ""),
         header.get("doc_received_by_educ", ""),
         header.get("received_pa_from_educ", ""),
         ts)
    )
    pa_id = cur.lastrowid

    for line in lines:
        conn.execute(
            f"""INSERT INTO {lines_tbl}
                (pa_id, student_id, jenis_pembayaran, semester,
                 tahun_ajaran, ipk_sem_sebelumnya, jumlah_pembayaran)
                VALUES (?,?,?,?,?,?,?)""",
            (pa_id,
             line.get("student_id"),
             line.get("jenis_pembayaran", ""),
             line.get("semester", ""),
             line.get("tahun_ajaran", ""),
             line.get("ipk_sem_sebelumnya") or 0,
             line.get("jumlah_pembayaran") or 0)
        )

    conn.commit()
    conn.close()
    return {"ok": True, "pa_id": pa_id, "pa_number": pa_number,
            "pesan": f"Payment Application {pa_number} berhasil dibuat."}


def update_pa(pa_id: int, company_id: int, data: dict, tab: str = "agri") -> dict:
    pa_tbl, lines_tbl, _, pam_prefix = _tbls(tab)
    conn = get_conn()
    row = conn.execute(
        f"SELECT id, status, nomor_pam FROM {pa_tbl} WHERE id=? AND company_id=?",
        (pa_id, company_id)
    ).fetchone()
    if not row:
        conn.close()
        return {"ok": False, "pesan": "PA tidak ditemukan."}

    new_status = data.get("status", row["status"])
    if data.get("tanggal_bayar"):
        new_status = "complete"
    nomor_pam = data.get("nomor_pam") or row["nomor_pam"]

    if new_status == "on_process" and not nomor_pam:
        pa_tbl2, _, _, pam_pfx = _tbls(tab)
        nomor_pam = _gen_nomor_pam(company_id, conn, pa_tbl2, pam_pfx)

    conn.execute(
        f"""UPDATE {pa_tbl} SET
             tgl_payment_application = ?,
             tgl_surat_pengajuan     = ?,
             doc_received_by_educ    = ?,
             received_pa_from_educ   = ?,
             checked_by_fincon       = ?,
             approved_by_htj_1       = ?,
             send_pa_back_to_educ    = ?,
             pa_received_by_po_fin   = ?,
             approval_by_htj_2       = ?,
             nomor_pam               = ?,
             tanggal_bayar           = ?,
             keterangan              = ?,
             status                  = ?,
             updated_at              = ?
            WHERE id=? AND company_id=?""",
        (data.get("tgl_payment_application", ""),
         data.get("tgl_surat_pengajuan", ""),
         data.get("doc_received_by_educ", ""),
         data.get("received_pa_from_educ", ""),
         data.get("checked_by_fincon", ""),
         data.get("approved_by_htj_1", ""),
         data.get("send_pa_back_to_educ", ""),
         data.get("pa_received_by_po_fin", ""),
         data.get("approval_by_htj_2", ""),
         nomor_pam,
         data.get("tanggal_bayar", ""),
         data.get("keterangan", ""),
         new_status,
         _ts(), pa_id, company_id)
    )
    line_id = data.get("line_id")
    if line_id:
        conn.execute(
            f"""UPDATE {lines_tbl} SET
                 jenis_pembayaran   = ?,
                 semester           = ?,
                 tahun_ajaran       = ?,
                 ipk_sem_sebelumnya = ?,
                 jumlah_pembayaran  = ?
                WHERE id=? AND pa_id=?""",
            (data.get("jenis_pembayaran", ""),
             data.get("semester", ""),
             data.get("tahun_ajaran", ""),
             data.get("ipk_sem_sebelumnya") or 0,
             data.get("jumlah_pembayaran") or 0,
             line_id, pa_id)
        )
    conn.commit()
    conn.close()

    msg = "PA berhasil diupdate."
    if new_status == "on_process" and nomor_pam and not row["nomor_pam"]:
        msg = f"PA pindah ke On Process. Nomor PAM: {nomor_pam}"
    return {"ok": True, "pesan": msg, "nomor_pam": nomor_pam}


def delete_pa(pa_id: int, company_id: int, tab: str = "agri") -> dict:
    pa_tbl, lines_tbl, *_ = _tbls(tab)
    conn = get_conn()
    pa = conn.execute(
        f"SELECT id, pa_number FROM {pa_tbl} WHERE id=? AND company_id=?",
        (pa_id, company_id)
    ).fetchone()
    if not pa:
        conn.close()
        return {"ok": False, "pesan": "PA tidak ditemukan."}

    linked = conn.execute(
        f"""SELECT COUNT(*) FROM payment_beasiswa pb
            JOIN {lines_tbl} l ON l.id = pb.etf_pa_line_id
            WHERE l.pa_id=?""",
        (pa_id,)
    ).fetchone()[0]
    if linked:
        conn.close()
        return {
            "ok": False,
            "pesan": f"PA ini tidak dapat dihapus — sudah terhubung ke {linked} record payment. "
                     "Hapus payment terkait terlebih dahulu di modul Payment Memo.",
        }

    conn.execute(f"DELETE FROM {lines_tbl} WHERE pa_id=?", (pa_id,))
    conn.execute(f"DELETE FROM {pa_tbl} WHERE id=? AND company_id=?", (pa_id, company_id))
    conn.commit()
    conn.close()
    return {"ok": True, "pesan": f"PA {pa['pa_number']} berhasil dihapus."}


def get_pa_summary(company_id: int) -> list:
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM pa_summary WHERE company_id=? ORDER BY pa_number",
        (company_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]
