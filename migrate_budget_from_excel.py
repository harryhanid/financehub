# migrate_budget_from_excel.py
# One-time migration: exported Budget_Dashboard2 Google Sheet (.xlsx) -> finance_hub.db
#
# NOTE on deviation from the original task brief: the brief's reference code routed
# Master_Budget/Realisasi rows through modules.budget.service.create_budget() /
# create_realisasi(). Both of those functions unconditionally auto-generate their own
# primary key (create_budget() via generate_budget_id(), create_realisasi() via a
# uuid4-based trx_id) and ignore any id/trx_id present in the payload. That means rows
# imported that way get a *new* id that never matches the source spreadsheet's
# Budget_ID/Trx_ID, so every Realisasi/Carryover_Logs row referencing the original
# Budget_ID would fail to resolve its parent budget. This script instead preserves the
# spreadsheet's original IDs via direct inserts (mirroring the same INSERT statements
# and validation used in create_budget()/create_realisasi()), so cross-sheet references
# stay intact. Company/mm/yy validation and deadline computation still reuse the
# service layer (VALID_COMPANIES, compute_deadline) rather than duplicating that logic.
import sys
import os
from datetime import datetime
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "app"))

from database import get_conn
from modules.budget.service import VALID_COMPANIES, compute_deadline


def _clean(v):
    if v is None:
        return ""
    return str(v).strip()


def _clean_date(v):
    if v is None or v == "":
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return str(v).strip()[:10]


def _ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


def migrate_budget_from_excel(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    errors = []
    budgets_imported = 0
    realisasi_imported = 0
    carryover_imported = 0

    if "Master_Budget" in wb.sheetnames:
        ws = wb["Master_Budget"]
        conn = get_conn()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            budget_id = _clean(row[0])
            company = _clean(row[3])
            try:
                mm = int(row[1])
                yy = int(row[2])
            except (TypeError, ValueError):
                errors.append(f"Master_Budget row {budget_id}: bulan/tahun tidak valid")
                continue
            if company not in VALID_COMPANIES:
                errors.append(f"Master_Budget row {budget_id}: company harus PO atau TF")
                continue
            if not (1 <= mm <= 12):
                errors.append(f"Master_Budget row {budget_id}: bulan harus 1-12")
                continue
            existing = conn.execute(
                "SELECT 1 FROM budget_master WHERE id = ?", (budget_id,)
            ).fetchone()
            if existing:
                errors.append(f"Master_Budget row {budget_id}: budget ID sudah ada, dilewati")
                continue
            try:
                amount = float(row[10] or 0)
            except (TypeError, ValueError):
                errors.append(f"Master_Budget row {budget_id}: jumlah tidak valid")
                continue
            deadline = _clean_date(row[11]) if len(row) > 11 else None
            if not deadline:
                deadline = compute_deadline(mm, yy)
            conn.execute(
                """INSERT INTO budget_master
                   (id, mm, yy, company, dept, gl_account, gl_description, budget_category,
                    activity, description, amount, deadline, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (budget_id, mm, yy, company, _clean(row[4]), _clean(row[5]), _clean(row[6]),
                 _clean(row[7]), _clean(row[8]), _clean(row[9]), amount, deadline, _ts())
            )
            budgets_imported += 1
        conn.commit()
        conn.close()

    if "Realisasi" in wb.sheetnames:
        ws = wb["Realisasi"]
        conn = get_conn()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            trx_id = _clean(row[0])
            budget_id = _clean(row[1])
            budget_row = conn.execute(
                "SELECT * FROM budget_master WHERE id = ?", (budget_id,)
            ).fetchone()
            if not budget_row:
                errors.append(f"Realisasi row {trx_id}: budget ID tidak ditemukan: {budget_id}")
                continue
            existing = conn.execute(
                "SELECT 1 FROM budget_realisasi WHERE trx_id = ?", (trx_id,)
            ).fetchone()
            if existing:
                errors.append(f"Realisasi row {trx_id}: trx ID sudah ada, dilewati")
                continue
            try:
                amount = float(row[11] or 0)
            except (TypeError, ValueError):
                errors.append(f"Realisasi row {trx_id}: jumlah tidak valid")
                continue
            if amount <= 0:
                errors.append(f"Realisasi row {trx_id}: jumlah harus lebih dari 0")
                continue
            tanggal = _clean_date(row[12]) or datetime.now().strftime("%Y-%m-%d")
            description = _clean(row[10]) or budget_row["description"]
            conn.execute(
                """INSERT INTO budget_realisasi
                   (trx_id, budget_id, mm, yy, company, dept, gl_account, gl_description,
                    budget_category, activity, description, amount, tanggal_realisasi, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (trx_id, budget_id, budget_row["mm"], budget_row["yy"], budget_row["company"],
                 budget_row["dept"], budget_row["gl_account"], budget_row["gl_description"],
                 budget_row["budget_category"], budget_row["activity"], description,
                 amount, tanggal, _ts())
            )
            realisasi_imported += 1
        conn.commit()
        conn.close()

    if "Carryover_Logs" in wb.sheetnames:
        ws = wb["Carryover_Logs"]
        conn = get_conn()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            budget_id = _clean(row[0])
            existing = conn.execute(
                "SELECT 1 FROM budget_master WHERE id = ?", (budget_id,)
            ).fetchone()
            if not existing:
                errors.append(f"Carryover_Logs row for {budget_id}: budget not found, skipped")
                continue
            conn.execute(
                """INSERT INTO budget_carryover_logs
                   (budget_id, requested_by, request_date, status, approval_date,
                    extension_months, reason, approved_by, type, additional_amount)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (budget_id, _clean(row[1]), _clean_date(row[2]), _clean(row[3]) or "Pending",
                 _clean_date(row[4]), row[5] or 12, _clean(row[6]), _clean(row[7]),
                 _clean(row[8]) or "Carryover", row[9] or 0)
            )
            carryover_imported += 1
        conn.commit()
        conn.close()

    return {
        "budgets": budgets_imported, "realisasi": realisasi_imported,
        "carryover_logs": carryover_imported, "errors": errors,
    }


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python migrate_budget_from_excel.py <path-to-exported.xlsx>")
        sys.exit(1)
    result = migrate_budget_from_excel(sys.argv[1])
    print(f"Budgets imported: {result['budgets']}")
    print(f"Realisasi imported: {result['realisasi']}")
    print(f"Carryover logs imported: {result['carryover_logs']}")
    if result["errors"]:
        print(f"\n{len(result['errors'])} errors:")
        for e in result["errors"]:
            print(f"  - {e}")
