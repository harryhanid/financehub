"""Sahabat ETF — 'REPORT PER [tanggal]' Excel export.

Replicates the legacy manually-maintained Book3.xlsx report layout using
openpyxl, computed entirely from modules.sahabat_etf.service.build_report_data()
(no Excel formulas — every value is a precomputed static number).
"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_INDO_MONTHS = ["", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
                "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]

_NAVY       = "1F4E79"
_DARK_NAVY  = "002060"
_GREEN      = "3C7D22"
_LIGHT_BLUE = "E3F6FD"
_WHITE      = "FFFFFF"

_BAND_A = "DEEAF1"
_BAND_B = "EBF5FB"

_FILL_NAVY       = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_FILL_DARK_NAVY  = PatternFill(start_color=_DARK_NAVY, end_color=_DARK_NAVY, fill_type="solid")
_FILL_GREEN      = PatternFill(start_color=_GREEN, end_color=_GREEN, fill_type="solid")
_FILL_LIGHT_BLUE = PatternFill(start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid")
_FILL_BAND_A     = PatternFill(start_color=_BAND_A, end_color=_BAND_A, fill_type="solid")
_FILL_BAND_B     = PatternFill(start_color=_BAND_B, end_color=_BAND_B, fill_type="solid")

_FONT_TITLE    = Font(bold=True, size=25)
_FONT_UNIT     = Font(bold=False, size=15)
_FONT_HEADER   = Font(bold=True, size=15, color=_WHITE)
_FONT_SECTION  = Font(bold=True, size=15, color=_WHITE)
_FONT_SUBTOTAL = Font(bold=True, size=15)
_FONT_NORMAL   = Font(bold=False, size=15)
_FONT_PILLAR   = Font(bold=True, size=15)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
_ALIGN_RIGHT  = Alignment(horizontal="right")

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_THICK  = Side(style="thick")

_MONEY_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'
_PCT_FMT   = "0.0%"

_COL = {
    "desc": "B", "periode_start": "C", "periode_dash": "D", "periode_end": "E", "pillar": "F",
    "cur_plafon": "G", "cur_klaim": "H", "cur_dibayar": "I", "cur_saldo": "J",
    "cum_plafon": "K", "cum_klaim": "L", "cum_dibayar": "M", "cum_saldo": "N",
    "catatan": "O",
}
_METRIC_COLS = {
    "cur": ["cur_plafon", "cur_klaim", "cur_dibayar", "cur_saldo"],
    "cum": ["cum_plafon", "cum_klaim", "cum_dibayar", "cum_saldo"],
}
_METRIC_KEYS = ["plafon", "klaim", "dibayar", "saldo"]

_ROW_COLUMNS = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
_LABEL_COLUMNS = ["B", "C", "D", "E", "F"]
# Vertical grid: thick brackets the "TAHUN X" column block (G-J), medium brackets
# the whole table (B left edge, O right edge) — matches Book3.xlsx's grid exactly.
_GRID_LEFT = {"B": _MEDIUM, "G": _THICK}
_GRID_RIGHT = {"J": _THICK, "N": _THICK, "O": _MEDIUM}


def _grid_border(col, top=_THIN, bottom=_THIN):
    return Border(left=_GRID_LEFT.get(col, _THIN), right=_GRID_RIGHT.get(col, _THIN),
                  top=top, bottom=bottom)


def _apply_row_grid(ws, row, columns=_ROW_COLUMNS, top=_THIN, bottom=_THIN, fill=None):
    for col in columns:
        c = ws[f"{col}{row}"]
        c.border = _grid_border(col, top=top, bottom=bottom)
        if fill:
            c.fill = fill


def _set(ws, coord, value, font=None, fill=None, align=None, border=None, numfmt=None):
    c = ws[coord]
    c.value = value
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    if numfmt: c.number_format = numfmt
    return c


def _today_title() -> str:
    now = datetime.now()
    return f"{now.day} {_INDO_MONTHS[now.month]} {now.year}"


def _set_column_widths(ws):
    widths = {"A": 2.38, "B": 44.84, "C": 10.54, "D": 1.54, "E": 10.84,
              "F": 21.0, "G": 13.92, "O": 16.69}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_title(ws, row):
    ws.merge_cells(f"B{row}:N{row}")
    _set(ws, f"B{row}", f"REPORT PER {_today_title()}", font=_FONT_TITLE, align=_ALIGN_LEFT)
    _set(ws, f"O{row}", "(Rp, Dalam Jutaan)", font=_FONT_UNIT, align=_ALIGN_RIGHT)
    return row + 2


def _write_header(ws, row, report_year):
    sub_row = row + 1
    for col in _ROW_COLUMNS:
        fill = _FILL_GREEN if col in ("K", "L", "M", "N") else _FILL_NAVY
        top_bottom = _THIN if col in ("G", "H", "I", "J", "K", "L", "M", "N") else None
        _apply_row_grid(ws, row, columns=[col], top=_MEDIUM, bottom=top_bottom, fill=fill)
        _apply_row_grid(ws, sub_row, columns=[col], top=None, bottom=_MEDIUM, fill=fill)

    ws.merge_cells(f"B{row}:B{sub_row}")
    _set(ws, f"B{row}", "Deskripsi", font=_FONT_HEADER, align=_ALIGN_CENTER)
    ws.merge_cells(f"C{row}:E{sub_row}")
    _set(ws, f"C{row}", "Periode", font=_FONT_HEADER, align=_ALIGN_CENTER)
    ws.merge_cells(f"F{row}:F{sub_row}")
    _set(ws, f"F{row}", "Pillar", font=_FONT_HEADER, align=_ALIGN_CENTER)
    ws.merge_cells(f"G{row}:J{row}")
    _set(ws, f"G{row}", f"TAHUN {report_year}", font=_FONT_HEADER, align=_ALIGN_CENTER)
    ws.merge_cells(f"K{row}:N{row}")
    _set(ws, f"K{row}", f"S/D TAHUN {report_year}", font=_FONT_HEADER, align=_ALIGN_CENTER)
    ws.merge_cells(f"O{row}:O{sub_row}")
    _set(ws, f"O{row}", "Catatan", font=_FONT_HEADER, align=_ALIGN_CENTER)

    for label, col in (("Plafon", "G"), ("Klaim", "H"), ("Dibayar", "I"), ("Saldo", "J")):
        _set(ws, f"{col}{sub_row}", label, font=_FONT_HEADER, align=_ALIGN_CENTER)
    for label, col in (("Plafon", "K"), ("Klaim", "L"), ("Dibayar", "M"), ("Saldo", "N")):
        _set(ws, f"{col}{sub_row}", label, font=_FONT_HEADER, align=_ALIGN_CENTER)
    return row + 2


def _write_metrics_row(ws, row, metrics, font, fill=None, top=_THIN, bottom=_THIN):
    for window in ("cur", "cum"):
        for i, metric_key in enumerate(_METRIC_KEYS):
            col = _COL[_METRIC_COLS[window][i]]
            value = metrics[window][metric_key] / 1_000_000
            _set(ws, f"{col}{row}", value, font=font, fill=fill,
                 align=_ALIGN_CENTER, border=_grid_border(col, top=top, bottom=bottom),
                 numfmt=_MONEY_FMT)


def _write_total_row(ws, row, label, total, fill):
    ws.merge_cells(f"B{row}:F{row}")
    _apply_row_grid(ws, row, columns=_LABEL_COLUMNS, fill=fill)
    _apply_row_grid(ws, row, columns=["O"], fill=fill)
    _set(ws, f"B{row}", label, font=_FONT_SECTION, align=_ALIGN_LEFT)
    _write_metrics_row(ws, row, total, font=_FONT_SECTION, fill=fill)
    return row + 1


def _write_section(ws, row, section):
    ws.merge_cells(f"B{row}:F{row}")
    _apply_row_grid(ws, row, fill=_FILL_NAVY)
    _set(ws, f"B{row}", section["label"], font=_FONT_SECTION, align=_ALIGN_LEFT)
    row += 1

    for group in section["groups"]:
        ws.merge_cells(f"B{row}:F{row}")
        _apply_row_grid(ws, row, columns=_LABEL_COLUMNS)
        _apply_row_grid(ws, row, columns=["O"])
        _set(ws, f"B{row}", group["label"], font=_FONT_SUBTOTAL, align=_ALIGN_LEFT)
        _write_metrics_row(ws, row, group["subtotal"], font=_FONT_SUBTOTAL)
        row += 1
        for siswa_row in group["siswa_rows"]:
            _apply_row_grid(ws, row, columns=_LABEL_COLUMNS)
            _apply_row_grid(ws, row, columns=["O"])
            _set(ws, f"B{row}", siswa_row["nama"], font=_FONT_NORMAL, align=_ALIGN_LEFT)
            _set(ws, f"F{row}", siswa_row["pillar"], font=_FONT_PILLAR, align=_ALIGN_CENTER)
            _write_metrics_row(ws, row, siswa_row, font=_FONT_NORMAL, fill=_FILL_LIGHT_BLUE)
            row += 1

    return _write_total_row(ws, row, f"TOTAL {section['label']}", section["total"], _FILL_NAVY)


def _write_recap_rows(ws, row, recap_rows, banded=False):
    for i, r in enumerate(recap_rows):
        fill = (_FILL_BAND_A if i % 2 == 0 else _FILL_BAND_B) if banded else None
        _apply_row_grid(ws, row, columns=_LABEL_COLUMNS, fill=fill)
        _apply_row_grid(ws, row, columns=["O"], fill=fill)
        _set(ws, f"B{row}", r["nama"], font=_FONT_NORMAL, align=_ALIGN_LEFT)
        _set(ws, f"F{row}", r["pillar"], font=_FONT_PILLAR, align=_ALIGN_CENTER)
        _write_metrics_row(ws, row, r, font=_FONT_NORMAL, fill=fill)
        row += 1
    return row


def _write_pillar_breakdown(ws, row, pillar_breakdown, grand_total):
    ws.merge_cells(f"B{row}:F{row}")
    _apply_row_grid(ws, row, fill=_FILL_NAVY)
    _set(ws, f"B{row}", "BREAKDOWN PILLAR", font=_FONT_SECTION, align=_ALIGN_LEFT)
    _write_metrics_row(ws, row, grand_total, font=_FONT_SECTION, fill=_FILL_NAVY)
    row += 1

    for i, p in enumerate(pillar_breakdown):
        band = _FILL_BAND_A if i % 2 == 0 else _FILL_BAND_B
        _apply_row_grid(ws, row, columns=_LABEL_COLUMNS, fill=band)
        _apply_row_grid(ws, row, columns=["O"], fill=band)
        _set(ws, f"B{row}", p["pillar_label"], font=_FONT_SUBTOTAL, align=_ALIGN_LEFT)
        _write_metrics_row(ws, row, p, font=_FONT_SUBTOTAL, fill=band)
        row += 1
        _apply_row_grid(ws, row, columns=_LABEL_COLUMNS, fill=band)
        _apply_row_grid(ws, row, columns=["O"], fill=band)
        _set(ws, f"B{row}", "      % to total", font=_FONT_NORMAL, align=_ALIGN_LEFT)
        for window in ("cur", "cum"):
            for i2, metric_key in enumerate(_METRIC_KEYS):
                col = _COL[_METRIC_COLS[window][i2]]
                pct_key = f"pct_{window}"
                _set(ws, f"{col}{row}", p[pct_key][metric_key], font=_FONT_NORMAL, fill=band,
                     align=_ALIGN_CENTER, border=_grid_border(col), numfmt=_PCT_FMT)
        row += 1
    return row


def build_report_workbook(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    _set_column_widths(ws)
    ws.sheet_format.defaultRowHeight = 20

    row = _write_title(ws, 1)
    for section in data["sections"]:
        row = _write_header(ws, row, data["report_year"])
        row = _write_section(ws, row, section)
        row = _write_recap_rows(ws, row, section["recap"])
        row += 1  # spacer before next section's header

    row = _write_total_row(ws, row, "TOTAL PENDIDIKAN & KESEHATAN", data["combined_total"], _FILL_DARK_NAVY)
    row = _write_recap_rows(ws, row, data["combined_recap"], banded=True)
    row = _write_total_row(ws, row, "GRAND TOTAL", data["grand_total"], _FILL_DARK_NAVY)
    row += 1
    row = _write_pillar_breakdown(ws, row, data["pillar_breakdown"], data["grand_total"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
