import io
import openpyxl
import pytest

_RP_FMT = '_-"Rp"* #,##0_-;\\-"Rp"* #,##0_-;_-"Rp"* "-"_-;_-@_-'


def _data():
    return {
        "pam_no": "PAM-TEST", "pam_date": "2026-06-15",
        "requestors_name": "Tester", "department": "-",
        "cost_center": "CC-001", "gl_account": "GL-001",
        "so_sc": "", "pt": "PT TEST",
        "bu_corporate": True, "bu_upstream": False, "bu_downstream": False,
        "type_invoice": True, "type_downpayment": False, "type_advance": False,
        "vendor_name": "Budi Santoso, Ali Rahman",
        "invoice_memo_no": "-",
        "total_amount": 5_000_000,
        "due_date": "2026-07-01",
        "bank_account_name": "Budi Santoso, Ali Rahman",
        "bank_name": "BNI, BCA",
        "bank_account_no": "111, 222",
        "approved_by_1": "Mgr", "approved_by_2": "Dir",
        "company_id": 1,
    }


def _payments():
    return [
        {"siswa_code": "S2A", "nama": "Ali", "bank": "BCA", "norek": "222",
         "namarek": "Ali Rahman", "jenjang": "S2", "amount": 2_000_000,
         "cat1": "By Pendidikan", "cat2": None, "tanggal": "2026-06-01"},
        {"siswa_code": "S1A", "nama": "Budi", "bank": "BNI", "norek": "111",
         "namarek": "Budi Santoso", "jenjang": "S1", "amount": 3_000_000,
         "cat1": "By Pendidikan", "cat2": None, "tanggal": "2026-06-01"},
    ]


def test_excel_pam_new_colons_center_aligned(monkeypatch):
    from app.modules.payment_memo import exports
    monkeypatch.setattr(exports, "get_pam_payments_detail", lambda *a: [])

    result = exports.export_pam_excel_custom(_data(), _payments())
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["PAM NEW"]

    for row in [4, 5, 6, 7, 8]:
        assert ws[f"E{row}"].alignment.horizontal == "center", f"E{row} colon not center"
    for row in [19, 20, 21, 22]:
        assert ws[f"H{row}"].alignment.horizontal == "center", f"H{row} colon not center"
    for row in [25, 26, 27]:
        assert ws[f"H{row}"].alignment.horizontal == "center", f"H{row} colon not center"


def test_excel_pam_new_vendor_bank_wrap_text(monkeypatch):
    from app.modules.payment_memo import exports
    monkeypatch.setattr(exports, "get_pam_payments_detail", lambda *a: [])

    result = exports.export_pam_excel_custom(_data(), _payments())
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["PAM NEW"]

    assert ws["I19"].alignment.wrap_text is True, "I19 vendor name: wrap_text harus True"
    assert ws["I25"].alignment.wrap_text is True, "I25 bank account name: wrap_text harus True"
    assert ws["I26"].alignment.wrap_text is True, "I26 bank name: wrap_text harus True"


def test_rangkuman_rp_accounting_format(monkeypatch):
    """Amount cells di Rangkuman PAM harus pakai Rp Accounting number format."""
    from app.modules.payment_memo import exports
    monkeypatch.setattr(exports, "get_pam_payments_detail", lambda *a: [])

    result = exports.export_pam_excel_custom(_data(), _payments())
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws2 = wb["Rangkuman PAM"]

    # Data row 1 at row 8 (header at rows 6-7), col 8 = amount
    assert ws2.cell(8, 8).number_format == _RP_FMT, "data row: bukan Rp Accounting format"
    # Total row at row 10 (2 data rows + total)
    assert ws2.cell(10, 8).number_format == _RP_FMT, "total row: bukan Rp Accounting format"


def test_detail_sheet_rp_accounting_format():
    """Amount cells di Detail PAM sheet harus pakai Rp Accounting format."""
    import openpyxl as _xl
    from app.modules.payment_memo.exports import _build_detail_sheet

    wb = _xl.Workbook()
    ws = wb.active
    pam = {"pam_no": "PAM-TEST", "pam_date": "2026-06-15",
           "cost_center": "CC", "gl_account": "GL"}
    detail = [{
        "no": 1, "siswa_code": "S001", "nama": "Budi", "angkatan": "2020",
        "jenjang": "S1", "program": "Teknik", "universitas": "UI",
        "fakultas": "FT", "bank": "BNI", "norek": "111", "namarek": "Budi",
        "total_pembayaran": 3_000_000.0, "sisa_pendidikan": 0.0,
        "sisa_tunjangan": 0.0, "sisa_penelitian": 0.0,
        "rows": [{"keterangan": "By Pendidikan",
                  "pendidikan": 3_000_000.0, "tunjangan": 0.0, "penelitian": 0.0}],
    }]

    _build_detail_sheet(ws, pam, detail)

    # Data row dimulai row 7, col 14 = total_pembayaran
    assert ws.cell(7, 14).number_format == _RP_FMT, "total_pembayaran: bukan Rp Accounting"
    # Col 11 = pendidikan amount
    assert ws.cell(7, 11).number_format == _RP_FMT, "pendidikan: bukan Rp Accounting"
    # Grand total row 8, col 14
    assert ws.cell(8, 14).number_format == _RP_FMT, "grand total: bukan Rp Accounting"


def test_pdf_detail_has_12_columns():
    """PDF Detail PAM table harus punya 12 kolom (tambah Jenjang Studi, sebelumnya 11)."""
    from app.modules.payment_memo.exports import _build_detail_pdf_table

    detail = [{
        "no": 1, "siswa_code": "S001", "nama": "Budi",
        "jenjang": "S2", "total_pembayaran": 5_000_000.0,
        "sisa_pendidikan": 0.0, "sisa_tunjangan": 0.0, "sisa_penelitian": 0.0,
        "norek": "111",
        "rows": [{"keterangan": "By Pendidikan",
                  "pendidikan": 5_000_000.0, "tunjangan": 0.0, "penelitian": 0.0}],
    }]

    table = _build_detail_pdf_table(detail)
    assert len(table._colWidths) == 12, f"Expected 12 cols, got {len(table._colWidths)}"


def test_pdf_detail_data_row_has_12_cells():
    """Setiap data row harus punya 12 cell (sama dengan jumlah kolom)."""
    from app.modules.payment_memo.exports import _build_detail_pdf_table

    detail = [{
        "no": 1, "siswa_code": "S001", "nama": "Budi",
        "jenjang": "S1", "total_pembayaran": 3_000_000.0,
        "sisa_pendidikan": 0.0, "sisa_tunjangan": 0.0, "sisa_penelitian": 0.0,
        "norek": "222",
        "rows": [{"keterangan": "By Pendidikan",
                  "pendidikan": 3_000_000.0, "tunjangan": 0.0, "penelitian": 0.0}],
    }]

    table = _build_detail_pdf_table(detail)
    # Row 0 = header, row 1 = first data row
    assert len(table._cellvalues[1]) == 12, f"Expected 12 cells in data row, got {len(table._cellvalues[1])}"
