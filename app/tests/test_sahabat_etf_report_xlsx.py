import io
import openpyxl
from modules.sahabat_etf.report_xlsx import build_report_workbook


def _empty_metrics():
    return {"plafon": 0.0, "klaim": 0.0, "dibayar": 0.0, "saldo": 0.0}


def _metrics(plafon, klaim, dibayar):
    return {"plafon": plafon, "klaim": klaim, "dibayar": dibayar, "saldo": plafon - dibayar}


def _sample_data():
    empty_total = {"cur": _empty_metrics(), "cum": _empty_metrics()}
    return {
        "report_year": 2026,
        "sections": [
            {"key": "PENDIDIKAN", "label": "PENDIDIKAN", "groups": [], "total": empty_total, "recap": []},
            {"key": "KESEHATAN", "label": "KESEHATAN", "groups": [], "total": empty_total, "recap": []},
        ],
        "combined_total": empty_total,
        "combined_recap": [],
        "grand_total": empty_total,
        "pillar_breakdown": [],
    }


def _sample_data_with_content():
    sd_row = {"nama": "Richard Widjaja", "pillar": "SETF",
              "cur": _metrics(45_000_000, 12_135_000, 12_135_000),
              "cum": _metrics(90_000_000, 24_270_000, 24_270_000)}
    sd_group = {"key": "SD", "label": "SD",
                "subtotal": {"cur": sd_row["cur"], "cum": sd_row["cum"]},
                "siswa_rows": [sd_row]}
    pend_total = {"cur": sd_row["cur"], "cum": sd_row["cum"]}
    empty_total = {"cur": _empty_metrics(), "cum": _empty_metrics()}
    return {
        "report_year": 2026,
        "sections": [
            {"key": "PENDIDIKAN", "label": "PENDIDIKAN", "groups": [sd_group],
             "total": pend_total, "recap": []},
            {"key": "KESEHATAN", "label": "KESEHATAN", "groups": [], "total": empty_total, "recap": []},
        ],
        "combined_total": empty_total,
        "combined_recap": [],
        "grand_total": empty_total,
        "pillar_breakdown": [],
    }


def test_build_report_workbook_writes_title_and_unit_note():
    xlsx_bytes = build_report_workbook(_sample_data())
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws["B1"].value.startswith("REPORT PER ")
    assert ws["O1"].value == "(Rp, Dalam Jutaan)"


def test_build_report_workbook_writes_column_header_labels():
    xlsx_bytes = build_report_workbook(_sample_data())
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws["B3"].value == "Deskripsi"
    assert ws["C3"].value == "Periode"
    assert ws["F3"].value == "Pillar"
    assert ws["G3"].value == "TAHUN 2026"
    assert ws["K3"].value == "S/D TAHUN 2026"
    assert ws["O3"].value == "Catatan"
    assert ws["G4"].value == "Plafon"
    assert ws["H4"].value == "Klaim"
    assert ws["I4"].value == "Dibayar"
    assert ws["J4"].value == "Saldo"


def test_build_report_workbook_writes_section_group_and_siswa_row():
    xlsx_bytes = build_report_workbook(_sample_data_with_content())
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    values = {}
    for row in ws.iter_rows(min_row=5, max_row=20):
        for cell in row:
            if cell.value is not None:
                values.setdefault(cell.row, {})[cell.column_letter] = cell.value

    all_b_values = [v.get("B") for v in values.values()]
    assert "PENDIDIKAN" in all_b_values
    assert "SD" in all_b_values
    assert "Richard Widjaja" in all_b_values
    assert "TOTAL PENDIDIKAN" in all_b_values

    richard_row = next(r for r, v in values.items() if v.get("B") == "Richard Widjaja")
    assert values[richard_row]["G"] == 45.0   # cur.plafon = 45,000,000 / 1,000,000
    assert values[richard_row]["K"] == 90.0   # cum.plafon = 90,000,000 / 1,000,000 (ensures cur/cum not swapped)
    assert values[richard_row]["F"] == "SETF"
