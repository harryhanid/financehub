import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.bank.service import get_bank_setf_rows


def classify_rows(rows: list) -> dict:
    penerimaan = {}
    bank = {}
    sahabat_etf = {}

    for r in rows:
        ket = (r["keterangan"] or "").strip()
        ket_lower = ket.lower()
        if ket_lower.startswith("setoran awal"):
            continue

        month_key = r["tanggal"][:7]
        label = ket if ket else "(Tanpa Keterangan)"
        jumlah = r["jumlah"] or 0

        if r["source"] == "pam":
            sahabat_etf.setdefault(label, {}).setdefault(month_key, 0)
            sahabat_etf[label][month_key] += jumlah
            continue

        if ket_lower == "bank admin & bunga":
            bank.setdefault("Bunga dan Admin", {}).setdefault(month_key, 0)
            if r["jenis"] == "pengeluaran":
                bank["Bunga dan Admin"][month_key] += jumlah
            else:
                bank["Bunga dan Admin"][month_key] -= jumlah
        elif r["jenis"] == "pemasukan":
            penerimaan.setdefault(label, {}).setdefault(month_key, 0)
            penerimaan[label][month_key] += jumlah
        else:
            bank.setdefault(label, {}).setdefault(month_key, 0)
            bank[label][month_key] += jumlah

    return {"penerimaan": penerimaan, "bank": bank, "sahabat_etf": sahabat_etf}


def month_range(rows: list, today=None) -> list:
    today = today or datetime.now()
    current_key = f"{today.year:04d}-{today.month:02d}"

    eligible = [
        r["tanggal"] for r in rows
        if not (r["keterangan"] or "").strip().lower().startswith("setoran awal")
    ]
    if not eligible:
        return [current_key]

    earliest = min(eligible)[:7]
    y, m = int(earliest[:4]), int(earliest[5:7])

    months = []
    while (y, m) <= (today.year, today.month):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


_MONTH_NAMES_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]
_MONTH_ABBR_ID = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
_NUM_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'
_FIRST_MONTH_COL = 3  # column C
_EMPTY_SECTION_FORMULA = "=SUM(A1)"  # column A is never written elsewhere on this sheet, so this reliably evaluates to 0


def _write_section(ws, start_row, title, data, months):
    n_months = len(months)
    total_col = _FIRST_MONTH_COL + n_months
    section_font = Font(name="Arial", bold=True)
    normal_font = Font(name="Arial")

    r = start_row
    ws.cell(r, 2, title).font = section_font
    r += 1

    label_rows = []
    for label, month_vals in data.items():
        for k, mkey in enumerate(months):
            val = month_vals.get(mkey, 0) / 1_000_000
            c = ws.cell(r, _FIRST_MONTH_COL + k, val)
            c.number_format = _NUM_FMT
            c.font = normal_font
        ws.cell(r, 2, label).font = normal_font
        first_letter = get_column_letter(_FIRST_MONTH_COL)
        last_letter = get_column_letter(_FIRST_MONTH_COL + n_months - 1)
        tc = ws.cell(r, total_col, f"=SUM({first_letter}{r}:{last_letter}{r})")
        tc.number_format = _NUM_FMT
        tc.font = normal_font
        label_rows.append(r)
        r += 1

    subtotal_row = r
    ws.cell(r, 2, "Subtotal").font = section_font
    for k in range(n_months):
        col_letter = get_column_letter(_FIRST_MONTH_COL + k)
        formula = f"=SUM({col_letter}{label_rows[0]}:{col_letter}{label_rows[-1]})" if label_rows else _EMPTY_SECTION_FORMULA
        c = ws.cell(r, _FIRST_MONTH_COL + k, formula)
        c.number_format = _NUM_FMT
        c.font = section_font
    total_letter = get_column_letter(total_col)
    if label_rows:
        formula = f"=SUM({total_letter}{label_rows[0]}:{total_letter}{label_rows[-1]})"
    else:
        formula = _EMPTY_SECTION_FORMULA
    c = ws.cell(r, total_col, formula)
    c.number_format = _NUM_FMT
    c.font = section_font

    return subtotal_row, r + 1


def build_laporan_mutasi_excel(company_id: int, today=None) -> bytes:
    today = today or datetime.now()
    rows = get_bank_setf_rows(company_id)
    classified = classify_rows(rows)
    months = month_range(rows, today)
    n_months = len(months)
    total_col = _FIRST_MONTH_COL + n_months

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ALL"

    title_font = Font(name="Arial", bold=True, size=14)
    subtitle_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", bold=True)
    section_font = Font(name="Arial", bold=True)
    saldo_fill = PatternFill("solid", fgColor="004E9A")
    saldo_font = Font(name="Arial", bold=True, color="FFFFFF")

    ws.cell(2, 2, "Laporan Mutasi Bank Sahabat ETF").font = title_font
    subtitle = f"Dalam Jutaan Rp (sd {today.day} {_MONTH_NAMES_ID[today.month - 1]} {today.year})"
    ws.cell(3, 2, subtitle).font = subtitle_font

    ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
    ws.cell(5, 2, "Deskripsi").font = header_font

    col = _FIRST_MONTH_COL
    i = 0
    while i < n_months:
        yr = months[i][:4]
        j = i
        while j < n_months and months[j][:4] == yr:
            j += 1
        span = j - i
        if span > 1:
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + span - 1)
        ws.cell(5, col, yr).font = header_font
        for k in range(i, j):
            m = int(months[k][5:7])
            ws.cell(6, col + (k - i), _MONTH_ABBR_ID[m - 1]).font = header_font
        col += span
        i = j

    ws.merge_cells(start_row=5, start_column=total_col, end_row=6, end_column=total_col)
    ws.cell(5, total_col, "Total").font = header_font

    row = 8
    penerimaan_subtotal_row, row = _write_section(ws, row, "PENERIMAAN", classified["penerimaan"], months)
    row += 1

    ws.cell(row, 2, "PENGELUARAN").font = section_font
    row += 1
    bank_subtotal_row, row = _write_section(ws, row, "Bank", classified["bank"], months)
    row += 1

    sahabat_subtotal_row, row = _write_section(ws, row, "Sahabat ETF", classified["sahabat_etf"], months)
    row += 1

    saldo_row = row
    ws.cell(saldo_row, 2, "SALDO AKHIR").font = saldo_font
    ws.cell(saldo_row, 2).fill = saldo_fill
    for k in range(n_months):
        col_letter = get_column_letter(_FIRST_MONTH_COL + k)
        if k == 0:
            formula = f"={col_letter}{penerimaan_subtotal_row}-{col_letter}{bank_subtotal_row}-{col_letter}{sahabat_subtotal_row}"
        else:
            prev_letter = get_column_letter(_FIRST_MONTH_COL + k - 1)
            formula = (
                f"={prev_letter}{saldo_row}+{col_letter}{penerimaan_subtotal_row}"
                f"-{col_letter}{bank_subtotal_row}-{col_letter}{sahabat_subtotal_row}"
            )
        c = ws.cell(saldo_row, _FIRST_MONTH_COL + k, formula)
        c.number_format = _NUM_FMT
        c.font = saldo_font
        c.fill = saldo_fill

    last_month_letter = get_column_letter(_FIRST_MONTH_COL + n_months - 1)
    c = ws.cell(saldo_row, total_col, f"={last_month_letter}{saldo_row}")
    c.number_format = _NUM_FMT
    c.font = saldo_font
    c.fill = saldo_fill

    ws.column_dimensions["B"].width = 35
    for k in range(n_months + 1):
        ws.column_dimensions[get_column_letter(_FIRST_MONTH_COL + k)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
