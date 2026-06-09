# tests/test_pam_exports.py
import os, sys, io, pytest
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import config
config.DB_PATH = os.path.join(os.path.dirname(__file__), "test_finance_hub.db")

from database import init_db, get_conn
from modules.payment_memo.exports import export_pam_pdf

COMPANY_ID   = 2
COMPANY_CODE = "ETF"

@pytest.fixture(autouse=True)
def clean_db():
    if os.path.exists(config.DB_PATH):
        os.remove(config.DB_PATH)
    init_db()
    _seed()
    yield
    if os.path.exists(config.DB_PATH):
        os.remove(config.DB_PATH)


def _seed():
    conn = get_conn()
    conn.execute(
        """INSERT INTO pam_records
           (company_id, pam_no, pam_date, gl_account, cost_center, pt,
            requestors_name, keterangan, total_amount, due_date, status, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,'open',?)""",
        (COMPANY_ID, "PAM-001-ETF-05-2026", "2026-05-26", "70110230",
         "1008C1POFF", "PT. SMART Tbk", "Jany Turkanda",
         "Harry Santoso", 5000000, "2026-06-26", "2026-05-26T10:00:00")
    )
    conn.execute(
        """INSERT INTO siswa (company_id, code, nama, bank, norek, namarek,
           jenjang, program, status)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (COMPANY_ID, "S001", "Harry Santoso", "BCA", "1234567890",
         "Harry Santoso", "S1", "SMART", "Aktif")
    )
    conn.execute(
        """INSERT INTO payment_beasiswa
           (company_id, siswa_code, cat1, cat2, tanggal, amount,
            pillar, perusahaan, pam, status)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (COMPANY_ID, "S001", "General", "Sem 1", "2026-05-26",
         5000000, "ETF", "PT. SMART Tbk", "PAM-001-ETF-05-2026", "open")
    )
    conn.commit()
    conn.close()


def test_export_pam_pdf_returns_bytes():
    result = export_pam_pdf(1, COMPANY_ID, "Hong Tjhin", "Tenti Kidjo")
    assert isinstance(result, bytes)
    assert len(result) > 1000
    assert result[:4] == b'%PDF'


def test_export_pam_pdf_not_found_raises():
    with pytest.raises(ValueError, match="PAM record tidak ditemukan"):
        export_pam_pdf(999, COMPANY_ID, "A", "B")


from modules.payment_memo.exports import export_pam_excel
import zipfile


def test_export_pam_excel_returns_bytes():
    result = export_pam_excel(1, COMPANY_ID, "Hong Tjhin", "Tenti Kidjo")
    assert isinstance(result, bytes)
    assert len(result) > 500
    assert zipfile.is_zipfile(io.BytesIO(result))


def test_export_pam_excel_has_three_sheets():
    import openpyxl
    result = export_pam_excel(1, COMPANY_ID, "Hong Tjhin", "Tenti Kidjo")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert wb.sheetnames == ["PAM NEW", "Rangkuman PAM", "Detail PAM"]


def test_export_pam_excel_pam_no_in_sheet():
    import openpyxl
    result = export_pam_excel(1, COMPANY_ID, "Hong Tjhin", "Tenti Kidjo")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["PAM NEW"]
    values = [ws.cell(r, c).value for r in range(1, 30) for c in range(1, 10)]
    assert "PAM-001-ETF-05-2026" in values


def test_export_pam_excel_not_found_raises():
    with pytest.raises(ValueError, match="PAM record tidak ditemukan"):
        export_pam_excel(999, COMPANY_ID, "A", "B")


from modules.payment_memo.exports import export_pam_pdf_custom

_CUSTOM_DATA = {
    "pam_no":           "PAM-001-ETF-05-2026",
    "pam_date":         "2026-05-26",
    "requestors_name":  "Jany Turkanda",
    "department":       "HR",
    "cost_center":      "1008C1POFF",
    "gl_account":       "70110230",
    "so_sc":            "",
    "pt":               "PT. SMART Tbk",
    "bu_upstream":      False,
    "bu_downstream":    False,
    "bu_corporate":     True,
    "type_downpayment": False,
    "type_invoice":     True,
    "type_advance":     False,
    "vendor_name":      "Terlampir",
    "invoice_memo_no":  "-",
    "total_amount":     5000000,
    "due_date":         "2026-06-26",
    "bank_account_name":"Terlampir",
    "bank_name":        "Terlampir",
    "bank_account_no":  "Terlampir",
    "approved_by_1":    "Hong Tjhin",
    "approved_by_2":    "Tenti Kidjo",
}

_PAYMENTS = [
    {"siswa_code": "S001", "nama": "Harry Santoso", "bank": "BCA",
     "norek": "1234567890", "namarek": "Harry Santoso",
     "cat1": "General", "cat2": "Sem 1", "amount": 5000000},
]

def test_export_pam_pdf_custom_returns_pdf():
    result = export_pam_pdf_custom(_CUSTOM_DATA, _PAYMENTS)
    assert isinstance(result, bytes)
    assert result[:4] == b'%PDF'
    assert len(result) > 1000

def test_export_pam_pdf_custom_empty_payments():
    result = export_pam_pdf_custom(_CUSTOM_DATA, [])
    assert isinstance(result, bytes)
    assert result[:4] == b'%PDF'

def test_export_pam_pdf_custom_alternate_checkboxes_does_not_crash():
    data = {**_CUSTOM_DATA, "bu_upstream": True, "bu_corporate": False}
    result = export_pam_pdf_custom(data, [])
    assert isinstance(result, bytes)
    assert result[:4] == b'%PDF'

def test_export_pam_pdf_custom_real_vendor_name():
    data = {**_CUSTOM_DATA, "vendor_name": "PT. Maju Jaya", "bank_account_name": "Hendra Wijaya"}
    result = export_pam_pdf_custom(data, [])
    assert isinstance(result, bytes)
    assert result[:4] == b'%PDF'

from modules.payment_memo.exports import export_pam_excel_custom

def test_export_pam_excel_custom_returns_xlsx():
    import zipfile, io as _io
    result = export_pam_excel_custom(_CUSTOM_DATA, _PAYMENTS)
    assert isinstance(result, bytes)
    assert zipfile.is_zipfile(_io.BytesIO(result))

def test_export_pam_excel_custom_has_two_sheets():
    import openpyxl, io as _io
    result = export_pam_excel_custom(_CUSTOM_DATA, _PAYMENTS)
    wb = openpyxl.load_workbook(_io.BytesIO(result))
    assert wb.sheetnames == ["PAM NEW", "Rangkuman PAM"]

def test_export_pam_excel_custom_pam_no_in_sheet():
    import openpyxl, io as _io
    result = export_pam_excel_custom(_CUSTOM_DATA, _PAYMENTS)
    wb = openpyxl.load_workbook(_io.BytesIO(result))
    ws = wb["PAM NEW"]
    values = [ws.cell(r, c).value for r in range(1, 15) for c in range(1, 18)]
    assert "PAM-001-ETF-05-2026" in values

def test_export_pam_excel_custom_approved_by_in_sheet():
    import openpyxl, io as _io
    result = export_pam_excel_custom(_CUSTOM_DATA, _PAYMENTS)
    wb = openpyxl.load_workbook(_io.BytesIO(result))
    ws = wb["PAM NEW"]
    values = [ws.cell(r, c).value for r in range(36, 50) for c in range(1, 12)]
    assert "Hong Tjhin" in values


# ── Detail PAM (Sheet 3) tests ───────────────────────────────────────────────

from modules.payment_memo.service import get_pam_payments_detail

PAM_NO = "PAM-001-ETF-05-2026"


def _seed_detail():
    conn = get_conn()
    conn.execute(
        "INSERT OR IGNORE INTO siswa (company_id, code, nama, bank, norek, namarek, "
        "jenjang, angkatan, program, universitas, fakultas, status) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (COMPANY_ID, "S002", "Budi Santoso", "BNI", "9876543210",
         "Budi Santoso", "S3", 2024, "Kejaksaan", "Univ. Gadjah Mada",
         "Ilmu Hukum", "Aktif")
    )
    for cat1, cat2, amt in [
        ("By Pendidikan", "Semester 1", 15000000),
        ("By Tunjangan",  "Semester 1", 5000000),
        ("By Pendidikan", "Semester 2", 15000000),
    ]:
        conn.execute(
            "INSERT INTO payment_beasiswa "
            "(company_id, siswa_code, cat1, cat2, tanggal, amount, pam, status) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (COMPANY_ID, "S002", cat1, cat2, "2026-05-26", amt, PAM_NO, "open")
        )
    conn.execute(
        "INSERT INTO budget_beasiswa "
        "(company_id, siswa_code, cat1, amount) VALUES (?,?,?,?)",
        (COMPANY_ID, "S002", "By Pendidikan", 100000000)
    )
    conn.execute(
        "INSERT INTO budget_beasiswa "
        "(company_id, siswa_code, cat1, amount) VALUES (?,?,?,?)",
        (COMPANY_ID, "S002", "By Tunjangan", 20000000)
    )
    conn.commit()
    conn.close()


def test_get_pam_payments_detail_groups_by_siswa():
    _seed_detail()
    rows = get_pam_payments_detail(PAM_NO, COMPANY_ID)
    codes = [r["siswa_code"] for r in rows]
    assert "S002" in codes
    budi = next(r for r in rows if r["siswa_code"] == "S002")
    assert len(budi["rows"]) == 2  # Semester 1, Semester 2


def test_get_pam_payments_detail_pivot_cat1():
    _seed_detail()
    rows = get_pam_payments_detail(PAM_NO, COMPANY_ID)
    budi = next(r for r in rows if r["siswa_code"] == "S002")
    sem1 = next(r for r in budi["rows"] if r["keterangan"] == "Semester 1")
    assert sem1["pendidikan"] == 15000000
    assert sem1["tunjangan"]  == 5000000
    assert sem1["penelitian"] == 0


def test_get_pam_payments_detail_sisa_saldo():
    _seed_detail()
    rows = get_pam_payments_detail(PAM_NO, COMPANY_ID)
    budi = next(r for r in rows if r["siswa_code"] == "S002")
    # budget 100000000 - paid 30000000 (both semesters pendidikan)
    assert budi["sisa_pendidikan"] == 100000000 - 30000000
    # budget 20000000 - paid 5000000 (tunjangan sem1)
    assert budi["sisa_tunjangan"] == 20000000 - 5000000
    assert budi["sisa_penelitian"] == 0


def test_export_pam_excel_detail_sheet_headers():
    import openpyxl
    _seed_detail()
    result = export_pam_excel(1, COMPANY_ID, "", "")
    wb     = openpyxl.load_workbook(io.BytesIO(result))
    assert "Detail PAM" in wb.sheetnames
    ws     = wb["Detail PAM"]
    all_vals = {ws.cell(r, c).value
                for r in range(1, 8) for c in range(1, 22)}
    assert "NO" in all_vals
    assert "NAMA SISWA" in all_vals
    assert "KETERANGAN" in all_vals
    assert "SISA SALDO" in all_vals
    assert PAM_NO in all_vals
