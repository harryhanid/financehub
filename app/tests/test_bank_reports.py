import io
import zipfile
from datetime import datetime

import openpyxl
from database import get_conn

from modules.bank.reports import classify_rows, month_range, build_laporan_mutasi_excel


def _row(tanggal, jenis, jumlah, keterangan, source="manual", pam_record_id=None, rid=1):
    return {
        "id": rid, "tanggal": tanggal, "jenis": jenis, "jumlah": jumlah,
        "keterangan": keterangan, "source": source, "pam_record_id": pam_record_id,
    }


def test_classify_rows_excludes_setoran_awal_case_insensitive():
    rows = [
        _row("2025-11-17", "pemasukan", 1500000, "Setoran Awal - 17 Nov 2025"),
        _row("2025-11-20", "pemasukan", 2000, "setoran awal lain"),
        _row("2025-11-24", "pemasukan", 2000000000, "AGRI - PT Cipta Inti"),
    ]
    result = classify_rows(rows)
    assert result["penerimaan"] == {"AGRI - PT Cipta Inti": {"2025-11": 2000000000}}
    assert result["bank"] == {}
    assert result["sahabat_etf"] == {}


def test_classify_rows_pam_rows_go_to_sahabat_etf_grouped_by_name():
    rows = [
        _row("2026-07-09", "pengeluaran", 5320346, "Budi Widjaja", source="pam", pam_record_id=302),
        _row("2026-01-19", "pengeluaran", 7602533, "Budi Widjaja", source="pam", pam_record_id=837),
    ]
    result = classify_rows(rows)
    assert result["sahabat_etf"] == {
        "Budi Widjaja": {"2026-07": 5320346, "2026-01": 7602533}
    }


def test_classify_rows_bank_admin_bunga_nets_pengeluaran_minus_pemasukan():
    rows = [
        _row("2025-12-15", "pengeluaran", 2500, "Bank Admin & Bunga"),
        _row("2025-12-23", "pemasukan", 511742, "Bank Admin & Bunga"),
    ]
    result = classify_rows(rows)
    assert result["bank"] == {"Bunga dan Admin": {"2025-12": 2500 - 511742}}


def test_classify_rows_bank_admin_bunga_matches_case_insensitively():
    rows = [_row("2026-01-05", "pengeluaran", 30000, "bank admin & bunga")]
    result = classify_rows(rows)
    assert result["bank"] == {"Bunga dan Admin": {"2026-01": 30000}}


def test_classify_rows_manual_pemasukan_falls_back_to_penerimaan():
    rows = [_row("2026-01-22", "pemasukan", 4000000000, "APP - PT Tirta Pasific Unity")]
    result = classify_rows(rows)
    assert result["penerimaan"] == {"APP - PT Tirta Pasific Unity": {"2026-01": 4000000000}}


def test_classify_rows_manual_pengeluaran_falls_back_to_bank():
    rows = [_row("2026-01-30", "pengeluaran", 3500000000, "Penempatan RD")]
    result = classify_rows(rows)
    assert result["bank"] == {"Penempatan RD": {"2026-01": 3500000000}}


def test_classify_rows_blank_keterangan_grouped_as_tanpa_keterangan():
    rows = [_row("2026-01-05", "pengeluaran", 1000, "")]
    result = classify_rows(rows)
    assert result["bank"] == {"(Tanpa Keterangan)": {"2026-01": 1000}}


def test_month_range_from_earliest_eligible_to_today():
    rows = [
        _row("2025-11-17", "pemasukan", 1500000, "Setoran Awal - 17 Nov 2025"),
        _row("2025-12-15", "pengeluaran", 2500, "Bank Admin & Bunga"),
    ]
    months = month_range(rows, today=datetime(2026, 3, 1))
    assert months == ["2025-12", "2026-01", "2026-02", "2026-03"]


def test_month_range_empty_data_falls_back_to_current_month():
    months = month_range([], today=datetime(2026, 7, 15))
    assert months == ["2026-07"]


def test_month_range_all_setoran_awal_falls_back_to_current_month():
    rows = [_row("2025-11-17", "pemasukan", 1500000, "Setoran Awal")]
    months = month_range(rows, today=datetime(2026, 7, 15))
    assert months == ["2026-07"]


def _insert_bank_setf(company_id, tanggal, jenis, jumlah, keterangan="Test", source="manual", pam_record_id=None):
    conn = get_conn()
    conn.execute(
        "INSERT INTO bank_setf (company_id, tanggal, jenis, jumlah, keterangan, source, pam_record_id) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (company_id, tanggal, jenis, jumlah, keterangan, source, pam_record_id),
    )
    conn.commit()
    conn.close()


def test_build_laporan_mutasi_excel_returns_valid_xlsx_bytes():
    result = build_laporan_mutasi_excel(2, today=datetime(2026, 7, 15))
    assert isinstance(result, bytes)
    assert zipfile.is_zipfile(io.BytesIO(result))


def test_build_laporan_mutasi_excel_has_all_sheet_with_title():
    result = build_laporan_mutasi_excel(2, today=datetime(2026, 7, 15))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert wb.sheetnames == ["ALL"]
    ws = wb["ALL"]
    assert ws["B2"].value == "Laporan Mutasi Bank Sahabat ETF"


def test_build_laporan_mutasi_excel_empty_data_still_has_saldo_akhir_row():
    result = build_laporan_mutasi_excel(2, today=datetime(2026, 7, 15))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["ALL"]
    values = [ws.cell(r, c).value for r in range(1, 20) for c in range(1, 6)]
    assert "SALDO AKHIR" in values


def test_build_laporan_mutasi_excel_excludes_setoran_awal_from_sheet():
    _insert_bank_setf(2, "2025-11-17", "pemasukan", 1500000, keterangan="Setoran Awal - 17 Nov 2025")
    _insert_bank_setf(2, "2025-11-24", "pemasukan", 2000000000, keterangan="AGRI - PT Cipta Inti")
    result = build_laporan_mutasi_excel(2, today=datetime(2025, 11, 30))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["ALL"]
    all_values = [ws.cell(r, c).value for r in range(1, 40) for c in range(1, 15)]
    assert not any(v and "Setoran Awal" in str(v) for v in all_values)
    assert "AGRI - PT Cipta Inti" in all_values


def test_build_laporan_mutasi_excel_penerimaan_value_in_millions():
    _insert_bank_setf(2, "2025-11-24", "pemasukan", 2000000000, keterangan="AGRI - PT Cipta Inti")
    result = build_laporan_mutasi_excel(2, today=datetime(2025, 11, 30))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["ALL"]
    row_idx = None
    for r in range(1, 20):
        if ws.cell(r, 2).value == "AGRI - PT Cipta Inti":
            row_idx = r
            break
    assert row_idx is not None
    assert ws.cell(row_idx, 3).value == 2000000000 / 1_000_000


def test_build_laporan_mutasi_excel_writes_sum_formulas_for_subtotal():
    _insert_bank_setf(2, "2025-11-24", "pemasukan", 2000000000, keterangan="AGRI - PT Cipta Inti")
    result = build_laporan_mutasi_excel(2, today=datetime(2025, 11, 30))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["ALL"]
    formula_cells = [
        ws.cell(r, c).value for r in range(1, 30) for c in range(1, 10)
        if isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.startswith("=SUM(")
    ]
    assert len(formula_cells) > 0


def test_build_laporan_mutasi_excel_only_returns_data_for_requested_company():
    _insert_bank_setf(2, "2025-11-24", "pemasukan", 2000000000, keterangan="AGRI - PT Cipta Inti")
    _insert_bank_setf(1, "2025-11-24", "pemasukan", 999000000, keterangan="SMT Only Entry")
    result = build_laporan_mutasi_excel(2, today=datetime(2025, 11, 30))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["ALL"]
    all_values = [ws.cell(r, c).value for r in range(1, 40) for c in range(1, 15)]
    assert "SMT Only Entry" not in all_values


def test_build_laporan_mutasi_excel_zero_activity_month_still_shows_as_column():
    # Nov has a transaction, Des has none, Jan has a transaction — Des must still
    # appear as its own column with 0, not be skipped from the month range.
    _insert_bank_setf(2, "2025-11-24", "pemasukan", 2000000000, keterangan="AGRI - PT Cipta Inti")
    _insert_bank_setf(2, "2026-01-22", "pemasukan", 4000000000, keterangan="APP - PT Tirta Pasific Unity")
    result = build_laporan_mutasi_excel(2, today=datetime(2026, 1, 31))
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["ALL"]
    header_row_values = [ws.cell(6, c).value for c in range(3, 8)]
    assert header_row_values == ["Nov", "Des", "Jan", None, None]
