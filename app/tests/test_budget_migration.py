# tests/test_budget_migration.py
import os
import sys
import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from migrate_budget_from_excel import migrate_budget_from_excel
from modules.budget.service import list_budgets, list_realisasi, get_carryover_data


@pytest.fixture
def fixture_xlsx(tmp_path):
    path = tmp_path / "budget_export.xlsx"
    wb = openpyxl.Workbook()
    ws_budget = wb.active
    ws_budget.title = "Master_Budget"
    ws_budget.append(["ID", "MM", "YY", "Company", "Dept", "GL_Acc", "GL_Desc",
                       "Category", "Activity", "Desc", "Amount", "Deadline"])
    ws_budget.append(["PO-FIN-26-01-ABCDE", 1, 2026, "PO", "Finance", "70110230",
                       "Scholarship Expense", "OpEx", "Audit Fee", "Annual audit",
                       50000000, "2026-03-31"])

    ws_real = wb.create_sheet("Realisasi")
    ws_real.append(["Trx_ID", "Budget_ID", "MM", "YY", "Company", "Dept", "GL_Account",
                     "GL_Desc", "Budget_Category", "Activity", "Description",
                     "Amount_Realisasi", "Tanggal_Realisasi"])
    ws_real.append(["TRX-ABCDE123", "PO-FIN-26-01-ABCDE", 1, 2026, "PO", "Finance",
                     "70110230", "Scholarship Expense", "OpEx", "Audit Fee",
                     "Partial payment", 20000000, "2026-01-15"])

    ws_carry = wb.create_sheet("Carryover_Logs")
    ws_carry.append(["Budget_ID", "Requested_By", "Request_Date", "Status", "Approval_Date",
                      "Extension_Months", "Reason", "Approved_By", "Type", "Additional_Amount"])
    ws_carry.append(["PO-FIN-26-01-ABCDE", "harry", "2026-02-01", "Approved", "2026-02-05",
                      12, "Belum sempat", "releaser1", "Carryover", 0])

    wb.save(path)
    return str(path)


def test_migrate_imports_all_three_sheets(fixture_xlsx):
    result = migrate_budget_from_excel(fixture_xlsx)
    assert result["budgets"] == 1
    assert result["realisasi"] == 1
    assert result["carryover_logs"] == 1
    assert result["errors"] == []

    budgets = list_budgets({"company": "PO"})
    assert any(b["id"] == "PO-FIN-26-01-ABCDE" for b in budgets)

    realisasi = list_realisasi({"company": "PO"})
    assert any(r["trx_id"] == "TRX-ABCDE123" for r in realisasi)

    logs = get_carryover_data()
    assert any(l["budget_id"] == "PO-FIN-26-01-ABCDE" and l["status"] == "Approved" for l in logs)


def test_migrate_is_idempotent_for_budgets(fixture_xlsx):
    migrate_budget_from_excel(fixture_xlsx)
    result = migrate_budget_from_excel(fixture_xlsx)
    # second run should not error, but also should not duplicate the budget row
    budgets = [b for b in list_budgets({"company": "PO"}) if b["id"] == "PO-FIN-26-01-ABCDE"]
    assert len(budgets) == 1
    assert result["errors"] == [] or "sudah ada" in result["errors"][0].lower()
